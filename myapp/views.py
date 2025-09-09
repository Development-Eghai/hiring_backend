# Import necessary modules and models
import base64
from decimal import Decimal, InvalidOperation
from django.core.mail import EmailMessage
from io import BytesIO
from openpyxl.utils import get_column_letter
import json
import openpyxl
from openpyxl.styles import Font
from urllib.parse import quote
import os
import random
from django.utils.html import strip_tags
import PyPDF2 as pdf
import jwt
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
# from django.contrib import messages
from django.contrib.auth import get_user_model
# from django.contrib.auth.models import Us
from django.contrib.auth.hashers import check_password
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.crypto import get_random_string
from django.views.decorators.csrf import csrf_exempt
from langchain_ollama import OllamaLLM
from django.db.models import F
import requests
from rest_framework import generics
from rest_framework import status
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
# from .utils import extract_info_from_resume  # Import the parsing function
from rest_framework.permissions import IsAuthenticated

from myapp.zoom_utils import schedule_zoom_meet
from datetime import timedelta, timezone as dt_timezone
from django.core.files.storage import default_storage




from .models import ApprovalStatus, Approver, BankingDetails, BgCheckRequest, BgPackage, BgPackageDetail, BgVendor, CandidateApproval, CandidateEducation, CandidateEmployment, CandidateFeedback, CandidateFormInvite, CandidateInterviewStages, CandidatePersonal, CandidateProfile, CandidateReference, CandidateReview, CandidateSubmission, ConfigHiringData, \
    ConfigPositionRole, ConfigScoreCard, ConfigScreeningType, DocumentItem, FinancialDocuments, GeneratedOffer, InsuranceDetail, InterviewDesignParameters, InterviewDesignScreen, InterviewPlanner, \
    InterviewReview, InterviewSchedule, InterviewSlot, Interviewer, Nominee, OfferNegotiation, OfferSalaryComponent, OfferVariablePayComponent, PersonalDetails, ReferenceCheck, RequisitionDetails, StageAlertResponsibility, \
    UserDetails, UserroleDetails
from .models import Candidate
from .models import InterviewRounds, HiringPlan
from .models import JobRequisition
from .serializers import ApproverDetailSerializer, ApproverDetailSerializer1, ApproverSerializer, BgCheckRequestSerializer, BgPackageSerializer, CandidateApprovalStatusSerializer, CandidateDetailWithInterviewSerializer, CandidateFeedbackEnrichedSerializer, CandidateFormInviteSerializer, CandidateInterviewStagesSerializer, CandidateOfferReportSerializer, CandidatePreOnboardingSerializer, CandidateReviewSerializer, CandidateSerializer, \
    CandidateSubmissionSerializer, ConfigHiringDataSerializer, ConfigPositionRoleSerializer, ConfigScoreCardSerializer, \
    ConfigScreeningTypeSerializer, InterviewDesignParametersSerializer, InterviewDesignScreenSerializer, InterviewPlannerSerializer, InterviewReviewSerializer, \
    InterviewerSerializer, JobRequisitionSerializer, JobTemplateSerializer, OfferNegotiationSerializer, \
    StageAlertResponsibilitySerializer
from .serializers import HiringInterviewRoundsSerializer, HiringSkillsSerializer, HiringPlanSerializer

SECRET_KEY = settings.SECRET_KEY
from rest_framework.decorators import api_view,permission_classes
from rest_framework_simplejwt.tokens import RefreshToken
from .jwt_token import api_json_response_format
from rest_framework.views import APIView
from rest_framework.response import Response
from django.utils import timezone


from datetime import datetime, date

import pytz
# from sentence_transformers import SentenceTransformer, util
# import torch
import re
from .serializers import JobRequisitionDetailSerializer

# ollama.base_url = "http://ollama:11434"


# required for JWT
User = get_user_model()

class ApproverCreateListView(generics.ListCreateAPIView):
    queryset = Approver.objects.all()
    serializer_class = ApproverSerializer

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset().select_related("requisition", "hiring_plan").order_by("requisition__RequisitionID")

            if not queryset.exists():
                return Response(api_json_response_format(False, "No approvers found.", 404, {}), status=200)

            requisition_groups = {}
            for approver in queryset:
                req = approver.requisition
                req_id = req.RequisitionID if req else "Unknown"

                if req_id not in requisition_groups:
                    requisition_groups[req_id] = {
                        "req_id": req_id,
                        "planning_id": approver.hiring_plan.hiring_plan_id if approver.hiring_plan else "",
                        "client_name": req.company_client_name if req else "",
                        "client_id": req.client_id if req else "",
                        "approvers": []
                    }

                requisition_groups[req_id]["approvers"].append({
                    "approver_id": approver.id,
                    "role": approver.role,
                    "job_title": approver.job_title,
                    "first_name": approver.first_name,
                    "last_name": approver.last_name,
                    "email": approver.email,
                    "contact_number": approver.contact_number,
                    "set_as_approver": "Yes" if approver.set_as_approver else "No"
                })

            # Add count per requisition
            for group in requisition_groups.values():
                group["no_of_approvers"] = len(group["approvers"])

            response_data = list(requisition_groups.values())

            return Response(api_json_response_format(True, "Approvers grouped by requisition.", 200, response_data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error fetching grouped approvers. {str(e)}", 500, {}), status=200)


    def create(self, request, *args, **kwargs):
        try:
            payload = request.data
            planning_id_str = payload.get("planning_id")
            hiring_plan_obj = None

            if planning_id_str:
                try:
                    hiring_plan_obj = HiringPlan.objects.get(hiring_plan_id=planning_id_str)
                except HiringPlan.DoesNotExist:
                    return Response(api_json_response_format(
                        False,
                        f"HiringPlan with ID '{planning_id_str}' not found.",
                        404,
                        {}
                    ), status=200)
            approver_list = payload.get("approvers", [])

            if not approver_list or not isinstance(approver_list, list):
                return Response(api_json_response_format(
                    False,
                    "Missing or invalid 'approvers' list in payload.",
                    400,
                    {}
                ), status=200)
            

            # Shared metadata
            shared_fields = {
                "requisition_id": payload.get("req_id"),
                "hiring_plan": hiring_plan_obj.id if hiring_plan_obj else None,  # ✅ Correct key
                "client_name": payload.get("client_name"),
                "client_id": payload.get("client_id"),
                "no_of_approvers": payload.get("no_of_approvers")
            }


            # Enrich each approver with shared metadata
            
            enriched_approvers = []
            for approver in approver_list:
                enriched = {
                    **approver,
                    "requisition": shared_fields["requisition_id"],  # 👈 use correct key name
                    "hiring_plan": shared_fields["hiring_plan"],
                    "client_name": shared_fields["client_name"],
                    "client_id": shared_fields["client_id"],
                    "no_of_approvers": shared_fields["no_of_approvers"]
                }
                enriched_approvers.append(enriched)

            serializer = self.get_serializer(data=enriched_approvers, many=True)
            serializer.is_valid(raise_exception=True)
            Approver.objects.bulk_create([Approver(**data) for data in serializer.validated_data])

            return Response(api_json_response_format(
                True,
                "Approvers created successfully!",
                200,
                serializer.data
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error creating approvers. {str(e)}",
                500,
                {}
            ), status=200)



    def put(self, request, *args, **kwargs):
        try:
            payload = request.data
            planning_id_str = payload.get("planning_id")
            hiring_plan_obj = None

            if planning_id_str:
                try:
                    hiring_plan_obj = HiringPlan.objects.get(hiring_plan_id=planning_id_str)
                except HiringPlan.DoesNotExist:
                    return Response(api_json_response_format(
                        False,
                        f"HiringPlan with ID '{planning_id_str}' not found.",
                        404,
                        {}
                    ), status=200)
            approver_list = payload.get("approvers", [])

            if not approver_list or not isinstance(approver_list, list):
                return Response(api_json_response_format(
                    False,
                    "Missing or invalid 'approvers' list in payload.",
                    400,
                    {}
                ), status=200)

            # Shared metadata remap
            shared_fields = {
                "requisition_id": payload.get("req_id"),
                "hiring_plan":hiring_plan_obj.id if hiring_plan_obj else None,
                "client_name": payload.get("client_name"),
                "client_id": payload.get("client_id"),
                "no_of_approvers": payload.get("no_of_approvers")
            }

            updated_data = []
            errors = []

            for approver in approver_list:
                approver_id = approver.get("approver_id")

                if not approver_id:
                    errors.append(f"Missing 'approver_id' in one of the approver entries.")
                    continue

                try:
                    instance = Approver.objects.get(id=approver_id)

                    # Merge individual fields with shared ones
                    update_fields = {
                        **approver,
                        "requisition_id": shared_fields["requisition_id"],
                        "hiring_plan": shared_fields["hiring_plan"],
                    }

                    serializer = self.get_serializer(instance, data=update_fields, partial=True)
                    serializer.is_valid(raise_exception=True)
                    serializer.save()
                    updated_data.append(serializer.data)

                except Approver.DoesNotExist:
                    errors.append(f"Approver ID {approver_id} not found.")
                except Exception as e:
                    errors.append(f"Error updating approver ID {approver_id}: {str(e)}")

            success = bool(updated_data)
            message = "Approvers updated successfully!" if success else "No approvers updated."

            return Response(api_json_response_format(
                success,
                message,
                200 if success else 400,
                {
                    "updated_approvers": updated_data,
                    "errors": errors
                }
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error processing update request. {str(e)}",
                500,
                {}
            ), status=200)



    def delete(self, request, *args, **kwargs):
        try:
            approver_id = request.data.get('id')
            approver = Approver.objects.get(id=approver_id)
            approver.delete()
            return Response(api_json_response_format(
                True,
                "Approver deleted successfully!",
                200,
                {}
            ), status=200)
        except Approver.DoesNotExist:
            return Response(api_json_response_format(
                False,
                "Approver not found.",
                404,
                {}
            ), status=200)
        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error deleting approver. {str(e)}",
                500,
                {}
            ), status=200)
        

@csrf_exempt
@api_view(['POST'])
def get_approvers_by_requisition(request):
    try:
        req_id = request.data.get("req_id")

        if not req_id:
            return Response(api_json_response_format(
                False,
                "Missing 'req_id' in payload.",
                400,
                {}
            ), status=200)

        # Fetch approvers for given requisition
        approvers_qs = Approver.objects.filter(requisition__RequisitionID=req_id)

        if not approvers_qs.exists():
            return Response(api_json_response_format(
                False,
                f"No approvers found for requisition {req_id}.",
                404,
                {}
            ), status=200)

        # Pull metadata from first approver’s related requisition/hiring plan
        first_approver = approvers_qs.first()
        requisition = first_approver.requisition
        hiring_plan = getattr(first_approver, "hiring_plan", None)

        response_data = {
            "req_id": requisition.RequisitionID if requisition else "Unknown",
            "planning_id": hiring_plan.hiring_plan_id if hiring_plan else "",
            "client_name": requisition.company_client_name if requisition else "",
            "client_id": requisition.client_id if requisition else "",
            "no_of_approvers": approvers_qs.count(),
            "approvers": [
                {
                    "approver_id": approver.id,
                    "role": approver.role,
                    "job_title": approver.job_title,
                    "first_name": approver.first_name,
                    "last_name": approver.last_name,
                    "email": approver.email,
                    "contact_number": approver.contact_number,
                    "set_as_approver": approver.set_as_approver
                }
                for approver in approvers_qs
            ]
        }

        return Response(api_json_response_format(
            True,
            "Approvers fetched successfully!",
            200,
            response_data
        ), status=200)

    except Exception as e:
        return Response(api_json_response_format(
            False,
            f"Error fetching approvers. {str(e)}",
            500,
            {}
        ), status=200)


class BgPackageSetupView(APIView):
    def get(self, request):
        packages = BgPackage.objects.select_related("vendor").all()

        data = []
        for pkg in packages:
            vendor = pkg.vendor

            # Fetch vendor-level add-ons
            vendor_addons = [
                {
                    "title": detail.title,
                    "description": detail.description,
                    "rate": str(detail.rate)
                }
                for detail in vendor.details.all()
            ]

            package_data = {
                "vendor_id": vendor.id,
                "vendor_name": vendor.name,
                "vendor_email": vendor.contact_email,
                "vendor_address": vendor.address,
                "mobile_no": vendor.mobile_no,
                "package_name": pkg.name,
                "package_rate": str(pkg.rate),
                "package_verification": pkg.verification_items.split(",") if pkg.verification_items else [],
                "details": vendor_addons
            }


            data.append(package_data)

        return Response(api_json_response_format(
            True,
            "All Vendor Package Details fetched successfully.",
            200,
            data
        ), status=200)


    def post(self, request):
        try:
            vendor_name = request.data.get("vendor_name")
            vendor_email = request.data.get("vendor_email")
            vendor_address = request.data.get("vendor_address")
            packages = request.data.get("packages", [])
            details = request.data.get("details", [])

            if not all([vendor_name, vendor_email, vendor_address]) or not packages:
                return Response(api_json_response_format(
                    False,
                    "Vendor or package data missing.",
                    400,
                    {}
                ), status=200)

            vendor, _ = BgVendor.objects.get_or_create(
                name=vendor_name,
                defaults={
                    "contact_email": vendor_email,
                    "address": vendor_address,
                    "mobile_no": request.data.get("mobile_no", "")

                }
            )

            saved = []

            for pkg in packages:
                pkg_name = pkg.get("package_name")
                pkg_rate = pkg.get("package_rate")
                pkg_desc = pkg.get("package_description")

                if not pkg_name or pkg_rate is None:
                    continue

                verification_list = pkg.get("package_verification", [])
                verification_str = ", ".join(verification_list) if isinstance(verification_list, list) else ""

                bg_package = BgPackage.objects.create(
                    vendor=vendor,
                    name=pkg_name,
                    rate=pkg_rate,
                    description=pkg_desc,
                    verification_items=verification_str
                )


                saved.append(BgPackageSerializer(bg_package).data)

            # Save vendor-level add-ons
            for detail in details:
                title = detail.get("add_on_check_title")
                description = detail.get("add_on_check_desc", "")
                rate_raw = detail.get("add_on_check_rate")

                try:
                    rate = Decimal(rate_raw)
                except (TypeError, InvalidOperation):
                    continue  # skip invalid rate

                if title:
                    BgPackageDetail.objects.create(
                        vendor=vendor,
                        title=title,
                        description=description,
                        rate=rate
                    )


            return Response(api_json_response_format(
                True,
                f"{len(saved)} packages and {len(details)} add-ons configured successfully.",
                201,
                saved
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error processing request: {str(e)}",
                500,
                {}
            ), status=200)

    
    def put(self, request):
        try:
            vendor_id = request.data.get("vendor_id")
            if not vendor_id:
                return Response(api_json_response_format(
                    False,
                    "Vendor ID is required.",
                    400,
                    {}
                ), status=200)

            vendor_name = request.data.get("vendor_name")
            vendor_email = request.data.get("vendor_email")
            vendor_address = request.data.get("vendor_address")
            packages = request.data.get("packages", [])
            details = request.data.get("details", [])

            if not all([vendor_name, vendor_email, vendor_address]) or not packages:
                return Response(api_json_response_format(
                    False,
                    "Missing vendor info or packages.",
                    400,
                    {}
                ), status=200)

            try:
                vendor = BgVendor.objects.get(id=vendor_id)
            except BgVendor.DoesNotExist:
                return Response(api_json_response_format(
                    False,
                    "Vendor not found.",
                    404,
                    {}
                ), status=200)

            # Update vendor info
            vendor.name = vendor_name
            vendor.contact_email = vendor_email
            vendor.address = vendor_address
            vendor.mobile_no = request.data.get("mobile_no", vendor.mobile_no)
            vendor.save()

            updated_packages = []

            for pkg_data in packages:
                pkg_name = pkg_data.get("package_name")
                pkg_rate = pkg_data.get("package_rate")
                pkg_desc = pkg_data.get("package_description")

                bg_package = BgPackage.objects.filter(vendor=vendor, name=pkg_name).first()

                verification_list = pkg_data.get("package_verification", [])
                verification_str = ", ".join(verification_list) if isinstance(verification_list, list) else ""

                if bg_package:
                    bg_package.rate = pkg_rate or bg_package.rate
                    bg_package.description = pkg_desc or bg_package.description
                    bg_package.verification_items = verification_str or bg_package.verification_items
                    bg_package.save()
                else:
                    bg_package = BgPackage.objects.create(
                        vendor=vendor,
                        name=pkg_name,
                        rate=pkg_rate,
                        description=pkg_desc,
                        verification_items=verification_str
                    )


                updated_packages.append(BgPackageSerializer(bg_package).data)

            # Replace vendor-level add-ons
            BgPackageDetail.objects.filter(vendor=vendor).delete()

            for detail in details:
                title = detail.get("add_on_check_title")
                description = detail.get("add_on_check_desc", "")
                rate = detail.get("add_on_check_rate")

                if not title or rate is None:
                    continue

                BgPackageDetail.objects.create(
                    vendor=vendor,
                    title=title,
                    description=description,
                    rate=rate
                )

            return Response(api_json_response_format(
                True,
                f"{len(updated_packages)} packages and {len(details)} add-ons updated successfully.",
                200,
                updated_packages
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error while updating packages: {str(e)}",
                500,
                {}
            ), status=200)



        
    def delete(self, request):
        try:
            vendor_id = request.data.get("vendor_id")

            if not vendor_id:
                return Response(api_json_response_format(
                    False,
                    "Vendor ID is required to delete packages.",
                    400,
                    {}
                ), status=200)

            try:
                vendor = BgVendor.objects.get(id=vendor_id)
            except BgVendor.DoesNotExist:
                return Response(api_json_response_format(
                    False,
                    "Vendor not found.",
                    404,
                    {}
                ), status=200)

            # Delete all details linked to this vendor
            BgPackageDetail.objects.filter(vendor=vendor).delete()

            # Delete all packages linked to this vendor
            deleted_count = BgPackage.objects.filter(vendor=vendor).count()
            BgPackage.objects.filter(vendor=vendor).delete()

            # Delete the vendor
            vendor.delete()

            return Response(api_json_response_format(
                True,
                f"Deleted {deleted_count} package(s), their vendor-linked details, and vendor successfully.",
                200,
                {}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error while deleting vendor packages: {str(e)}",
                500,
                {}
            ), status=200)



class VendorPackageView(APIView):
    def post(self, request):
        vendor_id = request.data.get("vendor_id")

        if not vendor_id:
            return Response(api_json_response_format(
                False, "Vendor ID is required.", 400, {}
            ), status=200)

        try:
            vendor = BgVendor.objects.get(id=vendor_id)
        except BgVendor.DoesNotExist:
            return Response(api_json_response_format(
                False, "Vendor not found.", 404, {}
            ), status=200)

        # Fetch packages linked to vendor
        packages = BgPackage.objects.filter(vendor=vendor)

        package_data = [
            {
                "package_name": pkg.name,
                "package_description": pkg.description,
                "package_rate": float(pkg.rate),
                "package_verification": pkg.verification_items.split(",") if pkg.verification_items else []
            }
            for pkg in packages
        ]

        # Fetch add-on details linked to vendor
        details = BgPackageDetail.objects.filter(vendor=vendor)

        detail_data = [
            {
                "add_on_check_title": detail.title,
                "add_on_check_desc": detail.description,
                "add_on_check_rate": float(detail.rate)
            }
            for detail in details
        ]

        response_data = {
            "vendor_id": vendor.id,
            "vendor_name": vendor.name,
            "vendor_address": vendor.address,
            "vendor_email": vendor.contact_email,
            "mobile_no": vendor.mobile_no,
            "packages": package_data,
            "details": detail_data
        }

        return Response(api_json_response_format(
            True, "Vendor packages and details fetched successfully.", 200, response_data
        ), status=200)






class InitiateBgCheckView(APIView):
    def post(self, request):
        try:
            # Extract incoming data
            requisition_id = request.data.get("requisition_id")
            candidate_id = request.data.get("candidate_id")
            vendor_id = request.data.get("vendor_id")
            selected_package_id = request.data.get("selected_package_id")
            custom_checks = request.data.get("custom_checks", [])  # expects list
            current_stage = request.data.get("current_stage")
            bg_status = request.data.get("bg_status", "Initiated")  # optional override

            # Validate required fields
            if not all([requisition_id, candidate_id, vendor_id]) or (
                not selected_package_id and not custom_checks
            ):
                return Response(api_json_response_format(
                    False,
                    "Missing required fields or no checks selected.",
                    400,
                    {}
                ), status=200)

            # Resolve foreign keys
            try:
                requisition = JobRequisition.objects.get(RequisitionID=requisition_id)
                candidate = Candidate.objects.get(id=candidate_id)
                vendor = BgVendor.objects.get(id=vendor_id)
                package = BgPackage.objects.get(id=selected_package_id) if selected_package_id else None
            except Exception as e:
                return Response(api_json_response_format(
                    False,
                    f"Invalid reference: {str(e)}",
                    404,
                    {}
                ), status=200)

            # Create BG check record
            bg_record = BgCheckRequest.objects.create(
                requisition=requisition,
                candidate=candidate,
                vendor=vendor,
                selected_packages=package,
                custom_checks=custom_checks,
                status=bg_status
            )

            response_data = {
                "bg_request_id": bg_record.id,
                "status": bg_record.status,
                "candidate_id": candidate.id,
                "vendor": vendor.name,
                "package": package.name if package else None,
                "custom_checks": custom_checks
            }

            return Response(api_json_response_format(
                True,
                "Background check initiated successfully.",
                201,
                response_data
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error initiating BG check: {str(e)}",
                500,
                {}
            ), status=200)



class GetCandidateReviewView(APIView):
    def post(self, request):
        candidate_id = request.data.get("candidate_id")
        if not candidate_id:
            return Response(api_json_response_format(False, "Missing candidate_id", 400, {}), status=200)

        reviews = InterviewReview.objects.filter(candidate_id=candidate_id)
        if not reviews.exists():
            return Response(api_json_response_format(False, "No reviews found", 404, {}), status=200)

        serializer = InterviewReviewSerializer(reviews, many=True)
        return Response(api_json_response_format(True, "Reviews retrieved", 200, serializer.data), status=200)

# class InitiateBgCheckView(APIView):
#     def post(self, request):
#         try:
#             requisition_id = request.data.get("requisition_id")
#             candidate_id = request.data.get("candidate_id")
#             vendor_id = request.data.get("vendor_id")
#             package_id = request.data.get("package_id")
#             custom_checks = request.data.get("custom_checks", [])  # optional list

#             if not (requisition_id and candidate_id and vendor_id):
#                 return Response(api_json_response_format(
#                     False,
#                     "Missing required fields: requisition_id, candidate_id, vendor_id.",
#                     400,
#                     {}
#                 ), status=200)

#             requisition = JobRequisition.objects.get(RequisitionID=requisition_id)
#             candidate = Candidate.objects.get(pk=candidate_id)
#             vendor = BgVendor.objects.get(pk=vendor_id)

#             selected_package = BgPackage.objects.get(pk=package_id) if package_id else None

#             bg_request = BgCheckRequest.objects.create(
#                 requisition=requisition,
#                 candidate=candidate,
#                 vendor=vendor,
#                 selected_package=selected_package,
#                 custom_checks=custom_checks,
#                 status="Initiated"
#             )

#             serializer = BgCheckRequestSerializer(bg_request)
#             return Response(api_json_response_format(
#                 True,
#                 "BG Check initiated.",
#                 201,
#                 serializer.data
#             ), status=200)

#         except JobRequisition.DoesNotExist:
#             return Response(api_json_response_format(False, "Requisition not found", 404, {}), status=200)
#         except Candidate.DoesNotExist:
#             return Response(api_json_response_format(False, "Candidate not found", 404, {}), status=200)
#         except BgVendor.DoesNotExist:
#             return Response(api_json_response_format(False, "Vendor not found", 404, {}), status=200)
#         except BgPackage.DoesNotExist:
#             return Response(api_json_response_format(False, "Package not found", 404, {}), status=200)
#         except Exception as e:
#             return Response(api_json_response_format(False, f"Error initiating BG check: {str(e)}", 500, {}), status=200)

class BgCheckRequestView(APIView):
    def post(self, request):
        try:
            data = request.data.copy()

            requisition = JobRequisition.objects.get(RequisitionID=data.get("requisition"))
            candidate = Candidate.objects.get(pk=data.get("candidate"))
            vendor = BgVendor.objects.get(pk=data.get("vendor"))

            # Accept multiple package IDs
            package_ids = data.get("selected_package", [])
            if not isinstance(package_ids, list):
                package_ids = [package_ids]

            packages = BgPackage.objects.filter(pk__in=package_ids)

            bg_request = BgCheckRequest.objects.create(
                requisition=requisition,
                candidate=candidate,
                vendor=vendor,
                custom_checks=data.get("custom_checks", []),
                status=data.get("status", "Initiated")
            )

            bg_request.selected_packages.set(packages)  # 👈 assign multiple packages

            serializer = BgCheckRequestSerializer(bg_request)
            return Response(api_json_response_format(True, "BG Check created.", 201, serializer.data), status=200)

        except JobRequisition.DoesNotExist:
            return Response(api_json_response_format(False, "Requisition not found", 404, {}), status=200)
        except Candidate.DoesNotExist:
            return Response(api_json_response_format(False, "Candidate not found", 404, {}), status=200)
        except BgVendor.DoesNotExist:
            return Response(api_json_response_format(False, "Vendor not found", 404, {}), status=200)
        except Exception as e:
            return Response(api_json_response_format(False, str(e), 500, {}), status=200)



    def get(self, request):
        bg_requests = BgCheckRequest.objects.select_related(
            "requisition", "candidate", "vendor"
        ).prefetch_related("selected_packages").all()


        data = [
            {
                "req_id": req.requisition.RequisitionID,
                "candidate_id": req.candidate.CandidateID,
                "first_name": req.candidate.candidate_first_name,
                "last_name": req.candidate.candidate_last_name,
                "email_id": req.candidate.Email,
                "location": getattr(req.requisition.position_information, "location", "Not Provided"),
                "vendor_name": req.vendor.name,
                "vendor_package": ", ".join([pkg.name for pkg in req.selected_packages.all()]) if req.selected_packages.exists() else "Custom",
                "add_on_check": req.custom_checks,
                "status": req.status,
                "decision": "Approved" if req.status == "Completed" else "Pending"  # Logic placeholder
            }
            for req in bg_requests
        ]

        return Response(api_json_response_format(
            True,
            "Dashboard fields loaded.",
            200,
            data
        ), status=200)

class BgCheckContextualDetailsView(APIView):
    def post(self, request):
        requisition_id = request.data.get("requisition_id")
        candidate_id = request.data.get("candidate_id")

        if not requisition_id:
            return Response(api_json_response_format(
                False, "Missing requisition_id.", 400, {}
            ), status=200)

        try:
            # Get candidate using candidate_id if present
            candidate = Candidate.objects.select_related('Req_id_fk').filter(
                CandidateID=candidate_id,
                Req_id_fk__RequisitionID=requisition_id
            ).first()

            stage = CandidateInterviewStages.objects.filter(
                candidate_id=candidate.CandidateID
            ).order_by('-interview_date').first() if candidate else None

            # Get the background check request
            bg_request = BgCheckRequest.objects.select_related(
                    "vendor", "requisition__position_information"
                ).filter(
                    requisition__RequisitionID=requisition_id
                ).first()


            # Get all vendors and packages for dropdown population
            vendors = BgVendor.objects.values_list("id", "name")
            # packages = BgPackage.objects.values_list("id", "name")
            location = "Not Provided"
            if requisition_id:
                try:
                    requisition_detail = RequisitionDetails.objects.select_related("requisition").get(
                        requisition__RequisitionID=requisition_id
                    )
                    location = requisition_detail.location or "Not Provided"
                except RequisitionDetails.DoesNotExist:
                    location = "Not Provided"
            

            data = {
                "vendor_options": [{"id": v[0], "name": v[1], "value": v[1]} for v in vendors],
                # "package_options": [{"id": p[0], "name": p[1], "value": p[1]} for p in packages],
                "location": location,
                "current_stage":  "Offer Generated",
                "candidate_email": candidate.Email if candidate else "Not Provided",
                "candidate_first_name": candidate.candidate_first_name if candidate else "Not Provided",
                "candidate_last_name": candidate.candidate_last_name if candidate else "Not Provided"
            }

            return Response(api_json_response_format(
                True, "Form data loaded successfully.", 200, data
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error loading data: {str(e)}", 500, {}
            ), status=200)
    

class BgPackageByVendorView(APIView):
    def post(self, request):
        vendor_id = request.data.get("vendor_id")

        if not vendor_id:
            return Response(api_json_response_format(
                False, "Missing vendor_id.", 400, {}
            ), status=200)

        try:
            packages = BgPackage.objects.filter(vendor_id=vendor_id).values("id", "name", "description", "rate")

            package_options = [
                {
                    "id": pkg["id"],
                    "name": pkg["name"],
                    "value": pkg["name"],
                    "description": pkg.get("description", ""),
                    "rate": pkg.get("rate", "N/A")
                }
                for pkg in packages
            ]

            return Response(api_json_response_format(
                True, "Packages loaded successfully.", 200, {
                    "package_options": package_options
                }
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error fetching packages: {str(e)}", 500, {}
            ), status=200)


class BgAddonChecksView(APIView):
    def post(self, request):
        vendor_id = request.data.get("vendor_id")
        if not vendor_id:
            return Response(api_json_response_format(
                False,
                "Missing vendor_id.",
                400,
                {}
            ), status=200)

        try:
            checks = BgPackageDetail.objects.filter(vendor_id=vendor_id)
            addon_checks = [
                { "name": check.title, "value": check.title }
                for check in checks
            ]
            return Response(api_json_response_format(
                True,
                "Checks loaded.",
                200,
                { "addon_checks": addon_checks }
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                str(e),
                500,
                {}
            ), status=200)



def resolve_offer_approval_status(approvals, expected_roles):
    decision_map = {
        approval.approver.role: approval.status
        for approval in approvals
        if approval.approver and hasattr(approval.approver, "role")
    }

    decisions = [decision_map.get(role, "Awaiting") for role in expected_roles]

    # Normalize statuses
    normalized = [d if d != "Pending" else "Awaiting" for d in decisions]

    if any(d == "Awaiting" for d in normalized):
        return "Awaiting Approval"
    if all(d == "Approved" for d in normalized):
        return "Approved"
    return "Partially Approved"



def compute_overall_status(approvals, expected_roles):
    decision_map = {a.role: a.decision for a in approvals}
    decisions = [decision_map.get(role, "Awaiting") for role in expected_roles]

    if any(d == "Awaiting" for d in decisions):
        return "Awaiting Approval"
    if all(d == "Approved" for d in decisions):
        return "Approved"
    return "Partially Approved"

def compute_overall_status1(approvals, expected_roles):
    decisions = {approval.approver.role: approval.decision for approval in approvals if approval.decision}

    if not decisions:
        return "Awaiting Approval"

    if all(decision == "Approved" for decision in decisions.values()):
        if set(decisions.keys()) == set(expected_roles):
            return "Approved"
        else:
            return "Partially Approved"

    if any(decision == "Reject" for decision in decisions.values()):
        return "Rejected"

    return "Awaiting Approval"



class CandidateApprovalStatusView(APIView):
    def get(self, request):
        try:
            candidates = Candidate.objects.select_related('Req_id_fk').all()
            grouped_results = {}

            for candidate in candidates:
                if candidate.Result in [None, "", "Pending"]:
                    continue

                requisition = candidate.Req_id_fk
                if not requisition:
                    continue

                details = getattr(requisition, 'position_information', None)
                approvals = CandidateApproval.objects.filter(candidate=candidate)
                approvers = Approver.objects.filter(requisition=requisition, set_as_approver="Yes")

                approver_details = []
                for approver in approvers:
                    approval = approvals.filter(approver=approver).first()

                    approver_details.append({
                        "role": approver.role,
                        "name": f"{approver.first_name} {approver.last_name}",
                        "email": approver.email,
                        "contact_number": approver.contact_number,
                        "job_title": approver.job_title,
                        "status": approver.set_as_approver,
                        "decision": approval.decision if approval else "Awaiting",
                        "comment": approval.comment if approval and approval.comment else ""
                    })

                expected_roles = [a.role for a in approvers]
                overall_status = compute_overall_status1(approvals, expected_roles)

                grouped_results[candidate.CandidateID] = {
                    "req_id": requisition.RequisitionID,
                    "client_name": details.company_client_name if details else "",
                    "client_id": requisition.client_id if requisition and requisition.client_id else "",
                    "candidate_id": candidate.CandidateID,
                    "candidate_first_name": candidate.candidate_first_name,
                    "candidate_last_name": candidate.candidate_last_name,
                    "screening_status": candidate.Result,
                    "approvers": approver_details,
                    "overall_status": overall_status,
                    "no_of_approvers": len(approver_details)
                }

            final_result = list(grouped_results.values())
            serializer = ApproverDetailSerializer1(final_result, many=True)

            return Response(api_json_response_format(
                True,
                "Grouped approval status retrieved!",
                200,
                serializer.data
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error retrieving approval status. {str(e)}",
                500,
                []
            ), status=200)


    def post(self, request):
        try:
            candidates = Candidate.objects.select_related('Req_id_fk').all()
            results = []

            for candidate in candidates:
                if candidate.Result in [None, "", "Pending"]:
                    continue  # ⛔ Skip candidates with null or pending results

                requisition = candidate.Req_id_fk
                if not requisition:
                    continue

                details = getattr(requisition, 'position_information', None)
                approvals = CandidateApproval.objects.filter(candidate=candidate)
                approvers = Approver.objects.filter(requisition=requisition, set_as_approver="Yes")

                approver_details = []
                for approver in approvers:
                    approval = approvals.filter(approver=approver).first()

                    approver_detail = {
                        "role": approver.role,
                        "name": f"{approver.first_name} {approver.last_name}",
                        "email": approver.email,
                        "contact_number": approver.contact_number,
                        "job_title": approver.job_title,
                        "status": approver.set_as_approver,
                        "decision": approval.decision if approval else "Awaiting",
                        "comment": approval.comment if approval and approval.comment else "",
                        # "reviewed_at": approval.reviewed_at if approval else "Not Reviewed",
                        "req_id": requisition.RequisitionID,
                        "client_name": details.company_client_name if details else "",
                        "client_id": requisition.client_id if requisition and requisition.client_id else "",
                        "candidate_id": candidate.CandidateID,
                        "candidate_first_name": candidate.candidate_first_name,
                        "candidate_last_name": candidate.candidate_last_name,
                        "screening_status": candidate.Result, 
                    }

                    results.append(approver_detail)  # Append each approver individually



                expected_roles = [a.role for a in approvers]
                overall_status = compute_overall_status(approvals, expected_roles)

                # results.append({
                #     "req_id": requisition.RequisitionID,
                #     "client_name": details.company_client_name if details else "",
                #     "client_id": details.client_id if details else "",
                #     "candidate_id": candidate.CandidateID,
                #     "candidate_first_name": candidate.candidate_first_name,
                #     "candidate_last_name": candidate.candidate_last_name,
                #     "overall_status": overall_status,
                #     "no_of_approvers": len(approvers),
                #     "approvers": approver_details
                # })
            # results.append({
            #     "req_id": requisition.RequisitionID,
            #     "client_name": details.company_client_name if details else "",
            #     "client_id": details.client_id if details else "",
            #     "approvers": approver_details
            # })

            serializer = ApproverDetailSerializer(results, many=True)
            return Response(api_json_response_format(
                True,
                "approval status retrieved!",
                200,
                serializer.data
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error retrieving approval status. {str(e)}",
                500,
                []
            ), status=200)

    def put(self, request):
        try:
            candidate_id = request.data.get("candidate_id")
            role_id = request.data.get("role_id")
            decision = request.data.get("decision")
            comment = request.data.get("comment", "")

            if not candidate_id or not role_id or not decision:
                raise ValueError("Missing candidate_id, role_id, or decision")

            role_map = {
                1: "HM",
                3: "FPNA"
            }
            role = role_map.get(int(role_id))
            if not role:
                raise ValueError("Invalid role_id")

            candidate = Candidate.objects.get(pk=candidate_id)
            approval = CandidateApproval.objects.filter(candidate=candidate, role=role).first()

            if not approval:
                return Response(api_json_response_format(
                    False,
                    f"No approval record found for role '{role}' and candidate ID {candidate_id}",
                    404,
                    {}
                ), status=200)

            approval.decision = decision
            approval.comment = comment
            approval.reviewed_at = timezone.now()
            approval.save()

            all_approvals = CandidateApproval.objects.filter(candidate=candidate)
            approvers = Approver.objects.filter(requisition=candidate.Req_id_fk, set_as_approver="Yes")
            expected_roles = [a.role for a in approvers]
            overall_status = compute_overall_status(all_approvals, expected_roles)

            return Response(api_json_response_format(
                True,
                f"{role} decision updated successfully",
                200,
                {
                    "candidate_id": candidate.CandidateID,
                    "role": role,
                    "decision": approval.decision,
                    "comment": approval.comment,
                    "reviewed_at": approval.reviewed_at,
                    "overall_status": overall_status
                }
            ), status=200)

        except Candidate.DoesNotExist:
            return Response(api_json_response_format(
                False,
                "Candidate not found",
                404,
                {}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error updating approval decision: {str(e)}",
                500,
                {}
            ), status=200)

class OfferApprovalStatusView(APIView):
    def get(self, request):
        try:
            negotiations = OfferNegotiation.objects.filter(
                negotiation_status="Successful"
            ).select_related("requisition").prefetch_related("approvals__approver")

            grouped_results = {}

            for offer in negotiations:
                requisition = offer.requisition
                candidate = Candidate.objects.filter(Req_id_fk=requisition.RequisitionID).first()
                approvals = ApprovalStatus.objects.filter(offer_negotiation=offer).select_related("approver")

                approver_details = []
                expected_roles = []

                for approval in approvals:
                    approver = approval.approver
                    if approver:
                        expected_roles.append(approver.role)
                        approver_details.append({
                            "role": approver.role,
                            "name": f"{approver.first_name} {approver.last_name}",
                            "email": approver.email,
                            "contact_number": approver.contact_number,
                            "job_title": approver.job_title,
                            "status": approver.set_as_approver,
                            "decision": approval.status if approval.status else "Awaiting",
                            # "comment": approval.comment or ""
                        })

                overall_status = resolve_offer_approval_status(approvals, expected_roles)

                candidate_id = candidate.CandidateID if candidate else None
                candidate_first_name = offer.first_name
                candidate_last_name = offer.last_name

                grouped_results[candidate_id] = {
                    "req_id": requisition.RequisitionID,
                    "client_id": requisition.client_id,
                    "client_name": offer.client_name,
                    "candidate_id": candidate_id,
                    "candidate_first_name": candidate_first_name,
                    "candidate_last_name": candidate_last_name,
                    "approvers": approver_details,
                    "overall_status": overall_status,
                    "no_of_approvers": len(approver_details)
                }

            final_result = list(grouped_results.values())

            return Response({
                "success": True,
                "message": "Grouped offer approval status retrieved!",
                "error_code": 200,
                "data": final_result
            }, status=200)

        except Exception as e:
            return Response({
                "success": False,
                "message": f"Error retrieving approval status. {str(e)}",
                "error_code": 500,
                "data": []
            }, status=200)





class ApproverFilterView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            hiring_plan_id = request.data.get("hiring_plan")
            if not hiring_plan_id:
                return Response(api_json_response_format(
                    False,
                    "hiring_plan is required in request body",
                    400,
                    []
                ), status=200)

            approvers = Approver.objects.filter(hiring_plan_id=hiring_plan_id)
            serializer = ApproverSerializer(approvers, many=True)
            return Response(api_json_response_format(
                True,
                "Approvers filtered by hiring plan retrieved successfully!",
                200,
                serializer.data
            ), status=200)
        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error filtering approvers. {str(e)}",
                500,
                []
            ), status=200)

class OfferNegotiationViewSet(viewsets.ModelViewSet):
    queryset = OfferNegotiation.objects.prefetch_related('benefits').all()
    serializer_class = OfferNegotiationSerializer

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            enriched_data = []

            for obj, item in zip(queryset, serializer.data):
                # Defensive fetch to avoid AttributeError
                candidate = getattr(obj, "candidate", None)
                requisition = getattr(obj, "requisition", None)

                item["Req ID"] = getattr(requisition, "RequisitionID", "") if requisition else "N/A"
                item["Candidate ID"] = getattr(candidate, "CandidateID", "") if candidate else "N/A"
                item["Candidate First Name"] = getattr(candidate, "candidate_first_name", "") if candidate else "N/A"
                item["Candidate Last Name"] = getattr(candidate, "candidate_last_name", "") if candidate else "N/A"

                enriched_data.append(item)
            
            return Response(api_json_response_format(
                True, "Job requisitions retrieved successfully!", 200, enriched_data
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error retrieving job requisitions. {str(e)}", 500, {}
            ), status=500)  # Make sure status reflects the actual error

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error retrieving job requisitions. {str(e)}", 500, {}
            ), status=200)

    def create(self, request, *args, **kwargs):
        try:
            offer_id = request.data.get("id")
            if not offer_id:
                return Response(api_json_response_format(False, "Missing offer ID", 400, {}), status=200)

            offer_data = request.data
            requisition_id = offer_data.get("requisition")

            # Step 1: Validate requisition lookup
            try:
                requisition = JobRequisition.objects.get(RequisitionID=requisition_id)
            except JobRequisition.DoesNotExist:
                return Response(api_json_response_format(False, "Requisition not found.", 404, {}), status=200)

            # Step 2: Get offer object
            try:
                offer = OfferNegotiation.objects.get(id=offer_id)
            except OfferNegotiation.DoesNotExist:
                return Response(api_json_response_format(False, "Offer negotiation not found.", 404, {}), status=200)

            # Step 3: Handle benefits safely
            benefit_data = offer_data.get("benefits", [])
            benefit_ids = []
            for item in benefit_data:
                try:
                    if isinstance(item, dict):
                        benefit_ids.append(int(item.get("id")))
                    else:
                        benefit_ids.append(int(item))
                except (ValueError, TypeError):
                    continue  # Could log invalid items if needed

            offer.benefits.set(benefit_ids)

            # Step 4: Assign other fields
            for field, value in offer_data.items():
                if hasattr(offer, field) and field not in ["id", "requisition", "benefits"]:
                    setattr(offer, field, value)

            offer.requisition = requisition
            offer.save()

            # Step 5: Serialize and respond
            serializer = self.get_serializer(offer)
            
            # Step 6: Auto-create ApprovalStatus records if status is 'Successful'
            if offer.negotiation_status == "Successful":
                approvers = Approver.objects.filter(
                    requisition__RequisitionID=requisition_id,
                    set_as_approver__in=["Yes", "Maybe"]
                )

                approval_instances = [
                    ApprovalStatus(
                        offer_negotiation=offer,
                        approver=approver,
                        status="Pending"
                    )
                    for approver in approvers
                    if not ApprovalStatus.objects.filter(offer_negotiation=offer, approver=approver).exists()
                ]

                ApprovalStatus.objects.bulk_create(approval_instances)
                offer.notify_pending_approvers()


            # Step 7: Serialize and respond
            serializer = self.get_serializer(offer)
            return Response(api_json_response_format(True, "Offer negotiation updated successfully!", 200, serializer.data), status=200)

            return Response(api_json_response_format(True, "Offer negotiation updated successfully!", 200, serializer.data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error updating offer negotiation. {str(e)}", 500, {}), status=200)



    @action(detail=False, methods=['get'], url_path='pending-approvals')
    def pending_approvals(self, request):
        try:
            negotiations = OfferNegotiation.objects.filter(
                negotiation_status='Successful'
            ).select_related('requisition').prefetch_related('approvals__approver')

            data = []

            for i, offer in enumerate(negotiations, start=1):
                candidate = Candidate.objects.filter(
                    Req_id_fk=offer.requisition.RequisitionID
                ).first()

                # Build { role: status } dictionary
                role_map = {a.approver.role: a.status for a in offer.approvals.all()}

                def get_display_status(status):
                    return "Yes" if status == "Approved" else "Awaiting" if status == "Pending" else status or "N/A"

                # Core static data
                row = {
                    "SNo": i,
                    "ReqID": offer.requisition.RequisitionID,
                    "ClientID": f"{offer.requisition.id:04}",
                    "ClientName": offer.client_name,
                    "CandidateID": candidate.CandidateID if candidate else None,
                    "CandidateFirstName": offer.first_name,
                    "CandidateLastName": offer.last_name,
                    "Status": (
                        "Waiting for approval"
                        if any(status == "Pending" for status in role_map.values())
                        else offer.negotiation_status
                    ),

                }

                # Add dynamic approver columns
                for role, status in role_map.items():
                    column_name = f"{role} Status" if "Status" not in role else role
                    row[column_name] = get_display_status(status)

                data.append(row)

            return Response(api_json_response_format(
                True, "Pending approvals fetched successfully.", 200, data
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error fetching pending approvals. {str(e)}", 500, {}
            ), status=200)

    @action(detail=False, methods=['post'], url_path='get-by-id')
    def get_by_id(self, request):
        try:
            offer_id = request.data.get("id")
            if not offer_id:
                return Response(api_json_response_format(
                    False, "Missing offer ID in request body.", 400, {}
                ), status=200)

            try:
                offer = OfferNegotiation.objects.prefetch_related('benefits').get(id=offer_id)
            except OfferNegotiation.DoesNotExist:
                return Response(api_json_response_format(
                    False, "Offer negotiation not found.", 404, {}
                ), status=200)

            serializer = self.get_serializer(offer)
            return Response(api_json_response_format(
                True, "Offer negotiation retrieved successfully!", 200, serializer.data
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error fetching offer negotiation. {str(e)}", 500, {}
            ), status=200)



    @action(detail=False, methods=['post'], url_path='resend-approval-emails')
    def resend_approval_emails(self, request):
        offer_id = request.data.get("offer_id")
        if not offer_id:
            return Response(api_json_response_format(False, "Missing offer_id", 400, {}), status=200)

        try:
            offer = OfferNegotiation.objects.get(id=offer_id)
            offer.notify_pending_approvers()
            return Response(api_json_response_format(True, "Approval emails resent successfully.", 200, {}), status=200)

        except OfferNegotiation.DoesNotExist:
            return Response(api_json_response_format(False, "Offer negotiation not found.", 404, {}), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error resending emails: {str(e)}", 500, {}), status=200)


def approve_offer(request, negotiation_id):
    email = request.GET.get('email')
    role = request.GET.get('role')

    if not email or not role:
        return HttpResponse("Invalid approval link. Missing parameters.", status=200)

    negotiation = get_object_or_404(OfferNegotiation, pk=negotiation_id)
    approver = get_object_or_404(Approver, email=email, role=role)
    approval_status = get_object_or_404(ApprovalStatus, offer_negotiation=negotiation, approver=approver)

    if approval_status.status != "Pending":
        return HttpResponse(f"You already responded: {approval_status.status}", status=200)

    approval_status.status = "Approved"
    approval_status.save()

    return HttpResponse(f"✅ Thank you! Your approval for {negotiation.first_name} {negotiation.last_name} has been recorded.", status=200)

def reject_offer(request, negotiation_id):
    email = request.GET.get('email')
    role = request.GET.get('role')

    if not email or not role:
        return HttpResponse("Invalid rejection link. Missing parameters.", status=200)

    negotiation = get_object_or_404(OfferNegotiation, pk=negotiation_id)
    approver = get_object_or_404(Approver, email=email, role=role)
    approval_status = get_object_or_404(ApprovalStatus, offer_negotiation=negotiation, approver=approver)

    if approval_status.status != "Pending":
        return HttpResponse(f"You already responded: {approval_status.status}", status=200)

    approval_status.status = "Rejected"
    approval_status.save()

    return HttpResponse(f"❌ Your rejection has been recorded for {negotiation.first_name} {negotiation.last_name}.", status=200)

class GetInterviewScheduleAPIView(APIView):
    def post(self, request):
        try:
            email = request.data.get("email")
            if not email:
                return Response(api_json_response_format(False, "Missing interviewer email", 400, {}), status=200)

            schedules = InterviewSchedule.objects.filter(
                interviewer__email=email
            ).select_related('candidate', 'candidate__Req_id_fk')

            data = []
            for schedule in schedules:
                candidate = schedule.candidate
                requisition = candidate.Req_id_fk

                data.append({
                    "Schedule ID": schedule.id,
                    "Candidate First Name": candidate.candidate_first_name,
                    "Candidate Last Name":  candidate.candidate_last_name,
                    "Requisition ID": requisition.RequisitionID,
                    "Position Title": requisition.PositionTitle,
                    "Interview Date": schedule.date.strftime("%Y-%m-%d"),
                    "Start Time": schedule.start_time.strftime("%H:%M"),
                    "End Time": schedule.end_time.strftime("%H:%M"),
                    "Round Name": schedule.round_name,
                    "Meet Link": schedule.meet_link
                })

            return Response(api_json_response_format(True, "Fetched scheduled interviews", 200, data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Failed to fetch interviews: {e}", 500, {}), status=200)
        
        
class InterviewReportAPIView(APIView):
    def get(self, request):
        try:
            interviewers = Interviewer.objects.all()
            data = []

            for interviewer in interviewers:
                req_id = interviewer.req_id.RequisitionID if interviewer.req_id else ""

                # 🔄 Sort interviews by date and start_time descending
                interviews = InterviewSchedule.objects.filter(interviewer=interviewer)\
                    .select_related('candidate')\
                    .order_by('-created_at')

                for interview in interviews:
                    candidate = interview.candidate
                    requisition = candidate.Req_id_fk if candidate else None

                    stage = CandidateInterviewStages.objects.filter(
                        candidate_id=candidate.CandidateID,
                        Req_id=req_id,
                        interview_stage=interview.round_name
                    ).first()

                    data.append({
                        "Schedule ID": interview.id,
                        "Req ID": req_id,
                        "Client ID": requisition.client_id if requisition and requisition.client_id else "",
                        "First Name": interviewer.first_name,
                        "Last Name": interviewer.last_name,
                        "Job Title": interviewer.job_title,
                        "Interview Mode": interviewer.interview_mode,
                        "Interviewer Stage": interviewer.interviewer_stage,
                        "Email ID": interviewer.email,
                        "Candidate ID": candidate.CandidateID if candidate else "",
                        "Candidate First Name": candidate.candidate_first_name if candidate else "",
                        "Candidate Last Name": candidate.candidate_last_name if candidate else "",
                        "Role": requisition.PositionTitle if requisition else "",
                        "Feedback": stage.feedback if stage else "",
                        "Interview Results": stage.result if stage else "",
                        "Contact Number": interviewer.contact_number,
                        "Round Name": interview.round_name,
                        "Interview Date": interview.date.strftime("%Y-%m-%d"),
                        "Start Time": interview.start_time.strftime("%H:%M"),
                        "End Time": interview.end_time.strftime("%H:%M"),
                        "Meet Link": interview.meet_link
                    })

            return Response(api_json_response_format(
                True, "Fetched full interview report", 200, data
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error fetching interview report: {str(e)}", 500, {}
            ), status=200)


class SubmitInterviewReviewView(APIView):
    def get(self, request):
        try:
            data = []

            schedules = InterviewSchedule.objects.select_related("candidate", "candidate__Req_id_fk").all()

            for schedule in schedules:
                reviews = InterviewReview.objects.filter(schedule=schedule, candidate=schedule.candidate)

                review_data = []
                for i, review in enumerate(reviews, start=1):
                    review_data.append({
                        "sno": i,
                        "parameterDefined": review.ParameterDefined,
                        "Guidelines": review.Guidelines,
                        "MinimumQuestions": str(review.MinimumQuestions),
                        "weightage": str(review.Weightage),
                        "ActualRating": review.ActualRating,
                        "Feedback": review.Feedback_param
                    })

                stage = CandidateInterviewStages.objects.filter(
                    candidate_id=schedule.candidate.CandidateID,
                    Req_id=schedule.candidate.Req_id_fk.RequisitionID,
                    interview_stage=schedule.round_name
                ).first()

                entry = {
                    "schedule_id": schedule.id,
                    "reqId": schedule.candidate.Req_id_fk.RequisitionID,
                    "candidate_id": schedule.candidate.CandidateID,
                    "final_rating": stage.final_rating if stage else 0,
                    "result": stage.result if stage else "",
                    "final_feedback": stage.feedback if stage else "",
                    "reviews": review_data
                }

                data.append(entry)

            return Response(api_json_response_format(True, "All interview reviews retrieved", 200, data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Failed to retrieve data: {e}", 500, {}), status=200)

    def post(self, request):
        try:
            req_id          = request.data.get("reqId")
            candidate_id    = request.data.get("candidate_id")
            final_rating    = request.data.get("final_rating", 0)
            result          = request.data.get("result", "")
            final_feedback  = request.data.get("final_feedback", "")
            reviews         = request.data.get("reviews", [])

            if not req_id or not candidate_id:
                return Response(api_json_response_format(False, "Missing reqId or candidate_id", 400, {}), status=200)

            # Fetch schedule based on candidate and requisition
            schedule = InterviewSchedule.objects.filter(
                candidate__CandidateID=candidate_id,
                candidate__Req_id_fk__RequisitionID=req_id
            ).first()

            if not schedule:
                return Response(api_json_response_format(False, "Schedule not found", 404, {}), status=200)

            # Create or update InterviewReview for each parameter
            for review in reviews:
                InterviewReview.objects.update_or_create(
                    schedule=schedule,
                    candidate=schedule.candidate,
                    ParameterDefined=review.get("parameterDefined", ""),

                    defaults={
                        "Guidelines": review.get("Guidelines", ""),
                        "MinimumQuestions": review.get("MinimumQuestions", ""),
                        "Weightage": review.get("weightage", 0),
                        "ActualRating": review.get("ActualRating", 0),
                        "Feedback_param": review.get("Feedback", "")
                    }

                )

            # Update CandidateInterviewStages with final feedback
            CandidateInterviewStages.objects.filter(
                candidate_id=schedule.candidate.CandidateID,
                Req_id=schedule.candidate.Req_id_fk.RequisitionID,
                interview_stage=schedule.round_name
            ).update(
                status="Completed",
                feedback=final_feedback,
                final_rating=final_rating,
                result=result
            )

            return Response(api_json_response_format(True, "Feedback saved", 200, {}), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Failed to save review: {e}", 500, {}), status=200)

    def put(self, request):
        try:
            schedule_id     = request.data.get("schedule_id")
            req_id          = request.data.get("reqId")
            candidate_id    = request.data.get("candidate_id")
            final_rating    = request.data.get("final_rating", 0)
            result          = request.data.get("result", "")
            final_feedback  = request.data.get("final_feedback", "")
            reviews         = request.data.get("reviews", [])

            if not schedule_id or not req_id or not candidate_id:
                return Response(api_json_response_format(False, "Missing schedule_id, reqId, or candidate_id", 400, {}), status=200)

            schedule = InterviewSchedule.objects.filter(
                id=schedule_id,
                candidate__CandidateID=candidate_id,
                candidate__Req_id_fk__RequisitionID=req_id
            ).first()

            if not schedule:
                return Response(api_json_response_format(False, "Schedule not found", 404, {}), status=200)

            # Update each review
            for review in reviews:
                InterviewReview.objects.update_or_create(
                    schedule=schedule,
                    candidate=schedule.candidate,
                    ParameterDefined=review.get("parameterDefined", ""),
                    defaults={
                        "Guidelines": review.get("Guidelines", ""),
                        "MinimumQuestions": review.get("MinimumQuestions", ""),
                        "Weightage": int(review.get("weightage", 0)),
                        "ActualRating": int(review.get("ActualRating", 0)),
                        "Feedback_param": review.get("Feedback", "")
                    }
                )

            # Update CandidateInterviewStages
            CandidateInterviewStages.objects.filter(
                candidate_id=schedule.candidate.CandidateID,
                Req_id=schedule.candidate.Req_id_fk.RequisitionID,
                interview_stage=schedule.round_name
            ).update(
                final_rating=final_rating,
                result=result,
                feedback=final_feedback,
                status="Completed"
            )

            return Response(api_json_response_format(True, "Interview review updated", 200, {}), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Failed to update review: {str(e)}", 500, {}), status=200)


class GetInterviewReviewsByCandidate(APIView):
    def post(self, request):
        try:
            candidate_id = request.data.get("candidate_id")
            if not candidate_id:
                return Response(api_json_response_format(False, "Missing candidate_id", 400, {}), status=200)

            reviews = InterviewReview.objects.filter(candidate_id=candidate_id)
            serializer = InterviewReviewSerializer(reviews, many=True)

            return Response(api_json_response_format(True, "Reviews fetched", 200, serializer.data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Failed to fetch reviews: {e}", 500, {}), status=200)


def get_fully_approved_candidates(instance):
    candidates = instance.candidates.all()
    approved_candidates = []

    for candidate in candidates:
        approvals = CandidateApproval.objects.filter(candidate_id=candidate.CandidateID)

        # Must have at least one approval, and all approvals must be "Approve"
        if approvals.exists() and not approvals.exclude(decision="Approved").exists():
            approved_candidates.append(candidate)

    return approved_candidates


class ScheduleContextAPIView(APIView):
    def post(self, request, *args, **kwargs):
        req_id = request.data.get("req_id")
        if not req_id:
            return Response(
                api_json_response_format(False, "Missing requisition_id", 400, {}),
                status=200
            )

        try:
            instance = JobRequisition.objects.prefetch_related(
                "candidates", "interviewer__slots", "interviewer"
            ).get(RequisitionID=req_id)
        except JobRequisition.DoesNotExist:
            return Response(
                api_json_response_format(False, "Requisition not found", 404, {}),
                status=200
            )

        # 🔍 Design screen and round metadata
        design_screen = InterviewDesignScreen.objects.filter(req_id=req_id).first()
        interview_design_id = design_screen.interview_design_id if design_screen else None

        round_metadata = []
        if interview_design_id:
            design_params = InterviewDesignParameters.objects.filter(
                interview_design_id=interview_design_id
            ).order_by("interview_desing_params_id")
            round_metadata = [
                {
                    "score_card": param.score_card,
                    "round_no": param.round_no.strip() if param.round_no else ""
                }
                for param in design_params
            ]


        candidate = instance.candidates.first()
        if not candidate:
            return Response(
                api_json_response_format(False, "No candidate linked to requisition", 404, {}),
                status=200
            )

        approved_candidates = get_fully_approved_candidates(instance)
        if not approved_candidates:
            return Response(
                api_json_response_format(False, "Candidate not approved by all approvers", 400, {}),
                status=200
            )

        candidate_name = f"{candidate.candidate_first_name} {candidate.candidate_last_name}"

        # 🔁 Build interviewer-specific schedules
        interviewer_payload = []
        for i, interviewer in enumerate(instance.interviewer.all()):
            slot_qs = InterviewSlot.objects.filter(interviewer=interviewer).order_by("date", "start_time")[:5]
            slots = [
                {
                    "date": slot.date.strftime("%Y-%m-%d"),
                    "time": slot.start_time.strftime("%I:%M %p")
                } for slot in slot_qs
            ]

            round_info = round_metadata[i] if i < len(round_metadata) else {}
            round_name = round_info.get("score_card", f"Round {i+1}")
            round_no = round_info.get("round_no", "").strip()
            round_no_suffix = f" R-{round_no}" if round_no.isdigit() else ""


            interviewer_payload.append({
                "name": f"{interviewer.first_name} {interviewer.last_name}{round_no_suffix}",
                "mode": interviewer.interview_mode if interviewer.interview_mode else "Not Specified",
                "round_name": round_name,
                "slots": slots
            })


        # 🎯 Final payload
        response_payload = {
            "req_id": instance.RequisitionID,
            "job_position": instance.PositionTitle,
            "planning_id": instance.Planning_id.hiring_plan_id if instance.Planning_id else "N/A",
            "candidate_name": candidate_name,
            "interviewers": interviewer_payload,
            "location": "Zoom",
            "time_zone": "IST",
            "durations": "30 mins",
            "purpose": "Technical Screening"
        }

        return Response(
            api_json_response_format(True, "Fetched scheduling context successfully", 200, {"data": response_payload}),
            status=200
        )

class InterviewerContextAPIView(APIView):
    def post(self, request):
        interviewer_id = request.data.get("interviewer_id")
        if not interviewer_id:
            return Response(api_json_response_format(False, "Missing interviewer ID", 400, {}), status=200)

        try:
            interviewer = Interviewer.objects.get(interviewer_id=interviewer_id)
        except Interviewer.DoesNotExist:
            return Response(api_json_response_format(False, "Interviewer not found", 404, {}), status=200)

        req_id = getattr(interviewer, "req_id_id", None)
        rounds_payload = []

        # Step 1: Get all slots for the given interviewer
        all_slots = list(InterviewSlot.objects.filter(
            interviewer=interviewer
        ).order_by("date", "start_time"))

        if req_id:
            try:
                # Step 2: Get requisition and list of interviewers
                requisition = JobRequisition.objects.prefetch_related("interviewer").get(RequisitionID=req_id)
                interviewer_list = list(requisition.interviewer.all())

                if interviewer in interviewer_list:
                    index = interviewer_list.index(interviewer)

                    # Step 3: Get design rounds for this requisition
                    design_screen = InterviewDesignScreen.objects.filter(req_id=req_id).first()
                    interview_design_id = getattr(design_screen, "interview_design_id", None)

                    if interview_design_id:
                        round_list = list(InterviewDesignParameters.objects.filter(
                            interview_design_id=interview_design_id
                        ).order_by("interview_desing_params_id"))

                        # Step 4: Pick round by interviewer index only
                        if index < len(round_list):
                            round_obj = round_list[index]
                            round_name = round_obj.score_card or f"Round {index + 1}"

                            # Assign first 5 available slots for that interviewer
                            slots_chunk = all_slots[:5]

                            formatted_slots = [
                                {
                                    "date": slot.date.strftime("%Y-%m-%d"),
                                    "time": slot.start_time.strftime("%I:%M %p")
                                } for slot in slots_chunk
                            ]

                            rounds_payload.append({
                                "round_name": round_name,
                                "slots": formatted_slots
                            })

            except JobRequisition.DoesNotExist:
                pass

        response_payload = {
            "mode": interviewer.interview_mode or "Not Specified",
            "rounds": rounds_payload
        }

        return Response(api_json_response_format(True, "Interviewer context retrieved", 200, response_payload), status=200)

class InterviewerListCreateAPIView(generics.ListCreateAPIView):
    queryset = Interviewer.objects.prefetch_related('slots').all()
    serializer_class = InterviewerSerializer

    def list(self, request, *args, **kwargs):
        try:
            queryset = Interviewer.objects.select_related('req_id').prefetch_related('slots').order_by('-created_at')
            serializer = self.get_serializer(queryset, many=True)
            data = serializer.data

            for idx, interviewer in enumerate(queryset):
                requisition_id = interviewer.req_id.RequisitionID
                job_req = JobRequisition.objects.filter(RequisitionID=requisition_id).select_related('Planning_id').first()
                planning_id = job_req.Planning_id.hiring_plan_id if job_req and job_req.Planning_id else None
                data[idx]['planning_id'] = planning_id

            return Response(api_json_response_format(
                True,
                "Interviewer list retrieved successfully",
                200,
                data
            ), status=200)


        except Exception as e:
            return Response(api_json_response_format(
                False,
                str(e),
                500,
                []
            ), status=200)


    def create(self, request, *args, **kwargs):
        try:
            if 'file' in request.FILES:
                # ✅ Save manually to desired folder
                file = request.FILES['file']
                os.makedirs(RESUME_STORAGE_FOLDER, exist_ok=True)
                file_path = os.path.join(RESUME_STORAGE_FOLDER, file.name)
                
                with open(file_path, "wb") as f:
                    f.write(file.read())

                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

                created_count = 0
                skipped_rows = []
                
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    excel_date = row[9].value
                    parsed_date = excel_date.date() if isinstance(excel_date, datetime) else excel_date

                    data = {
                        "req_id": row[0].value,
                        "client_id": row[1].value,
                        "first_name": row[2].value,
                        "last_name": row[3].value,
                        "job_title": row[4].value,
                        "interview_mode": row[5].value,
                        "interviewer_stage": row[6].value,
                        "email": row[7].value,
                        "contact_number": row[8].value,
                        "slots": [{
                            "date": parsed_date,
                            "start_time": row[10].value,
                            "end_time": row[11].value
                        }]

                    }

                    exists = Interviewer.objects.filter(
                        req_id__RequisitionID=data["req_id"],
                        email=data["email"],
                        interviewer_stage=data["interviewer_stage"],
                        slots__date=data["slots"][0]["date"],
                        slots__start_time=data["slots"][0]["start_time"]
                    ).exists()

                    if exists:
                        skipped_rows.append(row_idx)
                        continue

                    serializer = self.get_serializer(data=data)
                    if serializer.is_valid():
                        serializer.save()
                        created_count += 1
                    else:
                        return Response(api_json_response_format(
                            False,
                            f"Error at row {row_idx}: {serializer.errors}",
                            400,
                            {}
                        ), status=200)

                return Response(api_json_response_format(
                    True,
                    f"{created_count} interviewers created. Skipped rows: {skipped_rows}",
                    200,
                    {}
                ), status=200)


            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(api_json_response_format(True, "Created successfully", 200, {}), status=200)
        except Exception as e:
            return Response(api_json_response_format(False, str(e), 400, {}), status=200)

    def put(self, request, *args, **kwargs):
        try:
            interviewer_id = request.data.get("interviewer_id")
            if not interviewer_id:
                return Response(api_json_response_format(False, "Missing 'id' in request body", 400, {}), status=200)

            interviewer = Interviewer.objects.get(pk=interviewer_id)
            serializer = self.get_serializer(interviewer, data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(api_json_response_format(True, "Updated successfully", 200, serializer.data), status=200)
        except Interviewer.DoesNotExist:
            return Response(api_json_response_format(False, "Interviewer not found", 404, {}), status=200)
        except Exception as e:
            return Response(api_json_response_format(False, str(e), 500, {}), status=200)

    def delete(self, request, *args, **kwargs):
        try:
            interviewer_id = request.data.get("interviewer_id")
            if not interviewer_id:
                return Response(api_json_response_format(False, "Missing 'id' in request body", 400, {}), status=200)

            interviewer = Interviewer.objects.get(pk=interviewer_id)

            # 🧹 Delete Interview Slots
            InterviewSlot.objects.filter(interviewer=interviewer).delete()

            # 🧹 Find Related Schedules
            schedules = InterviewSchedule.objects.filter(interviewer=interviewer)

            # 🧹 Delete Reviews linked to those schedules
            InterviewReview.objects.filter(schedule__in=schedules).delete()

            # 🧹 Delete Interview Schedules
            schedules.delete()

            # 🧹 Optionally delete CandidateInterviewStages based on stage match
            stage = interviewer.interviewer_stage.strip().lower()
            CandidateInterviewStages.objects.filter(
                Req_id=interviewer.req_id,
                interview_stage__iexact=stage
            ).delete()


            # 🧹 Finally delete the Interviewer
            interviewer.delete()

            return Response(api_json_response_format(True, "Deleted successfully with related records", 200, {}), status=200)

        except Interviewer.DoesNotExist:
            return Response(api_json_response_format(False, "Interviewer not found", 404, {}), status=200)
        except Exception as e:
            return Response(api_json_response_format(False, str(e), 500, {}), status=200)

class ScheduleMeetView(APIView):
    def get(self, request):
        try:
            schedules = InterviewSchedule.objects.select_related("candidate", "interviewer").all().order_by("-date", "-start_time")
            result = []

            for schedule in schedules:
                candidate_full_name = f"{schedule.candidate.candidate_first_name} {schedule.candidate.candidate_last_name}"
                interviewer_full_name = f"{schedule.interviewer.first_name} {schedule.interviewer.last_name}"

                result.append({
                    "schedule_id": schedule.id,
                    "req_id": schedule.interviewer.req_id.RequisitionID if schedule.interviewer.req_id else "Not Linked",
                    "planning_id": (
                        schedule.candidate.Req_id_fk.Planning_id.hiring_plan_id
                        if schedule.candidate.Req_id_fk and schedule.candidate.Req_id_fk.Planning_id else "N/A"
                    ),
                    "candidate_name": candidate_full_name,
                    "interviewers_id": schedule.interviewer.interviewer_id,
                    "interviewer_name": interviewer_full_name,
                    "job_position": schedule.candidate.Req_id_fk.PositionTitle if schedule.candidate.Req_id_fk else "Unknown",
                    "location": schedule.location,
                    "time_zone": schedule.time_zone,
                    "durations": schedule.durations,
                    "purpose": schedule.purpose,
                    "mode": schedule.mode,
                    "no_of_rounds": 1,  # Assuming 1 by default; adjust if stored differently
                    "meet_link": schedule.meet_link,
                    "guests": schedule.guests or [],
                    "schedule_slots": [
                        {
                            "date": str(schedule.date),
                            "time": schedule.start_time.strftime("%I:%M %p"),
                            "guests": [guest.get("email") for guest in schedule.guests] if schedule.guests else []
                        }
                    ]
                })

            return Response(api_json_response_format(True, "All schedules retrieved", 200, result), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error retrieving schedule list: {str(e)}", 500, {}), status=200)

    def post(self, request):
        try:
            payload = request.data
            req_id        = payload.get("req_id")
            planning_id   = payload.get("planning_id")
            candidate_name = payload.get("candidate_name", "")
            interviewer_name = payload.get("interviewer_name", "")
            round_name    = payload.get("purpose", "Interview")
            summary       = payload.get("purpose", f"{round_name} Interview")

            # Extract first schedule slot
            slot = payload.get("schedule_slots", [{}])[0]
            date = slot.get("date")
            time = slot.get("time")
            duration_raw = payload.get("durations", "30 mins")
            duration_mins = int(duration_raw.split()[0]) if duration_raw else 30

            if not all([req_id, candidate_name, interviewer_name, date, time]):
                return Response(api_json_response_format(False, "Missing required fields", 400, {}), status=200)

            tz    = pytz.timezone("Asia/Kolkata")
            start = tz.localize(datetime.strptime(f"{date} {time}", "%Y-%m-%d %I:%M %p"))
            end   = start + timedelta(minutes=duration_mins)

            # Resolve Candidate and Interviewer by name + requisition
            candidate = Candidate.objects.get(
                candidate_first_name__iexact=candidate_name.split()[0],
                candidate_last_name__iexact=candidate_name.split()[1],
                Req_id_fk__RequisitionID=req_id
            )

            matches = Interviewer.objects.filter(
                first_name__iexact=interviewer_name.split()[0],
                last_name__iexact=interviewer_name.split()[1],
                req_id__RequisitionID=req_id,
                interviewer_stage__iexact=round_name
            )


            # Refine by round_name if multiple interviewers matched
            # if matches.count() > 1:
            #     matches = matches.filter(interview_stage__iexact=round_name).distinct()

            if matches.count() == 0:
                return Response(api_json_response_format(
                    False,
                    "No matching interviewer found with the given name, requisition, and round name.",
                    404,
                    {}
                ), status=200)

            interviewer = matches.first()


            guest_emails = [g.get("email") for g in payload.get("guests", []) if g.get("email")]

            join_url = schedule_zoom_meet(
                topic=summary,
                start_time_iso=start.isoformat(),
                duration_minutes=duration_mins,
                timezone="Asia/Kolkata"
            )

            InterviewSchedule.objects.create(
                candidate=candidate,
                interviewer=interviewer,
                round_name=round_name,
                date=start.date(),
                start_time=start.time(),
                end_time=end.time(),
                meet_link=join_url,
                location=payload.get("location"),
                time_zone=payload.get("time_zone"),
                purpose=payload.get("purpose"),
                mode=payload.get("mode"),
                guests=payload.get("guests"),
                durations=payload.get("durations")
            )
            CandidateInterviewStages.objects.create(
                candidate_id=candidate.CandidateID,
                Req_id_id=candidate.Req_id_fk.RequisitionID,
                interview_date=start.date(),
                interview_stage=round_name,
                mode_of_interview=payload.get("mode", ""),
                status="Stage Scheduled",
                feedback=""
            )
            subject = f"{summary} - Scheduled on {start.strftime('%d %b %Y at %I:%M %p')}"
            plain_message = (
                f"Dear Participant,\n\n"
                f"You have a scheduled meeting:\n"
                f"Topic: {summary}\n"
                f"Date: {start.strftime('%Y-%m-%d')}\n"
                f"Time: {start.strftime('%I:%M %p')} IST\n"
                f"Duration: {duration_mins} minutes\n"
                f"Join Link: {join_url}\n\n"
                f"Please be ready at the scheduled time.\n\nBest regards."
            )

            html_message = f"""
            <p>Dear Participant,</p>
            <p>You have a scheduled meeting:</p>
            <ul>
            <li><strong>Topic:</strong> {summary}</li>
            <li><strong>Date:</strong> {start.strftime('%Y-%m-%d')}</li>
            <li><strong>Time:</strong> {start.strftime('%I:%M %p')} IST</li>
            <li><strong>Duration:</strong> {duration_mins} minutes</li>
            <li><strong>Join Link:</strong> <a href="{join_url}">{join_url}</a></li>
            </ul>
            <p>Please be ready at the scheduled time.</p>
            <p>Best regards,<br>Your Hiring Team</p>
            """

            # Collect all recipients
            recipients = list(set([candidate.Email, interviewer.email] + guest_emails))


            # Send email
            email = EmailMultiAlternatives(
                subject=subject,
                body=plain_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=recipients
            )
            email.attach_alternative(html_message, "text/html")
            email.send(fail_silently=True)




            return Response(api_json_response_format(
                True,
                "Interview scheduled successfully",
                200,
                {"meet_link": join_url}
            ), status=200)

        except Candidate.DoesNotExist:
            return Response(api_json_response_format(False, "Candidate not found using given name and requisition.", 404, {}), status=200)
        except Interviewer.DoesNotExist:
            return Response(api_json_response_format(False, "Interviewer not found using given name and requisition.", 404, {}), status=200)
        except ValueError as e:
            return Response(api_json_response_format(False, f"Invalid datetime format: {e}", 400, {}), status=200)
        except Exception as e:
            return Response(api_json_response_format(False, f"Failed to schedule: {str(e)}", 500, {}), status=200)



@api_view(['POST'])
def add_candidate_feedback(request):
    base_data = request.data.copy()
    candidate_id = base_data.get('candidate')

    # 🔍 Validate candidate existence
    try:
        candidate = Candidate.objects.get(CandidateID=candidate_id)
    except Candidate.DoesNotExist:
        return Response(api_json_response_format(False, "Candidate not found", 400, {}), status=200)

    # 🔄 Fetch all interviews for candidate
    interviews = InterviewSchedule.objects.filter(candidate=candidate).order_by('date')
    if not interviews.exists():
        return Response(api_json_response_format(False, "No interviews found for candidate", 400, {}), status=200)

    feedback_records = []

    for interview in interviews:
        data = base_data.copy()

        # 🎯 Recruiter name
        interviewer = interview.interviewer
        first = interviewer.first_name if interviewer and interviewer.first_name else ""
        last = interviewer.last_name if interviewer and interviewer.last_name else ""
        full_name = f"{first} {last}".strip()
        data['recruiter_name'] = full_name if full_name else "Unknown Interviewer"

        # 📅 Interview date
        data['interview_date'] = interview.date

        # 🧠 Pull InterviewReview for this schedule
        review = InterviewReview.objects.filter(schedule=interview, candidate=candidate).order_by('-created_at').first()
        if review:
            data['assessment_score'] = str(review.ActualRating) if review.ActualRating is not None else ""
            data['interviewer_feedback'] = review.Feedback_param or ""
        else:
            data['assessment_score'] = ""
            data['interviewer_feedback'] = ""

        # 📝 Recruiter feedback from Candidate
        data['recruiter_feedback'] = candidate.Result or ""

        # 🚀 Validate and save each feedback record
        serializer = CandidateFeedbackEnrichedSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            feedback_records.append(serializer.data)
        else:
            return Response(api_json_response_format(False, "Validation error", 400, {"errors": serializer.errors}), status=200)

    return Response(api_json_response_format(True, "Feedback saved", 200, feedback_records), status=200)

@api_view(['POST'])
def get_all_candidate_feedback(request):
    data = request.data
    client_name = data.get("client_name")
    position = data.get("position_considered_for")
    recruiter = data.get("recruiter_name")
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    skills = data.get("skills_expertise")  # Expecting comma-separated string or list

    feedbacks = CandidateFeedback.objects.select_related(
        'candidate',
        'candidate__Req_id_fk'
    ).order_by('-created_at')

    # Apply filters
    if client_name:
        feedbacks.filter(candidate__Req_id_fk__company_client_name__icontains=client_name)


    if position:
        feedbacks = feedbacks.filter(
            candidate__Req_id_fk__position_information__job_position__icontains=position
        )

    if recruiter:
        feedbacks = feedbacks.filter(
            recruiter_name__icontains=recruiter
        )

    if start_date and end_date:
        feedbacks = feedbacks.filter(interview_date__range=[start_date, end_date])
    elif start_date:
        feedbacks = feedbacks.filter(interview_date__gte=start_date)
    elif end_date:
        feedbacks = feedbacks.filter(interview_date__lte=end_date)

    if skills:
        if isinstance(skills, str):
            skills = [s.strip() for s in skills.split(",")]
        for skill in skills:
            feedbacks = feedbacks.filter(skills__icontains=skill)

    if not feedbacks.exists():
        return Response(api_json_response_format(False, "No feedback records found", 404, {}), status=200)

    serializer = CandidateFeedbackEnrichedSerializer(feedbacks, many=True)
    return Response(api_json_response_format(True, "Filtered feedback retrieved", 200, serializer.data), status=200)

@api_view(['GET'])
def get_declined_offer_report(request):
    client = request.GET.get('client_name')
    recruiter = request.GET.get('recruiter_name')
    position = request.GET.get('position_offered')
    location = request.GET.get('location')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    feedbacks = CandidateFeedback.objects.filter(status='Declined')

    if client:
        feedbacks = feedbacks.filter(candidate__Req_id_fk__company_client_name__icontains=client)
    if recruiter:
        feedbacks = feedbacks.filter(recruiter_name__icontains=recruiter)
    if position:
        feedbacks = feedbacks.filter(candidate__Req_id_fk__PositionTitle__icontains=position)
    if location:
        feedbacks = feedbacks.filter(current_location__icontains=location)
    if start_date and end_date:
        feedbacks = feedbacks.filter(candidate__offernegotiation__offered_doj__range=[start_date, end_date])
    elif start_date:
        feedbacks = feedbacks.filter(candidate__offernegotiation__offered_doj__gte=start_date)
    elif end_date:
        feedbacks = feedbacks.filter(candidate__offernegotiation__offered_doj__lte=end_date)

    feedbacks = feedbacks.select_related(
        'candidate',
        'candidate__Req_id_fk',
        'candidate__Req_id_fk__Planning_id'
    ).prefetch_related(
        'candidate__offer_negotiations','candidate__generated_offers'  # 👈 This is the correct way
    )

    def get_offer_negotiation(candidate):
        return candidate.offer_negotiations.first()

    def get_department(candidate):
        details = getattr(candidate.Req_id_fk, "position_information", None)
        return getattr(details, "department", "N/A") if details else "N/A"

    def get_recruiter_name(candidate):
        return candidate.Req_id_fk.Recruiter if candidate.Req_id_fk else "N/A"

    def get_location(candidate):
        planning = getattr(candidate.Req_id_fk, "Planning_id", None)
        return getattr(planning, "location", "N/A") if planning else "N/A"

    def get_offer_date(candidate):
        offer = candidate.generated_offers.first()  # Safely fetch first related offer
        return offer.created_at.strftime("%Y-%m-%d") if offer and offer.created_at else None



    def get_offered_salary(candidate):
        offer = get_offer_negotiation(candidate)
        return str(offer.offered_salary) if offer and offer.offered_salary else None



    report_data = []
    for fb in feedbacks:
        candidate = fb.candidate
        requisition = candidate.Req_id_fk

        report_data.append({
            "Client Name": requisition.company_client_name if requisition else "N/A",
            "Candidate Name": candidate.candidate_first_name + " " + candidate.candidate_last_name,
            "Position Offered": requisition.PositionTitle if requisition else "N/A",
            "Department": get_department(candidate),
            "Recruiter Name": get_recruiter_name(candidate),
            "Offer Date": get_offer_date(candidate),
            "Decline Date": fb.created_at.strftime("%Y-%m-%d") if fb.created_at else None,
            "Assessment Score / Rating": fb.assessment_score,
            "Interviewer Feedback": fb.interviewer_feedback,
            "Reason for Decline": fb.reason_not_selected,
            "Current Employer": fb.current_employer,
            "Location": get_location(candidate),
            "CTC Offered": get_offered_salary(candidate),
        })

    return Response(
        api_json_response_format(True, "Filtered feedback retrieved", 200, report_data),status=200)


@api_view(['GET'])
def get_yet_to_join_offer_report(request):
    today = timezone.now().date()

    offers = GeneratedOffer.objects.filter(
        estimated_start_date__gt=today
    ).select_related(
        'candidate',
        'requisition',
        'requisition__Planning_id'
    ).prefetch_related(
        'candidate__offer_negotiations'
    )

    def get_department(requisition):
        details = getattr(requisition, "position_information", None)
        return getattr(details, "department", "N/A") if details else "N/A"

    def get_location(requisition):
        planning = getattr(requisition, "Planning_id", None)
        return getattr(planning, "location", "N/A") if planning else "N/A"

    def get_offer_negotiation(candidate):
        return candidate.offer_negotiations.first()

    def get_offered_salary(candidate):
        offer = get_offer_negotiation(candidate)
        return str(offer.offered_salary) if offer and offer.offered_salary else None

    report_data = []
    for offer in offers:
        candidate = offer.candidate
        requisition = offer.requisition

        # 🔍 Interview Review
        review = InterviewReview.objects.filter(
            schedule__candidate=candidate,
            candidate=candidate
        ).order_by('-created_at').first()

        assessment_score = str(review.ActualRating) if review and review.ActualRating is not None else "N/A"
        interviewer_feedback = review.Feedback_param if review and review.Feedback_param else "N/A"
        recruiter_comments = candidate.Result if candidate.Result else "N/A"

        report_data.append({
            "Client Name": requisition.company_client_name if requisition else "N/A",
            "Candidate Name": f"{candidate.candidate_first_name} {candidate.candidate_last_name}",
            "Position Offered": requisition.PositionTitle if requisition else "N/A",
            "Department": get_department(requisition),
            "Recruiter Name": requisition.Recruiter if requisition else "N/A",
            "Offer Date": offer.created_at.strftime("%Y-%m-%d") if offer.created_at else None,
            "Expected Date Of Joining": offer.estimated_start_date.strftime("%d-%m-%Y") if offer.estimated_start_date else None,
            "Current Employer": candidate.current_employer if hasattr(candidate, 'current_employer') else "N/A",
            "Location": get_location(requisition),
            "CTC Offered": get_offered_salary(candidate),
            "Status": "Yet to join",
            "Assessment Score / Rating": assessment_score,
            "Interviewer Feedback": interviewer_feedback,
            "Recruiter Comments": recruiter_comments,
        })

    return Response(
        api_json_response_format(True, "Yet-to-join offers retrieved", 200, report_data),
        status=200
    )


@api_view(['POST'])
def get_candidate_offer_report(request):
    data = request.data
    client_name = data.get("client_name")
    min_salary = data.get("min_salary")
    max_salary = data.get("max_salary")
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    location = data.get("location")

    # Base queryset: candidates with at least one offer
    candidates_with_offer = Candidate.objects.filter(
        offer_negotiations__isnull=False
    ).select_related(
        "Req_id_fk",
        "Req_id_fk__position_information"
    ).prefetch_related("offer_negotiations").distinct()

    # Apply filters on related OfferNegotiation fields
    if client_name:
        candidates_with_offer = candidates_with_offer.filter(
            offer_negotiations__client_name__icontains=client_name
        )

    if location:
        candidates_with_offer = candidates_with_offer.filter(
            offer_negotiations__offered_location__icontains=location
        )

    if min_salary:
        candidates_with_offer = candidates_with_offer.filter(
            offer_negotiations__offered_salary__gte=min_salary
        )

    if max_salary:
        candidates_with_offer = candidates_with_offer.filter(
            offer_negotiations__offered_salary__lte=max_salary
        )

    if start_date and end_date:
        candidates_with_offer = candidates_with_offer.filter(
            offer_negotiations__offered_doj__range=[start_date, end_date]
        )
    elif start_date:
        candidates_with_offer = candidates_with_offer.filter(
            offer_negotiations__offered_doj__gte=start_date
        )
    elif end_date:
        candidates_with_offer = candidates_with_offer.filter(
            offer_negotiations__offered_doj__lte=end_date
        )

    serializer = CandidateOfferReportSerializer(candidates_with_offer, many=True)
    return Response(api_json_response_format(True, "Filtered offer-generated candidates retrieved", 200, serializer.data), status=200)


@api_view(['GET'])
def export_candidate_feedback_excel(request):
    feedbacks = CandidateFeedback.objects.select_related(
        "candidate",
        "candidate__Req_id_fk",
        "candidate__Req_id_fk__position_information",
        "candidate__Req_id_fk__HiringManager"
    ).order_by("-created_at")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidate Feedback"

    headers = [
        "Client ID", "Client Name", "Candidate First Name", "Candidate Last Name", "Position Considered For",
        "Hiring Manager", "Recruiter Name", "Interview Date(s)", "Assessment Score / Rating",
        "Interviewer Feedback Summary", "Recruiter Feedback Summary", "Reason Not Selected/Joined",
        "Skills / Expertise", "Current Employer", "Current Location", "Last CTC", "Status",
        "Follow-Up Date", "Notes"
    ]
    ws.append(headers)

    for fb in feedbacks:
        c = fb.candidate
        req = getattr(c, "Req_id_fk", None)
        position_info = getattr(req, "position_information", None)
        manager = getattr(req, "HiringManager", None)

        row = [
            getattr(req, "client_id", "N/A"),
            getattr(req, "company_client_name", "N/A"),
            getattr(c, "candidate_first_name", "N/A"),
            getattr(c, "candidate_last_name", "N/A"),
            getattr(position_info, "job_position", "N/A") if position_info else "N/A",
            getattr(manager, "Name", "N/A") if manager else "N/A",
            fb.recruiter_name,
            fb.interview_date.strftime("%Y-%m-%d") if fb.interview_date else "",
            fb.assessment_score,
            fb.interviewer_feedback,
            fb.recruiter_feedback,
            fb.reason_not_selected,
            fb.skills,
            fb.current_employer,
            fb.current_location,
            fb.last_ctc,
            fb.status,
            fb.follow_up_date.strftime("%Y-%m-%d") if fb.follow_up_date else "",
            fb.notes
        ]
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="candidate_feedback.xlsx"'
    wb.save(response)
    return response
class ScheduleCandidateRecruiterMeetView(APIView):
    def post(self, request):
        try:
            payload = request.data
            candidate_email = payload.get("candidate_email")
            recruiter_email = payload.get("recruiter_email")
            round_name = payload.get("purpose", "Interview")
            summary = payload.get("summary", f"{round_name} Interview")

            slot = payload.get("schedule_slots", [{}])[0]
            date = slot.get("date")
            time = slot.get("time")
            duration_raw = payload.get("durations", "30 mins")
            duration_mins = int(duration_raw.split()[0]) if duration_raw else 30

            if not all([candidate_email, recruiter_email, date, time]):
                return Response(api_json_response_format(False, "Missing required fields", 400, {}), status=200)

            tz = pytz.timezone("Asia/Kolkata")
            start = tz.localize(datetime.strptime(f"{date} {time}", "%Y-%m-%d %I:%M %p"))
            end = start + timedelta(minutes=duration_mins)

            # Meeting Link Generation
            join_url = schedule_zoom_meet(
                topic=summary,
                start_time_iso=start.isoformat(),
                duration_minutes=duration_mins,
                timezone=tz.zone
            )

            # Compose mail content
            subject = f"{summary} - Scheduled on {start.strftime('%d %b %Y at %I:%M %p')}"
            message = (
                f"Dear Participant,\n\n"
                f"You have a scheduled meeting:\n"
                f"Topic: {summary}\n"
                f"Date: {start.strftime('%Y-%m-%d')}\n"
                f"Time: {start.strftime('%I:%M %p')} IST\n"
                f"Duration: {duration_mins} minutes\n"
                f"Join Link: {join_url}\n\n"
                f"Please be ready at the scheduled time.\n\nBest regards."
            )

            recipients = [candidate_email]

            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[candidate_email],
                cc=[recruiter_email]  # Recruiter added in CC
            )

            email.send(fail_silently=True)

            return Response(api_json_response_format(True, "Meeting link sent via email", 200, {
                "meet_link": join_url
            }), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Failed to send email: {str(e)}", 500, {}), status=200)

@api_view(["POST"])
def get_interview_schedule_by_id(request):
    try:
        schedule_id = request.data.get("schedule_id")
        if not schedule_id:
            return Response(api_json_response_format(False, "Missing 'schedule_id' in query params.", 400, {}), status=200)

        schedule = InterviewSchedule.objects.select_related(
            "interviewer__req_id", "candidate__Req_id_fk__Planning_id"
        ).get(id=schedule_id)

        candidate_full_name = f"{schedule.candidate.candidate_first_name} {schedule.candidate.candidate_last_name}" if schedule.candidate else "N/A"
        interviewer_full_name = f"{schedule.interviewer.first_name} {schedule.interviewer.last_name}" if schedule.interviewer else "N/A"

        result = {
            "schedule_id": schedule.id,
            "req_id": schedule.interviewer.req_id.RequisitionID if schedule.interviewer and schedule.interviewer.req_id else "Not Linked",
            "planning_id": (
                schedule.candidate.Req_id_fk.Planning_id.hiring_plan_id
                if schedule.candidate and schedule.candidate.Req_id_fk and schedule.candidate.Req_id_fk.Planning_id
                else "N/A"
            ),
            "candidate_name": candidate_full_name,
            "interviewer_name": interviewer_full_name,
            "job_position": schedule.candidate.Req_id_fk.PositionTitle if schedule.candidate and schedule.candidate.Req_id_fk else "Unknown",
            "interviewers_id": schedule.interviewer.interviewer_id,
            "location": schedule.location,
            "time_zone": schedule.time_zone,
            "durations": schedule.durations,
            "purpose": schedule.purpose,
            "mode": schedule.mode,
            "no_of_rounds": 1,  # Static for now; customize if needed
            "meet_link": schedule.meet_link,
            "guests": schedule.guests or [],
            "schedule_slots": [{
                "date": str(schedule.date),
                "time": schedule.start_time.strftime("%I:%M %p"),
                "guests": [guest.get("email") for guest in schedule.guests] if schedule.guests else []
            }]
        }

        return Response(api_json_response_format(True, "Schedule retrieved successfully!", 200, result), status=200)

    except InterviewSchedule.DoesNotExist:
        return Response(api_json_response_format(False, "Schedule not found.", 404, {}), status=200)
    except Exception as e:
        return Response(api_json_response_format(False, f"Error retrieving schedule: {str(e)}", 500, {}), status=200)



@api_view(["PUT"])
def update_interview_schedule(request):
    try:
        schedule_id = request.data.get("schedule_id")
        if not schedule_id:
            return Response(api_json_response_format(False, "Missing 'schedule_id' in request.", 400, {}), status=200)

        schedule = InterviewSchedule.objects.get(id=schedule_id)

        # Update fields (only if provided)
        updatable_fields = [
            "location", "time_zone", "purpose", "mode", "guests", "durations",
            "round_name", "meet_link"
        ]
        for field in updatable_fields:
            if field in request.data:
                setattr(schedule, field, request.data[field])

        # Update schedule slot timing
        slot = request.data.get("schedule_slots", [{}])[0]
        if slot.get("date") and slot.get("time"):
            tz = pytz.timezone("Asia/Kolkata")
            start = tz.localize(datetime.strptime(f"{slot['date']} {slot['time']}", "%Y-%m-%d %I:%M %p"))
            duration_mins = int(schedule.durations.split()[0]) if schedule.durations else 30
            end = start + timedelta(minutes=duration_mins)

            schedule.date = start.date()
            schedule.start_time = start.time()
            schedule.end_time = end.time()

        schedule.save()

        return Response(api_json_response_format(True, "Interview schedule updated successfully!", 200, {
            "interviewers_id": schedule.interviewer.interviewer_id,"schedule_id": schedule.id
        }), status=200)

    except InterviewSchedule.DoesNotExist:
        return Response(api_json_response_format(False, "Interview schedule not found.", 404, {}), status=200)
    except Exception as e:
        return Response(api_json_response_format(False, f"Error updating interview: {str(e)}", 500, {}), status=200)

@api_view(["DELETE"])
def delete_interview_schedule(request):
    try:
        schedule_id = request.data.get("schedule_id")
        if not schedule_id:
            return Response(api_json_response_format(False, "Missing 'schedule_id' in request.", 400, {}), status=200)

        schedule = InterviewSchedule.objects.select_related("candidate", "interviewer").get(id=schedule_id)

        # Delete related Interviewer slot(s)
        # InterviewSlot.objects.filter(interviewer=schedule.interviewer).delete()

        # Delete related CandidateInterviewStage entry for this round
        CandidateInterviewStages.objects.filter(
            candidate_id=schedule.candidate.CandidateID,
            Req_id_id=schedule.candidate.Req_id_fk.RequisitionID,
            interview_stage=schedule.round_name
        ).delete()

        # Delete the actual schedule
        schedule.delete()

        return Response(api_json_response_format(
            True,
            "Interview schedule, and candidate stage deleted successfully!",
            200,
            {}
        ), status=200)

    except InterviewSchedule.DoesNotExist:
        return Response(api_json_response_format(False, "Schedule not found.", 404, {}), status=200)
    except Exception as e:
        return Response(api_json_response_format(False, f"Error deleting interview: {str(e)}", 500, {}), status=200)


RESUME_STORAGE_FOLDER = "media/resumes"
DISPLAY_TO_MODEL_FIELD1 = {
    "job_position": "job_position",
    "Tech": "tech_stacks",
    "JD": "jd_details",
    "Experience": "experience_range",
    "Designation": "designation",
    "Target": "target_companies",
    "Compensation/Benefits": "compensation",
    "Working": "working_model",
    "Interview Status": "interview_status",
    "Place": "location",
    "Educational": "education_decision",
    "Relocation": "relocation",
    "Travel": "travel_opportunities",
    "Visa": "visa_requirements",
    "Domain": "domain_knowledge",
    "Background": "background_verification",
    "Shift": "shift_timings",
    "Role Type": "role_type",
    "Job Type": "job_type",
    "Communication": "communication_language",
    "Notice Period": "notice_period",
    "Career Gap": "career_gap",
    "Sabbatical": "sabbatical",
    "Screening Questions": "screening_questions",
    "Job Health Requirements": "job_health_requirements",
    "Social Media": "social_media_links",
    "Language Proficiency": "language_proficiency",
    "Additional Compensation": "additional_comp",
    "Citizenship Requirement": "citizen_requirement"
}
DISPLAY_TO_MODEL_FIELD = {
    "id": "RequisitionID",
    "Planning_id": "Planning_id__hiring_plan_id",
    "job_position": "position_information__job_position",
    "Tech": "Planning_id__tech_stacks",
    "JD": "Planning_id__jd_details",
    "Experience": "Planning_id__experience_range",
    "Designation": "Planning_id__designation",
    "Target": "Planning_id__target_companies",
    "Interviewer": "position_information__interviewer_teammate_employee_id",
    "Interview": "position_information__client_interview",
    "Compensation/Benefits": "Planning_id__compensation",
    "Duration/Timeline": ["position_information__contract_start_date", "position_information__contract_end_date"],
    "Place": "position_information__location",
    "Working": "position_information__working_model",
    "Educational": "Planning_id__education_decision",
    "Relocation": "Planning_id__relocation",
    "Travel": "Planning_id__travel_opportunities",
    "Visa": "Planning_id__visa_requirements",
    "Domain": "Planning_id__domain_knowledge",
    "Background": "Planning_id__background_verification",
    "Shift": "Planning_id__shift_timings",
    "Role Type": "Planning_id__role_type",
    "Job Type": "Planning_id__job_type",
    "Communication": "Planning_id__communication_language",
    "Notice Period": "Planning_id__notice_period",
    "Career Gap": "Planning_id__career_gap",
    "Sabbatical": "Planning_id__sabbatical",
    "Screening Questions": "Planning_id__screening_questions",
    "Job Health Requirements": "Planning_id__job_health_requirements",
    "Social Media": "Planning_id__social_media_links",
    "Language Proficiency": "Planning_id__language_proficiency",
    "Recruiter": "Recruiter",
    "division": "position_information__division",
    "department": "position_information__department",
    "location": "position_information__location",
    "status": "Status",
    "Client_name": "company_client_name",
    "Client_id": "client_id",
    "Requisition_date": "requisition_date",
    "Due_requisition_date": "due_requisition_date",
    "No_of_positions": "No_of_positions"

}

# Initialize Ollama model
# ollama_model = OllamaLLM(base_url='http://localhost:11434', model='ats_model')
ollama_model = OllamaLLM(base_url='http://ollama:11434', model='ats_model')
# model = SentenceTransformer("models/paraphrase-MiniLM-L6-v2")

def extract_text_from_pdf(uploaded_file):
    reader = pdf.PdfReader(uploaded_file)
    text = ""
    for page in reader.pages:
        text += str(page.extract_text())
    return text


def extract_info_from_text(text):
    """Attempts AI-based name & email extraction, with fallback regex."""
    prompt = f"""
    Extract only the full name and email address from this resume.
    Return data in JSON format: 
    {{
        "name": "[Full Name]",
        "email": "[Email Address]"
    }}
    Resume Content: {text}"""

    try:
        ai_output = ollama_model.invoke(prompt)
    except Exception as e:
        print("Ollama Error:", e)
        ai_output = ""

    except Exception as e:
        print("Ollama Error:", e)
        ai_output = ""

    print("AI Response:\n", ai_output)

    # Try extracting using AI response JSON
    name_match = re.search(r'"name":\s*"([^"]+)"', ai_output)
    email_match = re.search(r'"email":\s*"([^"]+)"', ai_output)

    if name_match and email_match:
        return name_match.group(1), email_match.group(1)

    # Fallback regex extraction if AI fails
    name_fallback = re.search(r"Name:\s*(.*)", text)
    email_fallback = re.search(r"Email:\s*([\w\.-]+@[\w\.-]+)", text)

    name = name_fallback.group(1) if name_fallback else "Unknown"
    email = email_fallback.group(1) if email_fallback else "Not found"

    return name, email

def process_resume(file_req_tuple):
    file, req_id = file_req_tuple

    # Ensure the resume directory exists
    os.makedirs(RESUME_STORAGE_FOLDER, exist_ok=True)

    # Save file locally
    file_path = os.path.join(RESUME_STORAGE_FOLDER, file.name)
    with open(file_path, "wb") as f:
        f.write(file.read())

    # Extract text and info
    resume_text = extract_text_from_pdf(file)
    name, email = extract_info_from_text(resume_text)

    # Build Candidate instance
    return Candidate(
        Name=name,
        Email=email,
        Resume=file_path,
        Req_id_fk=req_id
    )


import os
from django.core.files.base import ContentFile

class ResumeAccessView(APIView):
    def post(self, request):
        try:
            candidate_id = request.data.get("candidate_id")
            if not candidate_id:
                return Response(api_json_response_format(False, "candidate_id is required.", 400, {}), status=200)

            candidate = Candidate.objects.select_related("Req_id_fk").filter(CandidateID=candidate_id).first()
            if not candidate:
                return Response(api_json_response_format(False, "Candidate not found.", 404, {}), status=200)

            # Resume
            resume_path = candidate.Resume
            resume_url = request.build_absolute_uri(settings.MEDIA_URL + "resumes/" + str(resume_path)) if resume_path else "N/A"

            # Cover Letter
            cover_path = candidate.CoverLetter
            cover_url = request.build_absolute_uri(settings.MEDIA_URL + "resumes/" + str(cover_path)) if cover_path else "N/A"

            # JD as HTML file
            jd_url = "N/A"
            if candidate.Req_id_fk and hasattr(candidate.Req_id_fk, "posting_details"):
                jd_html = candidate.Req_id_fk.posting_details.internal_job_description
                if jd_html:
                    # Create JD file dynamically
                    jd_filename = f"jd_candidate_{candidate_id}.html"
                    jd_dir = os.path.join(settings.MEDIA_ROOT, "jd_descriptions")
                    os.makedirs(jd_dir, exist_ok=True)
                    jd_path = os.path.join(jd_dir, jd_filename)

                    with open(jd_path, "w", encoding="utf-8") as jd_file:
                        jd_file.write(jd_html)

                    jd_url = request.build_absolute_uri(settings.MEDIA_URL + "jd_descriptions/" + jd_filename)

            return Response(api_json_response_format(True, "Resume and related details fetched successfully.", 200, {
                "resume_url": resume_url,
                "cover_letter_url": cover_url,
                "JD_url": jd_url
            }), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error fetching resume details. {str(e)}", 500, {}), status=200)


class CandidateDeleteView(APIView):
    def delete(self, request):
        try:
            candidate_id = request.data.get("candidate_id")

            if not candidate_id:
                return Response(api_json_response_format(
                    False, "Missing candidate_id in request body", 400, {}
                ), status=200)

            candidate = Candidate.objects.filter(CandidateID=candidate_id).first()
            if not candidate:
                return Response(api_json_response_format(
                    False, "Candidate not found", 404, {}
                ), status=200)

            candidate.delete()
            return Response(api_json_response_format(
                True, "Candidate deleted successfully", 200, {}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error deleting candidate: {str(e)}", 500, {}
            ), status=200)


class CandidateUpdateView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def put(self, request):
        try:
            candidate_id = request.data.get("candidate_id")
            if not candidate_id:
                return Response(api_json_response_format(False, "candidate_id is required.", 400, {}), status=200)

            candidate = Candidate.objects.filter(CandidateID=candidate_id).first()
            if not candidate:
                return Response(api_json_response_format(False, "Candidate not found.", 404, {}), status=200)

            # 🎯 Targeted normalization
            normalized_data = request.data.copy()
            rename_map = {
                "Candidate_First_Name": "candidate_first_name",
                "Candidate_Last_Name": "candidate_last_name",
                "source": "Source",
                "Resume": "resume",
                "CoverLetter": "CoverLetter"
            }
            for old_key, new_key in rename_map.items():
                if old_key in request.data:
                    normalized_data[new_key] = request.data[old_key]

            # 🔁 Update basic fields
            updatable_fields = ["candidate_first_name","candidate_last_name", "Email", "Feedback", "Final_rating", "Result", "source", "Score", "Phone_no"]
            for field in updatable_fields:
                if field in normalized_data:
                    setattr(candidate, field, normalized_data[field] or None)
            candidate.Source = request.data.get("source", candidate.Source)
            field_map = {
                "resume": "Resume",
                "CoverLetter": "CoverLetter"
            }

            # 📁 File uploads
            for field_key, model_field in field_map.items():
                upload_file = request.FILES.get(field_key)
                if upload_file:
                    # Save the file manually
                    folder_path = os.path.join(settings.MEDIA_ROOT, "resumes")
                    os.makedirs(folder_path, exist_ok=True)

                    filename = upload_file.name
                    full_path = os.path.join(folder_path, filename)

                    with default_storage.open(full_path, 'wb+') as destination:
                        for chunk in upload_file.chunks():
                            destination.write(chunk)

                    setattr(candidate, model_field, filename)


            candidate.save()

            return Response(api_json_response_format(True, "Candidate updated successfully with files.", 200, {}), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error during candidate update. {str(e)}", 500, {}), status=200)

class CandidateExcelExportView(APIView):
    def get(self, request):
        try:
            candidates = Candidate.objects.select_related("Req_id_fk").all()

            if not candidates.exists():
                return Response(api_json_response_format(
                    False,
                    "No candidates to export.",
                    404,
                    {}
                ), status=200)

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Candidate Details"

            headers = [
                "Candidate ID", "Candidate First Name", "Candidate Last Name", "Requisition ID", "Applied Position",
                "Job Description URL", "Resume URL", "Cover Letter URL", "Current Stage",
                "Next Stage", "Time in Stage", "Overall Stage", "Final Stage",
                "Source", "Score", "Phone Number"
            ]
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True)

            for candidate in candidates:
                jd = candidate.Req_id_fk.posting_details.internal_job_description if getattr(candidate.Req_id_fk, "posting_details", None) else None
                jd_url = request.build_absolute_uri(f"/api/candidates/{candidate.CandidateID}/jd/") if jd else "N/A"

                resume_url = request.build_absolute_uri(settings.MEDIA_URL + "resumes/" + quote(str(candidate.Resume))) if candidate.Resume else "N/A"
                cover_url = request.build_absolute_uri(settings.MEDIA_URL + "resumes/" + quote(str(candidate.CoverLetter))) if candidate.CoverLetter else "N/A"

                stage = CandidateInterviewStages.objects.filter(candidate_id=candidate.CandidateID).order_by('-interview_date').first()
                time_in_stage = f"{(datetime.now().date() - stage.interview_date).days} days" if stage and stage.interview_date else "N/A"
                current_stage = stage.interview_stage if stage else "N/A"

                applied_position = candidate.Req_id_fk.position_information.job_position if candidate.Req_id_fk and hasattr(candidate.Req_id_fk, "position_information") else "N/A"

                ws.append([
                    candidate.CandidateID,
                    candidate.candidate_first_name or "N/A",
                    candidate.candidate_last_name or "N/A",
                    candidate.Req_id_fk.RequisitionID if candidate.Req_id_fk else "N/A",
                    applied_position,
                    jd_url,
                    resume_url,
                    cover_url,
                    current_stage,
                    "N/A",  # Placeholder for Next Stage
                    time_in_stage,
                    candidate.Result or "N/A",
                    candidate.Final_rating if candidate.Final_rating is not None else "N/A",
                    candidate.Source or "N/A",
                    candidate.Score or "N/A",
                    candidate.Phone_no or "N/A"
                ])

            # Return as downloadable Excel file
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="candidate_details.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error exporting Excel. {str(e)}",
                500,
                {}
            ), status=200)

class CandidateInterviewStageView(APIView):
    def get(self, request):
        try:
            stages = CandidateInterviewStages.objects.select_related('candidate__Req_id_fk').all()
            if not stages.exists():
                return Response(api_json_response_format(False, "No interview stages available", 404, {}), status=200)

            enriched_data = []
            for stage in stages:
                candidate = Candidate.objects.select_related('Req_id_fk').filter(CandidateID=stage.candidate_id).first()
                requisition = candidate.Req_id_fk if candidate else None

                enriched_data.append({
                    "Req_ID": requisition.RequisitionID if requisition else "",
                    "Client": requisition.ClientName if requisition else "",
                    "Candidate_ID": candidate.CandidateID if candidate else "",
                    "Candidate_Name": candidate.CandidateName if candidate else "",
                    "Interview_Date": stage.interview_date,
                    "Last_Interview_Stage": stage.interview_stage,
                    "Current_Interview_Status": stage.status,
                    "Interview_Details": {
                        "Mode": stage.mode_of_interview,
                        "Feedback": stage.feedback,
                    }
                })

            return Response(api_json_response_format(True, "Candidate interview stage data retrieved", 200, enriched_data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error retrieving interview stage data: {str(e)}", 500, {}), status=200)

    def post(self, request):
        try:
            candidate_id = request.data.get("candidate_id")
            req_id = request.data.get("req_id")

            if not candidate_id:
                return Response(api_json_response_format(False, "candidate_id is required", 400, {}), status=200)

            filters = {"candidate_id": candidate_id}
            if req_id:
                filters["Req_id_id"] = req_id

            stages = CandidateInterviewStages.objects.filter(**filters)
            if not stages.exists():
                return Response(api_json_response_format(False, "No interview stages found for given criteria", 404, {}), status=200)

            serializer = CandidateInterviewStagesSerializer(stages, many=True)
            return Response(api_json_response_format(True, "Interview stages retrieved successfully", 200, serializer.data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error retrieving interview stages: {str(e)}", 500, {}), status=200)


class CandidateAllRequisitionsView(APIView):
    def get(self, request):
        try:
            candidates = Candidate.objects.select_related("Req_id_fk").order_by('-ProfileCreated')


            if not candidates.exists():
                return Response(api_json_response_format(False, "No candidate records found.", 404, {}), status=200)

            candidate_data = []
            for candidate in candidates:
                serializer = CandidateDetailWithInterviewSerializer(candidate)
                serialized_data = serializer.data

                # Optional cleanup: remove unused interview_stages field
                serialized_data.pop("interview_stages", None)

                candidate_data.append(serialized_data)

            return Response(api_json_response_format(True, "All candidate details retrieved successfully.", 200, candidate_data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error fetching all candidate data. {str(e)}", 500, {}), status=200)




class CandidateInterviewDetailView(APIView):
    def post(self, request):
        try:
            req_id = request.data.get("req_id")

            if not req_id:
                return Response(api_json_response_format(False, "req_id is required.", 400, {}), status=200)

            candidates = Candidate.objects.filter(Req_id_fk=req_id)

            if not candidates.exists():
                return Response(api_json_response_format(False, "No candidates found for the given req_id.", 404, {}), status=200)

            candidate_data = []
            for candidate in candidates:
                interview_exists = CandidateInterviewStages.objects.filter(candidate_id=candidate.CandidateID).exists()
                serializer = CandidateDetailWithInterviewSerializer(candidate)
                serialized_data = serializer.data

                # ✅ Convert resume filename to full URL
                resume_path = serialized_data.get("CV_Resume")
                resume_url = request.build_absolute_uri(settings.MEDIA_URL + "resumes/" + str(resume_path)) if resume_path else "N/A"
                serialized_data["CV_Resume"] = resume_url

                if not interview_exists:
                    serialized_data.pop("interview_stages", None)

                candidate_data.append(serialized_data)

            return Response(api_json_response_format(True, "Full candidate details retrieved successfully.", 200, candidate_data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error fetching candidate details. {str(e)}", 500, {}), status=200)
# class BulkUploadResumeView(APIView):
#     parser_classes = (MultiPartParser, FormParser)

#     def post(self, request, *args, **kwargs):
#         try:
#             req_id = request.data.get("req_id")
#             files = request.FILES.getlist("files")

#             if not files:
#                 return Response(api_json_response_format(
#                     False, "No files were uploaded.", 400, {}
#                 ), status=200)

#             try:
#                 job_req = JobRequisition.objects.get(RequisitionID=req_id)
#             except JobRequisition.DoesNotExist:
#                 return Response(api_json_response_format(
#                     False, "Invalid requisition ID.", 400, {}
#                 ), status=200)

#             file_req_pairs = [(file, job_req) for file in files]
#             num_workers = min(len(files), os.cpu_count())

#             with ThreadPoolExecutor(max_workers=num_workers) as executor:
#                 candidates = list(executor.map(process_resume, file_req_pairs))

#             Candidate.objects.bulk_create(candidates, batch_size=100)

#             return Response(api_json_response_format(
#                 True,
#                 f"{len(candidates)} resumes processed and candidates stored successfully.",
#                 200,
#                 {"processed_count": len(candidates)}
#             ), status=200)

#         except Exception as e:
#             return Response(api_json_response_format(
#                 False,
#                 "Failed to process bulk resume upload. " + str(e),
#                 500,
#                 {}
#             ), status=200)


class BulkUploadResumeView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        try:
            req_id = request.data.get("req_id")
            files = request.FILES.getlist("files")
            Source = request.data.get("Source")

            if not req_id:
                return Response(api_json_response_format(
                    False, "Requisition ID is required.", 400, {}
                ), status=200)

            if not files:
                return Response(api_json_response_format(
                    False, "No files uploaded.", 400, {}
                ), status=200)

            try:
                job_req = JobRequisition.objects.get(RequisitionID=req_id)
            except JobRequisition.DoesNotExist:
                return Response(api_json_response_format(
                    False, "Invalid Requisition ID.", 400, {}
                ), status=200)

            # ✅ Generate randomized candidate data
            candidates = []
            SOURCE_OPTIONS = ["LinkedIn", "Referral", "Job Board", "Walk-In", "Campus Drive", "Employee Network"]

            for i, file in enumerate(files, start=1):
                unique_id = random.randint(1000, 9999)
                random_source = random.choice(SOURCE_OPTIONS)
                first_name = f"CandidateFirst{unique_id}"
                last_name = f"CandidateLast{unique_id}"
                filename = f"{file.name}"
                folder_path = os.path.join(settings.MEDIA_ROOT, "resumes")
                os.makedirs(folder_path, exist_ok=True)

                full_path = os.path.join(folder_path, filename)
                with default_storage.open(full_path, 'wb+') as destination:
                    for chunk in file.chunks():
                        destination.write(chunk)


                candidate = Candidate(
                    candidate_first_name=first_name,
                    candidate_last_name=last_name,
                    Email=f"candidate{unique_id}@gmail.com",
                    Req_id_fk=job_req,
                    Resume=file.name,
                    CoverLetter=f"This is a sample cover letter for candidate{unique_id}.",
                    Phone_no = "9999999999",
                    Source=Source  # ✅ Injected random source
                )
                candidates.append(candidate)


            with transaction.atomic():
                Candidate.objects.bulk_create(candidates, batch_size=100)

            return Response(api_json_response_format(
                True,
                f"{len(candidates)} resumes uploaded and candidates stored successfully.",
                200,
                {"uploaded_count": len(candidates)}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Failed to upload resumes. {str(e)}",
                500,
                {}
            ), status=200)


# def extract_info_from_text(text):
#     """Attempts AI-based name & email extraction, with fallback regex."""
#     prompt = f"""
#     Extract only the full name and email address from this resume.
#     Return data in JSON format: 
#     {{
#         "name": "[Full Name]",
#         "email": "[Email Address]"
#     }}
#     Resume Content: {text}
#     """
#     # ollama_model = OllamaLLM(base_url='http://ollama:11434', model='ats_model')
#     # response = ollama.chat(model='ats_model', messages=[{"role": "user", "content": prompt}])
#     # ai_output = response['message']['content']
#     # ai_output = ollama_model.invoke(prompt)
#     # return response.strip()
#     # Print AI response for debugging
#     try:
#         ai_output = ollama_model.invoke(prompt)
#     except Exception as e:
#         print("Ollama Error:", e)

#     print("AI Response:\n", ai_output)

#     # Try extracting using AI response JSON
#     name_match = re.search(r'"name":\s*"([^"]+)"', ai_output)
#     email_match = re.search(r'"email":\s*"([^"]+)"', ai_output)

#     if name_match and email_match:
#         return name_match.group(1), email_match.group(1)

#     # Fallback regex extraction if AI fails
#     name_fallback = re.search(r"Name:\s*(.*)", text)
#     email_fallback = re.search(r"Email:\s*([\w\.-]+@[\w\.-]+)", text)

#     name = name_fallback.group(1) if name_fallback else "Unknown"
#     email = email_fallback.group(1) if email_fallback else "Not found"

#     return name, email

# def process_resume(file_req_tuple):
#     file, req_id = file_req_tuple

#     os.makedirs(RESUME_STORAGE_FOLDER, exist_ok=True)

#     file_path = os.path.join(RESUME_STORAGE_FOLDER, file.name)
#     with open(file_path, "wb") as f:
#         f.write(file.read())

#     text = extract_text_from_pdf(file)
#     name, email = extract_info_from_text(text)

#     return Candidate(
#         Name=name,
#         Email=email,
#         Resume=file_path,
#         Req_id_fk =req_id  # Set FK using ID to avoid query overhead
#     )


# class BulkUploadResumeView(APIView):
#     parser_classes = (MultiPartParser, FormParser)

#     def post(self, request, *args, **kwargs):
#         try:
#             req_id = request.data.get("req_id")
#             files = request.FILES.getlist("files")

#             if not files:
#                 return Response(api_json_response_format(
#                     False,
#                     "No files were uploaded.",
#                     400,
#                     {}
#                 ), status=200)
#             try:
#                 job_req = JobRequisition.objects.get(pk=req_id)
#                 print(job_req)
#             except JobRequisition.DoesNotExist:
#                 return Response(api_json_response_format(
#                     False, "Invalid requisition ID.", 400, {}
#                 ), status=200)

#             file_req_pairs = [(file, job_req) for file in files]
#             num_workers = min(len(files), os.cpu_count())

#             with ThreadPoolExecutor(max_workers=num_workers) as executor:
#                 candidates = list(executor.map(process_resume, file_req_pairs))

#             Candidate.objects.bulk_create(candidates, batch_size=100)

#             return Response(api_json_response_format(
#                 True,
#                 f"{len(candidates)} resumes processed and candidates stored successfully.",
#                 200,
#                 {"processed_count": len(candidates)}
#             ), status=200)

#         except Exception as e:
#             return Response(api_json_response_format(
#                 False,
#                 "Failed to process bulk resume upload. " + str(e),
#                 500,
#                 {}
            # ), status=200)

# model = SentenceTransformer("models/paraphrase-MiniLM-L6-v2")
# model.save("models/paraphrase-MiniLM-L6-v2")  # Save it locally
# print("Model saved to local folder.")


# def get_matching_score(job_description, resume_text, resume_name):
#     job_embedding = model.encode(job_description, convert_to_tensor=True)
#     resume_embedding = model.encode(resume_text[:2000], convert_to_tensor=True)
#     similarity = util.pytorch_cos_sim(job_embedding, resume_embedding).item()
#     percentage = round(similarity * 100)

#     return {
#         "resume_name": resume_name,
#         "percentage": percentage
#     }


# def extract_text_from_pdf(uploaded_file):
#     reader = pdf.PdfReader(uploaded_file)
#     text = ""
#     for page in reader.pages:
#         text += str(page.extract_text())
#     return text


# class ResumeMatchingAPI(APIView):
#     parser_classes = (MultiPartParser, FormParser)

#     def post(self, request):
#         try:
#             job_description = request.data.get('job_description')
#             uploaded_files = request.FILES.getlist('resumes')

#             if not job_description or not uploaded_files:
#                 return Response(api_json_response_format(
#                     False, "Missing job description or resume files.", 400, {}
#                 ), status=200)

#             results = []
#             for uploaded_file in uploaded_files:
#                 resume_text = extract_text_from_pdf(uploaded_file)
#                 score = get_matching_score(job_description, resume_text, uploaded_file.name)
#                 results.append(score)

#             return Response(api_json_response_format(
#                 True, "Matching scores calculated successfully!", 200, {"matching_scores": results}
#             ), status=200)

#         except Exception as e:
#             return Response(api_json_response_format(
#                 False, "Error while processing resumes. " + str(e), 500, {}
#             ), status=200)

@api_view(['POST'])
def send_form_invite(request):
    candidate_id = request.data.get("candidate_id")
    candidate = get_object_or_404(Candidate, CandidateID=candidate_id)

    invite = CandidateFormInvite.objects.create(candidate=candidate)
    form_link = f"https://hiring.pixeladvant.com/candidate/personal_details_form?token={invite.token}"

    # Send the email
    send_mail(
        subject="Complete Your Personal Details",
        message=f"Hi {candidate.candidate_first_name+" "+candidate.candidate_last_name},\n\nPlease complete your personal details form using the secure link below. This link will expire on {invite.expires_at.strftime('%Y-%m-%d %H:%M')}.\n\n{form_link}\n\nThank you!",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[candidate.Email],
        fail_silently=True
    )

    # Serialize the invite object if needed
    serialized_invite = CandidateFormInviteSerializer(invite).data
    serialized_invite["link"] = form_link
    return Response(api_json_response_format(
        True,
        "Form invite sent successfully.",
        200,
        serialized_invite  # This will include link, expiry, etc.
    ), status=200)

@api_view(['POST'])
def send_pre_onboarding_form_invite(request):
    candidate_id = request.data.get("candidate_id")
    candidate = get_object_or_404(Candidate, CandidateID=candidate_id)

    invite = CandidateFormInvite.objects.create(candidate=candidate)
    form_link = f"https://hiring.pixeladvant.com/candidate/pre_onboarding_form?token={invite.token}"

    # Send the email
    send_mail(
        subject="Start Your Pre-Onboarding Process",
        message=(
            f"Hi {candidate.candidate_first_name} {candidate.candidate_last_name},\n\n"
            f"Please complete your pre-onboarding form using the secure link below. "
            f"This link will expire on {invite.expires_at.strftime('%Y-%m-%d %H:%M')}.\n\n"
            f"{form_link}\n\n"
            "If you have any questions, feel free to reach out.\n\n"
            "Best regards,\nThe Hiring Team"
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[candidate.Email],
        fail_silently=True
    )

    # Optional: return the link and invite metadata
    serialized_invite = CandidateFormInviteSerializer(invite).data
    serialized_invite["link"] = form_link
    return Response(api_json_response_format(
        True,
        "Pre-onboarding form invite sent successfully.",
        200,
        serialized_invite
    ), status=200)




class CandidateSubmissionViewSet(viewsets.ModelViewSet):
    queryset = CandidateSubmission.objects.all()
    serializer_class = CandidateSubmissionSerializer

    def create(self, request, *args, **kwargs):
        token = request.data.get("token")
        if not token:
            return Response(api_json_response_format(False, "Token missing.", 400, {}), status=200)

        try:
            invite = CandidateFormInvite.objects.get(token=token)
        except CandidateFormInvite.DoesNotExist:
            return Response(api_json_response_format(False, "Invalid token.", 404, {}), status=200)

        if invite.is_expired():
            return Response(api_json_response_format(False, "Token expired.", 403, {}), status=200)

        payload = request.data.copy()
        payload['candidate'] = invite.candidate.CandidateID

        serializer = CandidateSubmissionSerializer(data=payload)
        if serializer.is_valid():
            validated_data = serializer.validated_data

            # Handle nested data manually
            personal_data = validated_data.pop('personal_detail', {})
            education_data = validated_data.pop('education_details', [])
            employment_data = validated_data.pop('employment_details', [])
            reference_data = validated_data.pop('references', [])

            # Save submission
            submission = CandidateSubmission.objects.create(**validated_data)

            if personal_data:
                CandidatePersonal.objects.create(submission=submission, **personal_data)

            for edu in education_data:
                CandidateEducation.objects.create(submission=submission, **edu)

            for emp in employment_data:
                CandidateEmployment.objects.create(submission=submission, **emp)

            for ref in reference_data:
                CandidateReference.objects.create(candidate_submission=submission, **ref)

            return Response(api_json_response_format(
                True,
                "Form submitted successfully.",
                200,
                CandidateSubmissionSerializer(submission).data
            ), status=200)

        return Response(api_json_response_format(False, "Validation error.", 400, serializer.errors), status=200)



class CandidateDetailView(APIView):
    def post(self, request):
        try:
            candidate_id = request.data.get("candidate_id")
            if not candidate_id:
                return Response(api_json_response_format(False, "Missing candidate_id", 400, {}), status=200)

            candidate = Candidate.objects.select_related("Req_id_fk__HiringManager").get(pk=candidate_id)
            requisition = candidate.Req_id_fk

            # Serialize base candidate data
            serializer = CandidateSerializer(candidate)
            candidate_data = serializer.data

            # Enrich with HR name and position title
            candidate_data["position_title"] = requisition.PositionTitle if requisition else "Not Provided"
            candidate_data["hr_name"] = requisition.HiringManager.Name if requisition and requisition.HiringManager else "Unknown"
            
            # Format profile creation date
            raw_created = candidate.ProfileCreated
            candidate_data["ProfileCreated"] = raw_created.strftime("%Y-%m-%d") if raw_created else None

            # Add interview_design_id based on req_id
            interview_design = InterviewDesignScreen.objects.filter(req_id=requisition.RequisitionID).first()
            candidate_data["interview_design_id"] = interview_design.interview_design_id if interview_design else None

            return Response(api_json_response_format(
                True,
                "Candidate detail retrieved",
                200,
                candidate_data
            ), status=200)

        except Candidate.DoesNotExist:
            return Response(api_json_response_format(False, "Candidate not found", 404, {}), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, str(e), 500, {}), status=200)



class CandidateListByRequisitionView(APIView):
    def post(self, request):
        try:
            requisition_id = request.data.get("requisition_id")
            if not requisition_id:
                return Response(api_json_response_format(False, "Missing requisition_id", 400, {}), status=200)

            candidates = Candidate.objects.filter(Req_id_fk__RequisitionID=requisition_id).select_related("Req_id_fk")
            approved_candidates = []

            for candidate in candidates:
                if candidate.Result in [None, "", "Pending"]:
                    continue

                requisition = candidate.Req_id_fk
                approvals = CandidateApproval.objects.filter(candidate=candidate)
                approvers = Approver.objects.filter(requisition=requisition, set_as_approver="Yes")
                expected_roles = [a.role for a in approvers]

                overall_status = compute_overall_status1(approvals, expected_roles)
                if overall_status == "Approved":
                    approved_candidates.append({
                        "id": candidate.CandidateID,
                        "name": f"{candidate.candidate_first_name} {candidate.candidate_last_name}"
                    })

            return Response(api_json_response_format(True, "Approved candidates fetched successfully", 200, approved_candidates), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, str(e), 500, {}), status=200)



class InterviewerListByRequisitionView(APIView):
    def post(self, request):
        try:
            requisition_id = request.data.get("requisition_id")
            if not requisition_id:
                return Response(api_json_response_format(False, "Missing requisition_id", 400, {}), status=200)

            # Step 1: Get interview_design_id from InterviewDesignScreen
            design_screen = InterviewDesignScreen.objects.filter(req_id=requisition_id).first()
            interview_design_id = design_screen.interview_design_id if design_screen else None

            # Step 2: Build score_card → round_no map
            round_map = {}
            if interview_design_id:
                design_params = InterviewDesignParameters.objects.filter(interview_design_id=interview_design_id)
                round_map = {
                    param.score_card.strip(): param.round_no.strip()
                    for param in design_params
                    if param.score_card and param.round_no
                }

            # Step 3: Fetch all interviewers for the requisition
            candidates = Interviewer.objects.filter(req_id__RequisitionID=requisition_id)
            data = []
            for c in candidates:
                stage = c.interviewer_stage.strip() if c.interviewer_stage else ""
                round_no = round_map.get(stage, "")
                round_suffix = f" R-{round_no}" if round_no.isdigit() else ""
                full_name = f"{c.first_name} {c.last_name}{round_suffix}".strip()

                # No InterviewSchedule check — include all interviewers
                data.append({
                    "id": c.interviewer_id,
                    "name": full_name
                })

            return Response(api_json_response_format(True, "interviewer fetched successfully", 200, data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, str(e), 500, {}), status=200)



class CandidateScreening(APIView):
    def post(self, request):
        req_id = request.data.get("req_id")
        candidates = Candidate.objects.select_related("Req_id_fk__Planning_id")

        if req_id:
            candidates = candidates.filter(Req_id_fk__RequisitionID=req_id)

        response_data = []

        for candidate in candidates:
            reviews = CandidateReview.objects.filter(CandidateID=candidate).values(
                "ParameterDefined", "Guidelines", "MinimumQuestions", "ActualRating", "Feedback"
            )

            approvals = CandidateApproval.objects.filter(candidate=candidate).values(
                "approver__email", "role", "decision"
            )

            resume_url = (
                request.build_absolute_uri(settings.MEDIA_URL + "resumes/" + str(candidate.Resume))
                if candidate.Resume else "N/A"
            )
            cover_url = (
                request.build_absolute_uri(settings.MEDIA_URL + "resumes/" + str(candidate.CoverLetter))
                if candidate.CoverLetter else "N/A"
            )

            jd_url = "N/A"
            jd_html_content = "N/A"
            if candidate.Req_id_fk and hasattr(candidate.Req_id_fk, "posting_details"):
                jd_html = candidate.Req_id_fk.posting_details.internal_job_description
                if jd_html:
                    jd_filename = f"jd_candidate_{candidate.pk}.html"
                    jd_dir = os.path.join(settings.MEDIA_ROOT, "jd_descriptions")
                    os.makedirs(jd_dir, exist_ok=True)
                    jd_path = os.path.join(jd_dir, jd_filename)

                    with open(jd_path, "w", encoding="utf-8") as jd_file:
                        jd_file.write(jd_html)

                    jd_url = request.build_absolute_uri(settings.MEDIA_URL + "jd_descriptions/" + jd_filename)
                    jd_html_content = jd_html

            stage = CandidateInterviewStages.objects.filter(candidate_id=candidate.CandidateID).order_by('-interview_date').first()
            time_in_stage = f"{(datetime.now().date() - stage.interview_date).days} days" if stage and stage.interview_date else "N/A"
            current_stage = stage.interview_stage if stage else "N/A"

            applied_position = (
                candidate.Req_id_fk.position_information.job_position
                if candidate.Req_id_fk and hasattr(candidate.Req_id_fk, "position_information")
                else "N/A"
            )

            candidate_info = {
                "candidate_id": candidate.pk,
                "candidate_first_name": candidate.candidate_first_name or "N/A",
                "candidate_last_name": candidate.candidate_last_name or "N/A",
                "req_id": candidate.Req_id_fk.RequisitionID if candidate.Req_id_fk else None,
                "planning_id": candidate.Req_id_fk.Planning_id.hiring_plan_id if candidate.Req_id_fk and candidate.Req_id_fk.Planning_id else None,
                "client_name": candidate.Req_id_fk.company_client_name if candidate.Req_id_fk else None,
                "final_rating": candidate.Final_rating,
                "final_feedback": candidate.Feedback,
                "result": candidate.Result,
                "score": candidate.Score,
                "applied_position": applied_position,
                "current_stage": current_stage,
                "time_in_stage": time_in_stage,
                "reviews": list(reviews),
                "approvals": list(approvals),
                "resume_url": resume_url,
                "cover_letter_url": cover_url,
                "JD_url": jd_url,
                "JD": jd_html_content
            }

            response_data.append(candidate_info)

        return Response(api_json_response_format(
            True,
            "Candidate screening data retrieved successfully",
            200,
            response_data
        ), status=200)


class CandidateScreeningView(APIView):
    def get(self, request):
        candidates = Candidate.objects.select_related("Req_id_fk__Planning_id").all()
        response_data = []

        for candidate in candidates:
            reviews = CandidateReview.objects.filter(CandidateID=candidate).values(
                "ReviewID", "ParameterDefined", "Guidelines", "MinimumQuestions", "ActualRating", "Feedback"
            )


            approvals = CandidateApproval.objects.filter(candidate=candidate).values(
                "approver__email", "role", "decision"
            )

            # Resume URL
            resume_path = candidate.Resume
            resume_url = request.build_absolute_uri(settings.MEDIA_URL + "resumes/" + str(resume_path)) if resume_path else "N/A"

            # Cover Letter URL
            cover_path = candidate.CoverLetter
            cover_url = request.build_absolute_uri(settings.MEDIA_URL + "resumes/" + str(cover_path)) if cover_path else "N/A"

            # JD HTML URL (if available)
            jd_url = "N/A"
            jd_html_content = "N/A"
            if candidate.Req_id_fk and hasattr(candidate.Req_id_fk, "posting_details"):
                jd_html = candidate.Req_id_fk.posting_details.internal_job_description
                if jd_html:
                    jd_filename = f"jd_candidate_{candidate.pk}.html"
                    jd_dir = os.path.join(settings.MEDIA_ROOT, "jd_descriptions")
                    os.makedirs(jd_dir, exist_ok=True)
                    jd_path = os.path.join(jd_dir, jd_filename)

                    with open(jd_path, "w", encoding="utf-8") as jd_file:
                        jd_file.write(jd_html)

                    jd_url = request.build_absolute_uri(settings.MEDIA_URL + "jd_descriptions/" + jd_filename)
                    jd_html_content = jd_html
            
            stage = CandidateInterviewStages.objects.filter(candidate_id=candidate.CandidateID).order_by('-interview_date').first()
            time_in_stage = f"{(datetime.now().date() - stage.interview_date).days} days" if stage and stage.interview_date else "N/A"
            current_stage = stage.interview_stage if stage else "N/A"

            applied_position = candidate.Req_id_fk.position_information.job_position if candidate.Req_id_fk and hasattr(candidate.Req_id_fk, "position_information") else "N/A"

            candidate_info = {
                "candidate_id": candidate.pk,
                "req_id": candidate.Req_id_fk.RequisitionID if candidate.Req_id_fk else None,
                "planning_id": candidate.Req_id_fk.Planning_id.hiring_plan_id if candidate.Req_id_fk and candidate.Req_id_fk.Planning_id else None,
                "client_name": candidate.Req_id_fk.company_client_name if candidate.Req_id_fk else None,
                "final_rating": candidate.Final_rating,
                "final_feedback": candidate.Feedback,
                "result": candidate.Result,
                "score": candidate.Score,
                "applied_position": applied_position,
                "current_stage": current_stage,
                "time_in_stage": time_in_stage,
                "reviews": list(reviews),
                "approvals": list(approvals),
                "resume_url": resume_url,
                "cover_letter_url": cover_url,
                "JD_url": jd_url,
                "JD": jd_html_content

            }

            response_data.append(candidate_info)

        return Response(api_json_response_format(
            True,
            "All candidate screening data retrieved successfully",
            200,
            response_data
        ), status=200)


    def put(self, request):
        candidate_id = request.data.get("candidate_id")

        if not candidate_id:
            return Response(api_json_response_format(
                False, "Missing candidate_id in request body", 400, {}
            ), status=200)

        try:
            candidate = Candidate.objects.get(pk=candidate_id)
        except Candidate.DoesNotExist:
            return Response(api_json_response_format(
                False, "Candidate not found", 404, {}
            ), status=200)

        reviews_data = request.data.get("reviews", [])
        final_rating = request.data.get("final_rating")
        final_feedback = request.data.get("final_feedback")
        result = request.data.get("result")

        # Parse screening date
        screening_date_str = request.data.get("screeningDate")
        screening_date = None
        if screening_date_str:
            try:
                screening_date = datetime.strptime(screening_date_str, "%d/%m/%Y").date()
            except ValueError:
                return Response(api_json_response_format(
                    False, "Invalid screeningDate format. Expected DD/MM/YYYY.", 400, {}
                ), status=200)

        try:
            with transaction.atomic():
                for review in reviews_data:
                    review_id = review.get("id")

                    if review_id:
                        try:
                            existing_review = CandidateReview.objects.get(pk=review_id, CandidateID=candidate)
                            existing_review.ParameterDefined = review.get("parameterDefined", existing_review.ParameterDefined)
                            existing_review.Guidelines = review.get("Guidelines", existing_review.Guidelines)
                            existing_review.MinimumQuestions = review.get("Skills", existing_review.MinimumQuestions)
                            existing_review.ActualRating = review.get("ActualRating", existing_review.ActualRating)
                            existing_review.Feedback = review.get("Feedback", existing_review.Feedback)
                            existing_review.Weightage = review.get("weightage", existing_review.Weightage)
                            existing_review.ScreeningDate = screening_date
                            existing_review.save()
                        except CandidateReview.DoesNotExist:
                            continue
                    else:
                        CandidateReview.objects.create(
                            CandidateID=candidate,
                            ParameterDefined=review.get("parameterDefined"),
                            Guidelines=review.get("Guidelines"),
                            MinimumQuestions=review.get("Skills"),
                            ActualRating=review.get("ActualRating"),
                            Feedback=review.get("Feedback"),
                            Weightage=review.get("weightage"),
                            ScreeningDate=screening_date
                        )

                if final_rating is not None:
                    candidate.Final_rating = final_rating
                if final_feedback:
                    candidate.Feedback = final_feedback
                if result:
                    candidate.Result = result
                candidate.save()

            return Response(api_json_response_format(
                True, "Candidate screening data updated successfully", 200, {}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error updating screening data: {e}", 500, {}
            ), status=200)



    def post(self, request):
        candidate_id = request.data.get("candidate_id")

        if not candidate_id:
            return Response(api_json_response_format(
                False, "Missing candidate_id in request body", 400, {}
            ), status=200)

        try:
            candidate = Candidate.objects.select_related("Req_id_fk", "Req_id_fk__position_information").get(pk=candidate_id)
        except Candidate.DoesNotExist:
            return Response(api_json_response_format(
                False, "Candidate not found", 404, {}
            ), status=200)

        requisition_id = candidate.Req_id_fk.RequisitionID if candidate.Req_id_fk else None
        if not requisition_id:
            return Response(api_json_response_format(
                False, "Candidate is not linked to any requisition", 400, {}
            ), status=200)

        reviews_data = request.data.get("reviews", [])
        final_rating = request.data.get("final_rating")
        final_feedback = request.data.get("final_feedback")
        result = request.data.get("result")
        screening_date_str = request.data.get("screeningDate")
        screening_date = None
        if screening_date_str:
            try:
                screening_date = datetime.strptime(screening_date_str, "%d/%m/%Y").date()
            except ValueError:
                return Response(api_json_response_format(
                    False, "Invalid screeningDate format. Expected DD/MM/YYYY.", 400, {}
                ), status=200)



        review_instances = [
            CandidateReview(
                CandidateID=candidate,
                ParameterDefined=review.get("parameterDefined"),
                Guidelines=review.get("Guidelines"),
                MinimumQuestions=review.get("Skills"),
                Weightage=review.get("weightage"),
                ActualRating=review.get("ActualRating"),
                Feedback=review.get("Feedback"),
                ScreeningDate=screening_date

            )
            for review in reviews_data
        ]

        try:
            with transaction.atomic():
                CandidateReview.objects.bulk_create(review_instances)

                if final_rating is not None:
                    candidate.Final_rating = final_rating
                if final_feedback:
                    candidate.Feedback = final_feedback
                if result:
                    candidate.Result = result
                candidate.save()

                approvers = Approver.objects.filter(
                    requisition__RequisitionID=requisition_id,
                    set_as_approver__in=["Yes", "Maybe"]
                )

                approval_instances = [
                    CandidateApproval(
                        candidate=candidate,
                        approver=approver,
                        role=approver.role,
                        decision="Awaiting"
                    )
                    for approver in approvers
                ]
                CandidateApproval.objects.bulk_create(approval_instances)

                # 📧 Send HTML emails to approvers
                for approver in approvers:
                    subject = f"Candidate Screening Decision Requested - {requisition_id}"
                    approve_url = f"https://api.pixeladvant.com/api/approve-decision?candidate_id={candidate.pk}&approver_id={approver.id}&decision=Approved"
                    reject_url = f"https://api.pixeladvant.com/api/approve-decision?candidate_id={candidate.pk}&approver_id={approver.id}&decision=Reject"

                    html_message = f"""
                        <html>
                        <body style="font-family: Arial, sans-serif;">
                            <h2>Candidate Screening Decision Requested</h2>
                            <p>Dear {approver.role},</p>
                            <p>A candidate has completed screening for requisition ID <strong>{requisition_id}</strong>. Please review the details below and submit your decision.</p>

                            <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
                                <tr><td><strong>Candidate Name:</strong></td><td>{candidate.candidate_first_name + " " + candidate.candidate_last_name}</td></tr>
                                <tr><td><strong>Email:</strong></td><td>{candidate.Email}</td></tr>
                                <tr><td><strong>Applied Position:</strong></td><td>{candidate.Req_id_fk.position_information.job_position}</td></tr>
                                <tr><td><strong>Rating:</strong></td><td>{candidate.Final_rating}</td></tr>
                                <tr><td><strong>Feedback:</strong></td><td>{candidate.Feedback}</td></tr>
                            </table>

                            <p style="margin-top: 20px;">
                                <a href="{approve_url}" style="padding: 10px 20px; background-color: #4CAF50; color: white; text-decoration: none; border-radius: 5px;">✅ Approve</a>
                                &nbsp;&nbsp;
                                <a href="{reject_url}" style="padding: 10px 20px; background-color: #f44336; color: white; text-decoration: none; border-radius: 5px;">❌ Reject</a>
                            </p>

                            <p>Thank you,<br/>Hiring System</p>
                        </body>
                        </html>
                        """


                    send_mail(
                        subject=subject,
                        message="",  # plain-text fallback
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[approver.email],
                        html_message=html_message
                    )

            return Response(api_json_response_format(
                True, "Screening data submitted and emails sent successfully", 200, {}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error saving screening data: {e}", 500, {}
            ), status=200)




@api_view(["POST"])
def get_reviews_by_candidate(request):
    candidate_id = request.data.get("candidate_id")

    if not candidate_id:
        return Response(
            api_json_response_format(
                False,
                "Missing candidate_id in request body",
                400,
                {}
            ),
            status=200
        )

    try:
        candidate = Candidate.objects.get(pk=candidate_id)
    except Candidate.DoesNotExist as e:
        return Response(
            api_json_response_format(
                False,
                f"Candidate not found: {e}",
                404,
                {}
            ),
            status=200
        )
    except Exception as e:
        return Response(
            api_json_response_format(
                False,
                f"Error retrieving candidate: {e}",
                500,
                {}
            ),
            status=200
        )

    try:
        reviews = CandidateReview.objects.filter(CandidateID=candidate).values(
            id=F("ReviewID"),
            score_card=F("ParameterDefined"),
            guideline=F("Guidelines"),
            min_questions=F("MinimumQuestions"),
            actual_rating=F("ActualRating"),
            feedback=F("Feedback"),
            weightage=F("Weightage"),
        )

        formatted_reviews = []
        for review in reviews:
            review["skills"] = review["min_questions"]
            formatted_reviews.append(review)

        response_payload = {
            "candidate_summary": {
                "final_rating": candidate.Final_rating,
                "final_feedback": candidate.Feedback,
                "result": candidate.Result
            },
            "reviews": formatted_reviews
        }

        return Response(
            api_json_response_format(
                True,
                "Candidate reviews retrieved successfully.",
                200,
                response_payload
            ),
            status=200
        )

    except Exception as e:
        return Response(
            api_json_response_format(
                False,
                f"Error retrieving candidate reviews: {e}",
                500,
                {}
            ),
            status=200
        )


        
@api_view(['POST'])
def resend_approval_emails(request):
    try:
        candidate_id = request.data.get("candidate_id")
        candidate = Candidate.objects.select_related("Req_id_fk", "Req_id_fk__position_information").get(pk=candidate_id)
        requisition_id = candidate.Req_id_fk.RequisitionID if candidate.Req_id_fk else None

        if not requisition_id:
            return Response(api_json_response_format(False, "Candidate not linked to requisition", 400, {}), status=200)

        approvers = Approver.objects.filter(
            requisition__RequisitionID=requisition_id,
            set_as_approver__in=["Yes", "Maybe"]
        )

        for approver in approvers:
            subject = f"🔁 Resend: Candidate Screening Decision Requested - {requisition_id}"
            approve_url = f"https://api.pixeladvant.com/api/approve-decision?candidate_id={candidate.pk}&approver_id={approver.id}&decision=Approved"
            reject_url = f"https://api.pixeladvant.com/api/approve-decision?candidate_id={candidate.pk}&approver_id={approver.id}&decision=Reject"

            html_message = f"""
                <html>
                <body style="font-family: Arial, sans-serif;">
                    <h2>Candidate Screening Decision Requested</h2>
                    <p>Dear {approver.role},</p>
                    <p>This is a reminder to review the candidate for requisition ID <strong>{requisition_id}</strong>.</p>

                    <table style="border-collapse: collapse; width: 100%; margin-bottom: 20px;">
                        <tr><td><strong>Candidate Name:</strong></td><td>{candidate.candidate_first_name + " " + candidate.candidate_last_name}</td></tr>
                        <tr><td><strong>Email:</strong></td><td>{candidate.Email}</td></tr>
                        <tr><td><strong>Applied Position:</strong></td><td>{candidate.Req_id_fk.position_information.job_position}</td></tr>
                        <tr><td><strong>Rating:</strong></td><td>{candidate.Final_rating}</td></tr>
                        <tr><td><strong>Feedback:</strong></td><td>{candidate.Feedback}</td></tr>
                    </table>

                    <p>
                        <a href="{approve_url}" style="padding: 10px 20px; background-color: #4CAF50; color: white; text-decoration: none; border-radius: 5px;">✅ Approve</a>
                        &nbsp;&nbsp;
                        <a href="{reject_url}" style="padding: 10px 20px; background-color: #f44336; color: white; text-decoration: none; border-radius: 5px;">❌ Reject</a>
                    </p>

                    <p>Thank you,<br/>Hiring System</p>
                </body>
                </html>
            """

            send_mail(
                subject=subject,
                message="",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[approver.email],
                html_message=html_message
            )

        return Response(api_json_response_format(True, "Emails resent successfully", 200, {}), status=200)

    except Candidate.DoesNotExist:
        return Response(api_json_response_format(False, "Candidate not found", 404, {}), status=200)
    except Exception as e:
        return Response(api_json_response_format(False, f"Error resending emails: {e}", 500, {}), status=200)


class CandidateApprovalDecisionView(APIView):
    def get(self, request):
        candidate_id = request.GET.get("candidate_id")
        approver_id = request.GET.get("approver_id")
        decision = request.GET.get("decision")  # Should be 'Approve' or 'Reject'

        if not candidate_id or not approver_id or not decision:
            error_message = "Missing required query parameters: candidate_id, approver_id, or decision."
            return HttpResponse(self.error_html(error_message))

        try:
            with transaction.atomic():
                approval = get_object_or_404(
                    CandidateApproval,
                    candidate_id=candidate_id,
                    approver_id=approver_id
                )

                approval.decision = decision
                approval.reviewed_at = timezone.now()
                approval.save()

            return HttpResponse(self.success_html(candidate_id, decision))

        except Exception as e:
            error_message = f"Error updating approval decision: {str(e)}"
            return HttpResponse(self.error_html(error_message))

    def success_html(self, candidate_id, decision):
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Decision Recorded</title>
            <style>
                body {{ font-family: Arial, sans-serif; background-color: #f0f8ff; padding: 30px; text-align: center; }}
                .box {{ background-color: #e6ffed; border: 2px solid #4caf50; padding: 20px; border-radius: 8px; display: inline-block; }}
                h2 {{ color: #2e7d32; }}
                .icon {{ font-size: 60px; color: #4caf50; }}
            </style>
        </head>
        <body>
            <div class="box">
                <div class="icon">✅</div>
                <h2>Decision Recorded Successfully</h2>
                <p>Your response: <strong>{decision}</strong> for Candidate ID <strong>{candidate_id}</strong>.</p>
                <p>Thanks for your review!</p>
            </div>
        </body>
        </html>
        """

    def error_html(self, error_message):
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Decision Failed</title>
            <style>
                body {{ font-family: Arial, sans-serif; background-color: #fff3f3; padding: 30px; text-align: center; }}
                .box {{ background-color: #ffe6e6; border: 2px solid #f44336; padding: 20px; border-radius: 8px; display: inline-block; }}
                h2 {{ color: #c62828; }}
                .icon {{ font-size: 60px; color: #f44336; }}
            </style>
        </head>
        <body>
            <div class="box">
                <div class="icon">❌</div>
                <h2>Oops! Something Went Wrong</h2>
                <p>{error_message}</p>
                <p>Please try again or contact support.</p>
            </div>
        </body>
        </html>
        """

class OfferDetailsViewSet(APIView):
    def get(self, request):
        try:
            negotiations = OfferNegotiation.objects.filter(
                negotiation_status__in=["Pending", "Generated", "Approved", "Successful", "open", "In progress"]
            ).select_related("requisition").prefetch_related("approvals__approver")

            data = []

            for offer in negotiations:
                requisition = offer.requisition
                candidate = Candidate.objects.filter(Req_id_fk=requisition.RequisitionID).first()
                approvals = offer.approvals.select_related("approver")

                approvers_data = []
                for a in approvals:
                    approver = a.approver
                    if approver:
                        approvers_data.append({
                            "role": approver.role,
                            "job_title": approver.job_title,
                            "first_name": approver.first_name,
                            "last_name": approver.last_name,
                            "email": approver.email,
                            "contact_number": approver.contact_number,
                            "set_as_approver": "Yes" if approver.set_as_approver else "No"
                        })

                generated_offer_data = []
                generated_offers = GeneratedOffer.objects.filter(
                    requisition=requisition,
                    candidate=candidate
                )

                for gen_offer in generated_offers:
                    salary_data = [
                        {"name": s.name, "value": s.value}
                        for s in gen_offer.salary_components.all()
                    ]
                    variable_pay_data = [
                        {"name": v.name, "value": v.value}
                        for v in gen_offer.variable_pay_components.all()
                    ]

                    generated_offer_data.append({
                        "first_name": candidate.candidate_first_name,
                        "last_name": candidate.candidate_last_name,
                        "candidate_email": candidate.Email,
                        "offer_id": gen_offer.id,
                        "job_title": gen_offer.job_title,
                        "job_city": {
                            "label": gen_offer.job_city,
                            "value": gen_offer.job_city
                        },
                        "job_country": {
                            "label": gen_offer.job_country,
                            "value": gen_offer.job_country
                        },
                        "currency": {
                            "label": gen_offer.currency,
                            "value": gen_offer.currency
                        },

                        "salary": salary_data,
                        "variable_pay": variable_pay_data,
                        "estimated_start_date": gen_offer.estimated_start_date,
                        "recruiter_email": gen_offer.recruiter_email,
                        "negotiation_status": gen_offer.negotiation_status,
                        "created_at": gen_offer.created_at,
                        "updated_at": gen_offer.updated_at
                    })

                data.append({
                    "Req ID": requisition.RequisitionID,
                    "Candidate ID": candidate.CandidateID if candidate else "",
                    "Candidate First Name": candidate.candidate_first_name if candidate else "",
                    "Candidate Last Name": candidate.candidate_last_name if candidate else "",
                    "Req_ID": requisition.RequisitionID,
                    "Client_Id": requisition.client_id,
                    "Client_Name": offer.client_name,
                    "Candidate_Id": candidate.CandidateID if candidate else "N/A",
                    "Candidate_First_Name": offer.first_name,
                    "Candidate_Last_Name": offer.last_name,
                    "Applied_Position": offer.position_applied,
                    "Approvers": approvers_data,
                    "Generate_Offer": generated_offer_data,
                    "Status": offer.negotiation_status
                })

            return Response(api_json_response_format(True, "Offer details grid retrieved successfully", 200, data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error loading grid. {str(e)}", 500, []), status=200)


class GenerateOfferPrefillView(APIView):
    def post(self, request):
        try:
            req_id = request.data.get("requisition_id")
            candidate_id = request.data.get("candidate_id")

            if not req_id or not candidate_id:
                return Response(api_json_response_format(False, "Missing requisition or candidate ID", 400, {}), status=200)

            requisition = JobRequisition.objects.filter(RequisitionID=req_id).first()
            candidate = Candidate.objects.filter(CandidateID=candidate_id).first()
            negotiation = OfferNegotiation.objects.filter(requisition=requisition).first()
            # Get recruiter name from requisition
            recruiter_name = requisition.Recruiter if requisition and requisition.Recruiter else None
            if "," in recruiter_name:
                recruiter_name = recruiter_name.split(",")[0].strip()

            # Lookup recruiter email from UserDetails
            recruiter = UserDetails.objects.filter(Name__icontains=recruiter_name).first()
            recruiter_email = recruiter.Email if recruiter and recruiter.Email else "N/A"


            if not requisition or not candidate or not negotiation:
                return Response(api_json_response_format(False, "Data not found for prefill", 404, {}), status=200)

            # Structured salary and variable pay format
            salary_components = [
                {
                    "name": "Base Salary",
                    "value": str(negotiation.offered_salary or negotiation.expected_salary or 0)
                }
            ]

            variable_pay_components = [
                {
                    "name": "Variable Pay",
                    "value": "0"
                }
            ]

            prefill_data = {
                "first_name": negotiation.first_name,
                "last_name": negotiation.last_name,
                "candidate_email": candidate.Email,
                "recruiter_email": recruiter_email,
                "job_title": negotiation.offered_title or negotiation.expected_title,
                "estimated_start_date": negotiation.offered_doj or negotiation.expected_doj,
                "job_city": {
                    "label": negotiation.offered_location or negotiation.expected_location,
                    "value": negotiation.offered_location or negotiation.expected_location
                },
                "job_country": {
                    "label": "India",
                    "value": "India"
                },
                "currency": {
                    "label": "INR",
                    "value": "INR"
                },
                "salary": salary_components,
                "variable_pay": variable_pay_components
            }

            return Response(api_json_response_format(True, "Prefill data retrieved successfully", 200, prefill_data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error retrieving prefill data. {str(e)}", 500, {}), status=200)


class GenerateOfferView(APIView):
    def post(self, request):
        try:
            data = request.data

            requisition = JobRequisition.objects.filter(RequisitionID=data.get("req_id")).first()
            candidate = Candidate.objects.filter(CandidateID=data.get("candidate_id")).first()

            if not requisition or not candidate:
                return Response(api_json_response_format(False, "Invalid requisition or candidate ID", 400, {}), status=200)

            offer, created = GeneratedOffer.objects.update_or_create(
                requisition=requisition,
                candidate=candidate,
                defaults={
                    "recruiter_email": data.get("recruiter_email"),
                    "job_title": data.get("job_title"),
                    "job_city": data.get("job_city"),
                    "job_country": data.get("job_country"),
                    "currency": data.get("currency"),
                    "estimated_start_date": data.get("estimated_start_date"),
                    "negotiation_status": "Generated"
                }
            )

            # 🔁 Clear old salary & variable pay components
            OfferSalaryComponent.objects.filter(offer=offer).delete()
            OfferVariablePayComponent.objects.filter(offer=offer).delete()

            # 💰 Save salary components
            for item in data.get("salary", []):
                if item.get("name") and item.get("value"):
                    OfferSalaryComponent.objects.create(
                        offer=offer,
                        name=item["name"],
                        value=item["value"]
                    )

            # 🎯 Save variable pay components
            for item in data.get("variable_pay", []):
                if item.get("name") and item.get("value"):
                    OfferVariablePayComponent.objects.create(
                        offer=offer,
                        name=item["name"],
                        value=item["value"]
                    )
            CandidateProfile.objects.update_or_create(
            candidate_id=candidate.CandidateID,
            defaults={
                "candidate_id": candidate.CandidateID,
                "first_name": candidate.candidate_first_name,
                "last_name": candidate.candidate_last_name,
                "date_of_joining": data.get("estimated_start_date")
                }
            )

            message = "Offer updated successfully" if not created else "Offer generated successfully"
            return Response(api_json_response_format(True, message, 200, {
                "offer_id": offer.id,
                "candidate_name": f"{candidate.candidate_first_name} {candidate.candidate_last_name}",
                "job_title": offer.job_title,
                "status": offer.negotiation_status
            }), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error saving offer. {str(e)}", 500, {}), status=200)

class OfferRetrievalView(APIView):
    def post(self, request):
        try:
            data = request.data
            offer_id = data.get("id")

            def extract_components(queryset):
                return [
                    {"name": component.name, "value": component.value}
                    for component in queryset
                ]

            if offer_id:
                offer = GeneratedOffer.objects.select_related("candidate", "requisition").filter(id=offer_id).first()
                if not offer:
                    return Response(api_json_response_format(False, "Offer not found", 404, {}), status=200)

                offer_data = {
                    "offer_id": offer.id,
                    "req_id": offer.requisition.RequisitionID,
                    "candidate_id": offer.candidate.CandidateID,
                    "Req ID": offer.requisition.RequisitionID,
                    "Candidate ID": offer.candidate.CandidateID ,
                    "Candidate First Name": offer.candidate.candidate_first_name,
                    "Candidate Last Name": offer.candidate.candidate_last_name,
                    "candidate_name": f"{offer.candidate.candidate_first_name} {offer.candidate.candidate_last_name}",
                    "job_title": offer.job_title,
                    "job_city": offer.job_city,
                    "job_country": offer.job_country,
                    "currency": offer.currency,
                    "salary": extract_components(offer.salary_components.all()),
                    "variable_pay": extract_components(offer.variable_pay_components.all()),
                    "estimated_start_date": offer.estimated_start_date,
                    "recruiter_email": offer.recruiter_email,
                    "negotiation_status": offer.negotiation_status,
                    "created_at": offer.created_at,
                    "updated_at": offer.updated_at
                }
                return Response(api_json_response_format(True, "Offer retrieved successfully", 200, offer_data), status=200)

            offers = GeneratedOffer.objects.select_related("candidate", "requisition").all()
            all_data = []
            for offer in offers:
                all_data.append({
                    "offer_id": offer.id,
                    "req_id": offer.requisition.RequisitionID,
                    "candidate_id": offer.candidate.CandidateID,
                    "candidate_name": f"{offer.candidate.candidate_first_name} {offer.candidate.candidate_last_name}",
                    "job_title": offer.job_title,
                    "job_city": offer.job_city,
                    "job_country": offer.job_country,
                    "currency": offer.currency,
                    "salary": extract_components(offer.salary_components.all()),
                    "variable_pay": extract_components(offer.variable_pay_components.all()),
                    "estimated_start_date": offer.estimated_start_date,
                    "recruiter_email": offer.recruiter_email,
                    "negotiation_status": offer.negotiation_status,
                    "created_at": offer.created_at,
                    "updated_at": offer.updated_at
                })
            return Response(api_json_response_format(True, "All offers retrieved successfully", 200, all_data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error retrieving offer(s). {str(e)}", 500, {}), status=200)


# class GenerateOfferView(APIView):
#     def post(self, request):
#         try:
#             requisition_id = request.data.get("requisition_id")
#             candidate_id = request.data.get("candidate_id")

#             if not requisition_id:
#                 return Response(api_json_response_format(False, "Missing requisition_id", 400, {}), status=200)

#             queryset = Candidate.objects.filter(Req_id_fk__RequisitionID=requisition_id)

#             if candidate_id:
#                 queryset = queryset.filter(CandidateID=candidate_id)

#             response_data = []
#             for candidate in queryset:
#                 job_req = candidate.Req_id_fk
#                 hm_approved = "Yes" if job_req.HiringManager and job_req.HiringManager.Approved else "No"
#                 fpna_approved = "Yes" if job_req.CommentFromBusinessOps != "" else "No"  # simple proxy

#                 offer_status = "Offer Sent" if candidate.Result == "Offer Accepted" else (
#                     "Need to Generate" if candidate.Result in ["Shortlisted", None] else "Unknown"
#                 )

#                 response_data.append({
#                     "req_id": job_req.RequisitionID,
#                     "client_id": job_req.client_id,
#                     "client_name": job_req.company_client_name,
#                     "candidate_id": candidate.CandidateID,
#                     "candidate_first_name": candidate.candidate_first_name,
#                     "candidate_last_name": candidate.candidate_last_name,
#                     "hm_approver": hm_approved,
#                     "fpna_approver": fpna_approved,
#                     "generate_offer": ["Generate", "Preview"],
#                     "status": offer_status,
#                     "actions": ["Edit", "Delete"]
#                 })

#             return Response(api_json_response_format(True, "Offer data generated", 200, response_data), status=200)

#         except Exception as e:
#             return Response(api_json_response_format(False, str(e), 500, {}), status=200)


# class JobRequisitionViewSetget(viewsets.ReadOnlyModelViewSet):
#     queryset = JobRequisition.objects.all()
#     serializer_class = JobRequisitionSerializerget

#     def list(self, request):
#         try:
#             queryset = self.filter_queryset(self.get_queryset())
#             serializer = self.get_serializer(queryset, many=True)
#             return Response({
#                 "data": serializer.data,
#                 "status_code": 200
#             })
#         except Exception as e:
#             return Response({
#                 "error": "Error fetching job requisitions",
#                 "details": str(e),
#                 "status_code": 500
#             }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def extract_requested_fields(queryset, selected_fields, field_map):
    model_fields = set()
    field_lookup = {}  # display_name → actual ORM field path(s)

    for display in selected_fields:
        mapping = field_map.get(display)
        if isinstance(mapping, str):
            model_fields.add(mapping)
            field_lookup[display] = [mapping]
        elif isinstance(mapping, list):
            model_fields.update(mapping)
            field_lookup[display] = mapping
        else:
            field_lookup[display] = []

    data = list(queryset.values(*model_fields))
    result = []

    for row in data:
        flat = {}
        for display, fields in field_lookup.items():
            if fields:
                values = [row.get(f) for f in fields if f in row]
                value = values[0] if len(values) == 1 else values if any(values) else None

                # 🧼 Strip HTML from JD-like fields
                if display.lower() in ["jd", "jd_details", "job_description"] and isinstance(value, str):
                    value = strip_tags(value)

                flat[display] = value
            else:
                flat[display] = None
        result.append(flat)

    return result


class JobRequisitionFlatViewSet(viewsets.ViewSet):
    def create(self, request):
        try:
            selected_fields = request.data.get("fields", [])
            user_role = request.data.get("user_role")  # 👈 get the role from request

            if not isinstance(selected_fields, list) or not selected_fields:
                return Response(
                    api_json_response_format(False, "Missing or invalid 'fields' parameter.", 400, {}),
                    status=200
                )

            queryset = JobRequisition.objects.select_related("position_information", "Planning_id").order_by("-RequisitionID")

            # 🔍 Filter 'Incomplete form' based on user_role
            if str(user_role) != "1":
                queryset = queryset.exclude(Status="Incomplete form")

            result_data = extract_requested_fields(queryset, selected_fields, DISPLAY_TO_MODEL_FIELD)

            return Response(
                api_json_response_format(True, "Dynamic field data fetched successfully!", 200, {
                    "data": result_data,
                    "selected_fields": selected_fields
                }),
                status=200
            )

        except Exception as e:
            return Response(
                api_json_response_format(False, "Failed to retrieve dynamic field data: " + str(e), 500, {}),
                status=200
            )


class JobRequisitionPublicViewSet(viewsets.ViewSet):
    def list(self, request):
        try:
            requisitions = JobRequisition.objects.prefetch_related(
                'interview_team', 'teams'
            ).select_related('position_information', 'billing_details', 'posting_details')

            data = []
            for obj in requisitions:
                internal_title = getattr(obj.position_information, "internal_title", "") if obj.position_information else ""
                location = getattr(obj.position_information, "location", "") if obj.position_information else ""
                template_name = f"{internal_title} - {location}".strip(" -")

                job_template_data = JobTemplateSerializer({
                    "requisition_details": getattr(obj, "position_information", None),
                    "billing": getattr(obj, "billing_details", None),
                    "posting": getattr(obj, "posting_details", None),
                    "interviewers": obj.interview_team.all(),
                    "functional_teams": obj.teams.all(),
                }).data

                job_template_data["template_name"] = template_name  # Injecting display label
                data.append(job_template_data)


            return Response(api_json_response_format(
                True,
                "Job templates retrieved successfully!",
                200,
                {"job_template": data}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                "Error while fetching data for template"+str(e),
                500,
                {}
            ), status=200)


# Mapping of payload keys to configuration categories
JOB_REQUISITION_CONFIG_MAP = {
    # position_information block
    "internal_title": "Internal Job Title",
    "external_title": "External Job Title",
    "job_position": "Position",
    "business_line": "Business Line",
    "business_unit": "Business Unit",
    "division": "Division",
    "department": "Department",
    "location": "Location",
    "geo_zone": "Geo Zone",
    "band": "Band",
    "sub_band": "Sub Band",
    "client_interview": "Client Interview",
    "requisition_type": "Requisition Type",
    "working_model": "Working Model",

    # posting_details block
    "experience": "Experience",
    "qualification": "Qualification",
    "designation": "Designation",
    "job_region": "Job Region",

    # skills_required block
    "primary_skills": "Primary Skills",
    "secondary_skills": "Secondary Skills",
}

# Simulated config store (replace with DB or persistent store)
CONFIG_STORE = {}

def extract_values(raw):
    if isinstance(raw, list):
        return [
            item.get("value") or item.get("label") if isinstance(item, dict)
            else item.strip()
            for item in raw if item
        ]
    elif isinstance(raw, dict):
        return [raw.get("value") or raw.get("label")]
    elif isinstance(raw, str):
        return [v.strip() for v in raw.split(",") if v.strip()]
    return []

def config_value_exists(category, value):
    """Check if a value already exists in the config store for a category."""
    existing = CONFIG_STORE.get(category, set())
    return value.strip().lower() in (v.lower() for v in existing)

def ensure_config_value(category, value):
    """Insert value into config store if not already present."""
    if category not in CONFIG_STORE:
        CONFIG_STORE[category] = set()
    CONFIG_STORE[category].add(value.strip())

def process_job_requisition_config(payload):
    # Merge all relevant blocks
    combined = {
        **payload.get("position_information", {}),
        **payload.get("posting_details", {}),
        **payload.get("skills_required", {})
    }

    # Iterate through mapped keys and ingest values
    for key, category in JOB_REQUISITION_CONFIG_MAP.items():
        values = extract_values(combined.get(key))
        for val in values:
            if val and not config_value_exists(category, val):
                ensure_config_value(category, val)


class JobRequisitionViewSet(viewsets.ModelViewSet):
    queryset = JobRequisition.objects.select_related("HiringManager").all()
    serializer_class = JobRequisitionSerializer

    def create(self, request, *args, **kwargs):
        try:
            serializer = JobRequisitionSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            job_requisition = serializer.save()

            return Response(api_json_response_format(True, "Initial requisition created.", 200, {
                "RequisitionID": job_requisition.RequisitionID
                # "Planning_id": job_requisition.Planning_id.hiring_plan_id
            }), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, "Failed to create initial requisition. " + str(e), 500, {}), status=200)
    
    @action(detail=False, methods=['put'], url_path='update-requisition')
    def update_requisition(self, request):
        try:
            requisition_id = request.data.get("requisition_id")
            user_role = request.data.get("user_role")

            if user_role != 1:
                return Response(api_json_response_format(False, "Only Hiring Managers can update requisitions.", 403, {}), status=200)
            if not requisition_id:
                return Response(api_json_response_format(False, "Requisition ID is required.", 400, {}), status=200)

            instance = get_object_or_404(JobRequisition, RequisitionID=requisition_id)

            fixed_payload = request.data.copy()
            position_title_raw = fixed_payload.get("PositionTitle")
            if isinstance(position_title_raw, list):
                fixed_payload["PositionTitle"] = ", ".join([item.get("value", "") for item in position_title_raw if isinstance(item, dict)])
            elif isinstance(position_title_raw, dict):
                fixed_payload["PositionTitle"] = position_title_raw.get("value", "")
            elif isinstance(position_title_raw, str):
                fixed_payload["PositionTitle"] = position_title_raw
            else:
                fixed_payload["PositionTitle"] = ""


            # 🔧 Normalize posting_details
            posting_block = fixed_payload.get("posting_details", {})
            def normalize_to_string(value):
                return ", ".join(value) if isinstance(value, list) else value or ""

            posting_block["experience"] = normalize_to_string(posting_block.get("experience"))
            posting_block["designation"] = normalize_to_string(posting_block.get("designation"))
            posting_block["job_region"] = normalize_to_string(posting_block.get("job_region"))
            posting_block["qualification"] = normalize_to_string(posting_block.get("qualification"))
            fixed_payload["posting_details"] = posting_block

            # 🔧 Normalize position_information multi-value fields
            position_block = fixed_payload.get("position_information", {})
            def extract_values(field):
                if isinstance(field, list):
                    return ", ".join([item.get("value", "") for item in field if isinstance(item, dict)])
                elif isinstance(field, dict):
                    return field.get("value", "")
                elif isinstance(field, str):
                    return field
                return ""

            multi_fields = [
                "internal_title", "external_title", "job_position",
                "business_unit", "business_line", "division", "department", "location",
                "geo_zone", "band", "sub_band", "working_model",
                "client_interview", "requisition_type"
            ]

            for field in multi_fields:
                position_block[field] = extract_values(position_block.get(field))

            fixed_payload["position_information"] = position_block

            # 🔁 Process config ingestion if needed
            process_job_requisition_config(fixed_payload)

            # 🔄 Update instance
            serializer = self.get_serializer(instance, data=fixed_payload, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            return Response(api_json_response_format(True, "Job requisition updated successfully!", 200, {
                "requisition_id": serializer.instance.RequisitionID
            }), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error while updating requisition. {str(e)}", 500, {}), status=200)

    
    
    @action(detail=False, methods=['delete'], url_path='delete-requisition')
    def delete_requisition(self, request):
        try:
            user_role = request.data.get("user_role")
            requisition_id = request.data.get("requisition_id")

            if user_role != 1:
                return Response(api_json_response_format(False, "Only Hiring Managers can delete requisitions.", 403, {}), status=200)
            if not requisition_id:
                return Response(api_json_response_format(False, "requisition_id is required.", 400, {}), status=200)

            try:
                requisition = JobRequisition.objects.get(RequisitionID=requisition_id)
            except JobRequisition.DoesNotExist:
                return Response(api_json_response_format(False, "Job requisition not found.", 404, {}), status=200)

            requisition.delete()
            return Response(api_json_response_format(True, "Job requisition deleted successfully!", 200, {}), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, "Error while deleting requisition. " + str(e), 500, {}), status=200)

    @action(detail=False, methods=["post"], url_path="list-requisitions")
    def list_requisitions(self, request):
        try:
            user_role = request.data.get("user_role")
            username = request.data.get("username")

            if user_role not in [1, 2, 3, 5]:
                return Response(api_json_response_format(False, "Unauthorized", 403, {}), status=200)

            # 🔍 Role-based filtering + ordering (latest first)
            if user_role == 2:  # Recruiter
                queryset = JobRequisition.objects.filter(Recruiter__icontains=username).order_by("-RequisitionID")

            elif user_role == 3:  # Business Ops
                queryset = JobRequisition.objects.all().order_by("-RequisitionID")
            elif user_role == 5:  # vendor
                queryset = JobRequisition.objects.filter(Recruiter__icontains=username).order_by("-RequisitionID")
            else:  # Hiring Manager
                queryset = JobRequisition.objects.filter(HiringManager=request.user).order_by("-RequisitionID")


            current_date = datetime.now(dt_timezone.utc).date()

            data = []
            for requisition in queryset:
                info = getattr(requisition, "position_information", None)
                billing = getattr(requisition, "billing_details", None)

                start_date = billing.contract_start_date if billing else None
                end_date = billing.contract_end_date if billing else None
                age = (current_date - start_date).days if start_date else "N/A"
                age = max(age, 0) if isinstance(age, int) else "N/A"
                planning_obj = getattr(requisition, "Planning_id", None)
                planning_id = planning_obj.hiring_plan_id if planning_obj else "Not Provided"

                data.append({
                    "PlanningID": planning_id,
                    "RequisitionID": requisition.RequisitionID,
                    "ClientName": requisition.company_client_name or "Not Provided",
                    "ClientID": requisition.client_id or "Not Provided",
                    "JobTitle": requisition.PositionTitle,
                    "HiringManager": getattr(requisition.HiringManager, "Name", "Unknown"),
                    "Recruiter": username,
                    "JobPosting": getattr(requisition, "PostingSource", "Not Specified"),
                    "StartDate": start_date,
                    "DueDate": end_date,
                    "HiringStatus": requisition.Status,
                    "Age(Days)": age
                })

            return Response(api_json_response_format(True, "Job requisitions retrieved successfully!", 200, data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error fetching job requisitions. {str(e)}", 500, {}), status=200)

    @action(detail=False, methods=["POST"], url_path='approve_requisition')
    def update_requisition_status(self, request):
        try:
            user_role = request.data.get("user_role")
            req_data = request.data.get("req_data")

            if user_role != 3:
                return Response(api_json_response_format(
                    False,
                    "Only Business Ops can update requisition statuses.",
                    403,
                    {}
                ), status=200)

            if not isinstance(req_data, list):
                return Response(api_json_response_format(
                    False,
                    "req_data should be a list of requisition update objects.",
                    400,
                    {}
                ), status=200)

            updated = []
            not_found = []

            for item in req_data:
                req_id = item.get("req_id")
                new_status = item.get("status")
                comment = item.get("comment", "")

                if not req_id or not new_status:
                    continue  # skip incomplete entries

                try:
                    requisition = JobRequisition.objects.get(RequisitionID=req_id)
                    requisition.Status = new_status
                    requisition.CommentFromBusinessOps = comment
                    requisition.save()
                    updated.append(req_id)
                except JobRequisition.DoesNotExist:
                    not_found.append(req_id)
            decision = "approved" if new_status.lower() == "approved" else "rejected"
            login_url = "https://hiring.pixeladvant.com/"
            # requisition_url = f"https://yourdomain.com/requisitions/{requisition.pk}/details/"

            email_body = f"""
            Hi Team,

            Your Job Requisition '{requisition.RequisitionID}' has been {decision} by Business Ops.

            📌 Status: {new_status}
            📝 Comment: {comment or "No comments provided."}

            You can login to review the requisition status:
            🔗 Login: {login_url}

            Regards,
            Business Ops Team
            """

            send_mail(
                subject=f"Requisition '{requisition.RequisitionID}' {decision.capitalize()}",
                message=email_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=['anand40593@outlook.com'],  # 🔄 Use requisition owner’s email if dynamic
                fail_silently=True,
            )

            return Response(api_json_response_format(
                True,
                "Requisitions updated successfully.",
                200,
                {
                    "updated": updated,
                    "not_found": not_found
                }
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error updating requisitions. {str(e)}",
                500,
                {}
            ), status=200)
    
    @action(detail=False, methods=["POST"])
    def get_approved_extra_details(self, request):
        try:
            user_role = request.data.get("user_role")

            if user_role not in [2, 3]:
                return Response(api_json_response_format(False, "Unauthorized access.", 403, {}), status=200)

            approved_ids = JobRequisition.objects.filter(Status="Approved").values_list("RequisitionID", flat=True)
            details = RequisitionDetails.objects.filter(requisition_id__in=approved_ids).order_by("requisition_id")

            data = [{
                "requisition_id": d.requisition_id,
                "internal_title": d.internal_title,
                "external_title": d.external_title,
                "position": d.job_position,
                "business_line": d.business_line,
                "business_unit": d.business_unit,
                "division": d.division,
                "department": d.department,
                "location": d.location,
                "geo_zone": d.geo_zone,
                "employee_group": d.employee_group,
                "employee_sub_group": d.employee_sub_group,
                "contract_start_date": d.contract_start_date.isoformat() if d.contract_start_date else "",
                "contract_end_date": d.contract_end_date.isoformat() if d.contract_end_date else "",
                "career_level": d.career_level,
                "band": d.band,
                "sub_band": d.sub_band,
                "primary_skills": d.primary_skills,
                "secondary_skills": d.secondary_skills,
                "mode_of_working": d.working_model,
                "requisition_type": d.requisition_type,
                "client_interview": d.client_interview,
                "required_score": d.required_score,
                "onb_coordinator": d.onb_coordinator,
                "onb_coordinator_team": d.onb_coordinator_team,
                "isg_team": d.isg_team,
                "interviewer_teammate_employee_id": d.interviewer_teammate_employee_id
            } for d in details]

            return Response(api_json_response_format(True, "Extra requisition details retrieved successfully!", 200, data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, "Error retrieving extra requisition details."+ str(e), 500, {}), status=200)

    @action(detail=False, methods=["POST"])
    def get_requisition_by_id(self, request):
        try:
            requisition_id = request.data.get("req_id")
            if not requisition_id:
                return Response(api_json_response_format(False, "Requisition ID is required.", 400, {}), status=200)

            instance = get_object_or_404(JobRequisition, RequisitionID=requisition_id)

            # Related objects
            details = getattr(instance, "position_information", None)
            billing = getattr(instance, "billing_details", None)
            posting = getattr(instance, "posting_details", None)
            asset = getattr(instance, "asset_details", None)
            plan = instance.Planning_id if instance.Planning_id else None

            def to_label_value_list(raw):
                if not raw:
                    return []
                if isinstance(raw, str):
                    return [{"label": v.strip(), "value": v.strip()} for v in raw.split(",") if v.strip()]
                return []

            response_payload = {
                "Planning_id": getattr(instance.Planning_id, "hiring_plan_id", "Not Provided") if instance.Planning_id else "Not Provided",
                "HiringManager": getattr(instance.HiringManager, "Name", "Unknown") if instance.HiringManager else "Unknown",
                "PositionTitle": to_label_value_list(instance.PositionTitle),
                "requisition_id": instance.RequisitionID or "Not Provided",

                "position_information": {
                    "internal_title": to_label_value_list(getattr(details, "internal_title", "")),
                    "external_title": to_label_value_list(getattr(details, "external_title", "")),
                    "job_position": to_label_value_list(getattr(details, "job_position", "")),
                    "company_client_name": instance.company_client_name or "Not Provided",
                    "business_unit": to_label_value_list(getattr(details, "business_unit", "")),
                    "business_line": to_label_value_list(getattr(details, "business_line", "")),
                    "division": to_label_value_list(getattr(details, "division", "")),
                    "department": to_label_value_list(getattr(details, "department", "")),
                    "location": to_label_value_list(getattr(details, "location", "") or (getattr(plan, "location", "") if plan else "")),
                    "geo_zone": to_label_value_list(getattr(details, "geo_zone", "")),
                    "career_level": getattr(details, "career_level", ""),
                    "band": to_label_value_list(getattr(details, "band", "")),
                    "sub_band": to_label_value_list(getattr(details, "sub_band", "")),
                    "working_model": to_label_value_list(getattr(details, "working_model", "") or (getattr(plan, "working_model", "") if plan else "")),
                    "client_interview": to_label_value_list("Yes" if getattr(details, "client_interview", "") == "Yes" else ""),
                    "requisition_type": to_label_value_list(getattr(details, "requisition_type", "")),
                    "date_of_requisition": instance.requisition_date,
                    "due_date_of_requisition": instance.due_requisition_date
                },

                "skills_required": {
                    "primary_skills": [s.strip() for s in getattr(details, "primary_skills", "").split(",")] if getattr(details, "primary_skills", "") else [],
                    "secondary_skills": [s.strip() for s in getattr(details, "secondary_skills", "").split(",")] if getattr(details, "secondary_skills", "") else []
                },

                "billing_details": {
                    "billing_type": getattr(billing, "billing_type", ""),
                    "billing_start_date": getattr(billing, "billing_start_date", ""),
                    "billing_end_date": getattr(billing, "billing_end_date", ""),
                    "contract_start_date": getattr(billing, "contract_start_date", ""),
                    "contract_end_date": getattr(billing, "contract_end_date", "")
                },

                "posting_details": {
                    "experience": [e.strip() for e in getattr(posting, "experience", "").split(",")] if getattr(posting, "experience", "") else [],
                    "qualification": [q.strip() for q in getattr(posting, "qualification", "").split(",")] if getattr(posting, "qualification", "") else [],
                    "designation": [d.strip() for d in getattr(posting, "designation", "").split(",")] if getattr(posting, "designation", "") else [],
                    "job_region": [r.strip() for r in getattr(posting, "job_region", "").split(",")] if getattr(posting, "job_region", "") else [],
                    "required_score": getattr(posting, "required_score", ""),
                    "internalDesc": getattr(posting, "internal_job_description", ""),
                    "externalDesc": getattr(posting, "external_job_description", ""),
                    "questions": [],
                    "Competencies": []
                },

                "asset_details": {
                    "laptop_type": getattr(asset, "laptop_type", ""),
                    "laptop_needed": "Yes" if getattr(asset, "laptop_needed", "") == "Yes" else "No",
                    "comments": getattr(asset, "comments", "")
                }
            }

            return Response(api_json_response_format(True, "Requisition retrieved successfully!", 200, response_payload), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error retrieving requisition. {str(e)}", 500, {}), status=200)
    @action(detail=False, methods=["get"], url_path="recruiter-vendor-dropdown")
    def recruiter_vendor_dropdown(self, request):
        users = UserDetails.objects.filter(RoleID__in=[2, 5])

        recruiter_list = []
        vendor_list = []

        for user in users:
            entry = {
                "label": user.Name,
                "value": user.Name,
                "email": user.Email
            }
            if user.RoleID == 2:
                recruiter_list.append(entry)
            elif user.RoleID == 5:
                vendor_list.append(entry)

        response_data = {
            "Recruiter": recruiter_list,
            "Vendor": vendor_list
        }

        return Response(api_json_response_format(True, "Recruiter and Vendor list fetched", 200, response_data), status=200)


@api_view(['POST'])
def assign_recruiter_to_requisition(request):
    requisition_id = request.data.get("requisition_id")
    recruiter_name = request.data.get("recruiter_name")

    if not requisition_id or not recruiter_name:
        return Response(api_json_response_format(
            False, "Missing requisition_id or recruiter_name", 400, {}
        ), status=200)

    try:
        requisition = JobRequisition.objects.get(RequisitionID=requisition_id)

        # ✅ Enforce status check
        if requisition.Status != "Approved":
            return Response(api_json_response_format(
                True, f"Recruiter can only be assigned to requisitions with status 'Approved'. Current status: '{requisition.Status}'", 200, {}
            ), status=200)

        requisition.Recruiter = recruiter_name
        requisition.save()

        return Response(api_json_response_format(
            True, "Recruiter name updated successfully", 200, {
                "RequisitionID": requisition.RequisitionID,
                "Recruiter": recruiter_name
            }
        ), status=200)

    except JobRequisition.DoesNotExist:
        return Response(api_json_response_format(
            False, "Requisition not found", 404, {}
        ), status=200)

    except Exception as e:
        return Response(api_json_response_format(
            False, f"Error updating recruiter: {str(e)}", 500, {}
        ), status=200)



def get_matching_score(job_description, resume_text, resume_name):
    """Send job description and resume to Ollama model and get only the matching score"""
    prompt = f"""
    You are an AI-powered resume analysis agent.
    Given a job description, compare multiple resumes.
    
    For each resume, **ONLY return the JSON output**:
    
    {{
        "resume_name": "{resume_name}",
        "percentage": 90
    }}

    **DO NOT** provide additional analysis, explanations, keywords, or recommendations.
    
    Job Description: {job_description}
    Resume Text:{resume_text[:2000]} 

    """
    # response = ollama.chat(model='ats_model:latest', messages=[{"role": "user", "content": prompt}])
    # ai_output = response['message']['content']
    response = ollama_model.invoke(prompt)
    return response.strip()  # Remove unnecessary formatting

class ResumeMatchingAPI(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        """Handles resume uploads and returns matching scores only"""
        try:
            job_description = request.data.get('job_description')
            uploaded_files = request.FILES.getlist('resumes')

            if not job_description or not uploaded_files:
                return Response(api_json_response_format(
                    False, "Missing job description or resume files.",400, {}
                ), status=200)

            results = []
            for uploaded_file in uploaded_files:
                resume_text = extract_text_from_pdf(uploaded_file)
                score = get_matching_score(job_description, resume_text, uploaded_file.name)
                results.append(score)

            return Response(api_json_response_format(
                True, "Matching scores calculated successfully!", 200, {"matching_scores": results}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, "Error while processing resumes."+ str(e), 500, {}
            ), status=200)

# def send_email():
#     send_mail(
#         'Job Requsition Added',
#         'Please verify and Approve the added job requsition.',
#         settings.EMAIL_HOST_USER,
#         ['anandsivakumar27@gmail.com'],
#         fail_silently=False,
#     )
# send_mail(
#     subject='Test Email',
#     message='This is a test email from Django using Zoho SMTP.',
#     from_email='hiring@pixeladvant.com',
#     recipient_list=['anand040593@gmail.com'],
#     fail_silently=False,
# )

    
def generate_jwt_token(user):
    """ Generate JWT Token """
    payload = {
        'user_id': user.UserID,
        'email': user.Email,
        'role': UserroleDetails.objects.get(RoleID=user.RoleID).RoleName,
        'exp': datetime.datetime.now() + datetime.timedelta(hours=12)  # Token expires in 1 hour
    }
    return jwt.encode(payload, SECRET_KEY, algorithm='HS256')


@csrf_exempt
@api_view(['POST'])
def login_page(request):
    auth_header = request.headers.get("Authorization")

    if not auth_header or not auth_header.startswith("Basic "):
        return Response(api_json_response_format(False, "Authorization header missing or incorrect", 401, {}))

    try:
        _, encoded_credentials = auth_header.split(" ")
        decoded_credentials = base64.b64decode(encoded_credentials).decode("utf-8")
        username, password = decoded_credentials.split(":")
    except Exception as e:
        return Response(api_json_response_format(False, "Invalid credentials encoding. " + str(e), 401, {}))

    try:
        user_details = UserDetails.objects.get(Email=username)
        userrole = UserroleDetails.objects.get(RoleID=user_details.RoleID)

        if not check_password(password, user_details.PasswordHash):
            raise UserDetails.DoesNotExist  # reuse the same error logic

        user, created = User.objects.get_or_create(
            email=username,
            defaults={
                'username': username,
                'password': make_password(password)  # user should be able to sign-in via Django
            }
        )

        user = User.objects.get(email=username)
        refresh = RefreshToken.for_user(user)

        response_data = {
            'role': userrole.RoleName,
            'user_id':user_details.RoleID,
            'Name' :user_details.Name,
            'username': user.get_username(),
            'access': str(refresh.access_token),
            'refresh': str(refresh)
        }
        return Response(api_json_response_format(True, "Login successful", 200, response_data))

    except UserDetails.DoesNotExist:
        return Response(api_json_response_format(False, "Invalid username or password USER_NOT_FOUND", 400, {}))



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    try:
        refresh_token = request.data.get("refresh")

        if not refresh_token:
            return Response(api_json_response_format(False, "Refresh token required for logout", 400, {}))

        token = RefreshToken(refresh_token)
        token.blacklist()

        return Response(api_json_response_format(True, "Logout successful. Token blacklisted.", 200, {}))

    except Exception as e:
        return Response(api_json_response_format(False, f"Error during logout: {str(e)}", 500, {}))
    
class ForgotPasswordView(APIView):
    def post(self, request):
        try:
            email = request.data.get("email")
            user = get_object_or_404(UserDetails, Email=email)

            reset_token = get_random_string(32)
            user.ResetToken = reset_token
            user.save()
            send_mail(
                subject='Test Email',
                message='This is a test email from Django using Zoho SMTP.',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=['anand040593@gmail.com'],
                fail_silently=True,
            )
            # send_mail(
            #     "Password Reset Request",
            #     f"Your reset token is: {reset_token}",
            #     "noreply@example.com",
            #     [user.Email],
            #     fail_silently=False,
            # )

            return Response(api_json_response_format(
                True,
                "Reset token sent to email",
                200,
                {"email": user.Email}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                "Error while sending reset token."+ str(e),500,
                {}
            ), status=200)

class ResetPasswordView(APIView):
    def post(self, request):
        try:
            reset_token = request.data.get("token")
            new_password = request.data.get("new_password")

            user = get_object_or_404(UserDetails, ResetToken=reset_token)
            user.PasswordHash = make_password(new_password)
            user.ResetToken = None
            user.save()

            return Response(api_json_response_format(
                True,
                "Password reset successful",
                200,
                {"email": user.Email}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                "Error while resetting password."+ str(e),
                500,
                {}
            ), status=200)
        
class InterviewPlannerCalculation(APIView):
    def get(self, request):
        try:   
            plan_id = request.data.get('plan_id')
            req_id = request.data.get('req_id') 
            if  req_id:
                filters = {}
                if req_id:
                    filters['requisition_id'] = req_id

                queryset = InterviewPlanner.objects.filter(**filters)
                serializer = InterviewPlannerSerializer(queryset, many=True)            
                return Response(api_json_response_format(True,"Interview planning calculation",200,serializer.data), status=200)
            else:
                excluded_plan_ids = InterviewPlanner.objects.values_list('hiring_plan_id', flat=True)            
                filtered_overview = HiringPlan.objects.exclude(hiring_plan_id__in=excluded_plan_ids).values_list('hiring_plan_id', flat=True)            
                return Response(api_json_response_format(True,"Hiring Plan ID ",0,{"plan_id":filtered_overview}), status=status.HTTP_200_OK)
        except Exception as e:
            return Response(api_json_response_format(False,"Error in interview planning calculation get method "+str(e),500,{}), status=status.HTTP_200_OK)



    def post(self, request):
        try:
            plan_id = request.data.get('plan_id') or ""
            req_id = request.data.get('req_id')
            if not req_id:
                return Response(api_json_response_format(False,"req_id is required.",status.HTTP_400_BAD_REQUEST,{}), status=200) 
            filters = {}
            if plan_id:
                filters['hiring_plan_id'] = plan_id
            if req_id:
                filters['requisition_id'] = req_id

            queryset = InterviewPlanner.objects.filter(**filters)
            if queryset.exists():
                    return Response(api_json_response_format(True, "Already Bandwidth created "+plan_id+" and "+req_id+" ", status.HTTP_200_OK, {}), status=200)
            
            dead_line_days = int(request.data.get('dead_line_days'))
            offer_decline = int(request.data.get('offer_decline'))
            working_hours_per_day = int(request.data.get('working_hours_per_day'))
            no_of_roles_to_hire = int(request.data.get('no_of_roles_to_hire'))
            conversion_ratio = int(request.data.get('conversion_ratio'))
            # elimination = int(request.data.get('elimination'))  # currently unused
            avg_interviewer_time_per_week_hrs = int(request.data.get('avg_interviewer_time_per_week_hrs'))
            interview_round = int(request.data.get('interview_round'))
            interview_time_per_round = int(request.data.get('interview_time_per_round'))
            interviewer_leave_days = int(request.data.get('interviewer_leave_days'))
            no_of_month_interview_happens = int(request.data.get('no_of_month_interview_happens'))  # currently unused
            working_hrs_per_week = int(request.data.get('working_hrs_per_week'))

            required_candidate = int(no_of_roles_to_hire * conversion_ratio)
            decline_adjust_count = (required_candidate * offer_decline) / 100
            total_candidate_pipline = required_candidate + decline_adjust_count
            total_interviews_needed = total_candidate_pipline * interview_round
            total_interview_hrs = total_interviews_needed * interview_time_per_round
            total_interview_weeks = total_interview_hrs / working_hrs_per_week
            no_of_interviewer_need = round(total_interview_hrs / dead_line_days)
            leave_adjustment = round(
                no_of_interviewer_need + (
                    ((interviewer_leave_days * avg_interviewer_time_per_week_hrs) /
                     (dead_line_days * working_hours_per_day)) * no_of_interviewer_need
                )
            )    
            print(f"[DEBUG] request.data type: {type(request.data)}")  # Should be <class 'dict'> or <class 'QueryDict'>
            data = request.data.copy()  
            data["hiring_plan_id"] = plan_id    
            data["requisition_id"] = req_id
            data["required_candidate"] = required_candidate
            data["decline_adjust_count"] = decline_adjust_count
            data["total_candidate_pipline"] = total_candidate_pipline
            data["total_interviews_needed"] = total_interviews_needed
            data["total_interview_hrs"] = total_interview_hrs
            data["total_interview_weeks"] = total_interview_weeks
            data["no_of_interviewer_need"] = no_of_interviewer_need
            data["leave_adjustment"] = leave_adjustment             
            serializer = InterviewPlannerSerializer(data=data)
            if serializer.is_valid():  # 🔥 will show the root cause
                serializer.save()
                interview_plan_data = {
                    "required_candidate": required_candidate,
                    "decline_adjust_count": decline_adjust_count,
                    "total_candidate_pipline": total_candidate_pipline,
                    "total_interviews_needed": total_interviews_needed,
                    "total_interview_hrs": total_interview_hrs,
                    "working_hrs_per_week": working_hrs_per_week,
                    "total_interview_weeks": total_interview_weeks,
                    "no_of_interviewer_need": no_of_interviewer_need,
                    "leave_adjustment": leave_adjustment
                }
                return Response(api_json_response_format(True,"Interview planning calculation completed.",201,interview_plan_data), status=200)

            return Response(api_json_response_format(True,"Interview planning calculation completed.",201,{
                    "required_candidate": required_candidate,
                    "decline_adjust_count": decline_adjust_count,
                    "total_candidate_pipline": total_candidate_pipline,
                    "total_interviews_needed": total_interviews_needed,
                    "total_interview_hrs": total_interview_hrs,
                    "working_hrs_per_week": working_hrs_per_week,
                    "total_interview_weeks": total_interview_weeks,
                    "no_of_interviewer_need": no_of_interviewer_need,
                    "leave_adjustment": leave_adjustment
                }
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(False,"Error in interview planning calculation post method "+str(e),500,{}), status=status.HTTP_200_OK)

    def put(self, request):
        try:
            interview_plan_id = request.data.get('interview_plan_id')
            if not interview_plan_id:
                return Response(api_json_response_format(False, "interview_plan_id is required.", status.HTTP_400_BAD_REQUEST, {}), status=200)

            obj = InterviewPlanner.objects.get(interview_plan_id=interview_plan_id)
            if not obj:
                return Response(api_json_response_format(False, "Interview plan not found.", 404, {}), status=200)

            # 🧮 Perform calculations
            dead_line_days = int(request.data.get('dead_line_days'))
            offer_decline = int(request.data.get('offer_decline'))
            working_hours_per_day = int(request.data.get('working_hours_per_day'))
            no_of_roles_to_hire = int(request.data.get('no_of_roles_to_hire'))
            conversion_ratio = int(request.data.get('conversion_ratio'))
            avg_interviewer_time_per_week_hrs = int(request.data.get('avg_interviewer_time_per_week_hrs'))
            interview_round = int(request.data.get('interview_round'))
            interview_time_per_round = int(request.data.get('interview_time_per_round'))
            interviewer_leave_days = int(request.data.get('interviewer_leave_days'))
            working_hrs_per_week = int(request.data.get('working_hrs_per_week'))

            required_candidate = int(no_of_roles_to_hire * conversion_ratio)
            decline_adjust_count = (required_candidate * offer_decline) / 100
            total_candidate_pipline = required_candidate + decline_adjust_count
            total_interviews_needed = total_candidate_pipline * interview_round
            total_interview_hrs = total_interviews_needed * interview_time_per_round
            total_interview_weeks = total_interview_hrs / working_hrs_per_week
            no_of_interviewer_need = round(total_interview_hrs / dead_line_days)
            leave_adjustment = round(
                no_of_interviewer_need + (
                    ((interviewer_leave_days * avg_interviewer_time_per_week_hrs) /
                    (dead_line_days * working_hours_per_day)) * no_of_interviewer_need
                )
            )

            update_data = request.data.copy()
            update_data["required_candidate"] = required_candidate
            update_data["decline_adjust_count"] = decline_adjust_count
            update_data["total_candidate_pipline"] = total_candidate_pipline
            update_data["total_interviews_needed"] = total_interviews_needed
            update_data["total_interview_hrs"] = total_interview_hrs
            update_data["total_interview_weeks"] = total_interview_weeks
            update_data["no_of_interviewer_need"] = no_of_interviewer_need
            update_data["leave_adjustment"] = leave_adjustment

            serializer = InterviewPlannerSerializer(obj, data=update_data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(
                    True,
                    "Interview planning calculation updated successfully.",
                    200,
                    {
                        "interview_plan_id": interview_plan_id,
                        "required_candidate": required_candidate,
                        "decline_adjust_count": decline_adjust_count,
                        "total_candidate_pipline": total_candidate_pipline,
                        "total_interviews_needed": total_interviews_needed,
                        "total_interview_hrs": total_interview_hrs,
                        "working_hrs_per_week": working_hrs_per_week,
                        "total_interview_weeks": total_interview_weeks,
                        "no_of_interviewer_need": no_of_interviewer_need,
                        "leave_adjustment": leave_adjustment
                    }
                ), status=200)

            return Response(api_json_response_format(False, f"Validation error: {serializer.errors}", 400, {}), status=200)

        except Exception as error:
            return Response(api_json_response_format(
                False,
                f"Error updating interview planning calculation: {str(error)}",
                500,
                {}
            ), status=200)

    def delete(self, request):
        try:
            plan_id = request.data.get('plan_id')
            req_id = request.data.get('req_id')            
            
            if not req_id:
                return Response(api_json_response_format(False,"req_id is required.",status.HTTP_400_BAD_REQUEST,{}), status=200)
            
            filters = {}
            if plan_id:
                filters['hiring_plan_id'] = plan_id
            if req_id:
                filters['requisition_id'] = req_id

            queryset = InterviewPlanner.objects.filter(**filters)
            queryset.delete()            
            return Response(api_json_response_format(True,"Record deleted successfully.",200,{}), status=200)             
        except Exception as error:
            return Response(api_json_response_format(False,"Could not delete data ."+str(error),500,{}), status=204)
 


class InterviewerBandwidthDashboard(APIView):
    def post(self, request):
        try:
            requested_fields = request.data.get('field_names', [
                "hiring_plan_id",
                "requisition_id",
                "leave_adjustment",
                "dead_line_days",
                "offer_decline",
                "no_of_roles_to_hire",
                "interview_round",
                "no_of_interviewer_need",
                "required_candidate"
            ])

            if not isinstance(requested_fields, list):
                return Response(api_json_response_format(False, "field_names must be a list", 400, {}), status=200)

            model_fields = [f.name for f in InterviewPlanner._meta.fields]
            allowed_virtual_fields = ["job_position", "client_name", "client_id"]

            for field in requested_fields:
                if field not in model_fields and field not in allowed_virtual_fields:
                    return Response(api_json_response_format(False, f"Invalid field: {field}", 400, {}), status=200)

            # 🔁 Append computed fields if needed
            for virtual_field in allowed_virtual_fields:
                if virtual_field not in requested_fields:
                    requested_fields.append(virtual_field)

            queryset = InterviewPlanner.objects.all()
            result = []

            for obj in queryset:
                row = {field: getattr(obj, field, None) for field in requested_fields if field in model_fields}

                # 🔍 Enrich with requisition info
                try:
                    requisition = JobRequisition.objects.filter(RequisitionID=obj.requisition_id).first()
                    if requisition:
                        row["job_position"] = requisition.PositionTitle or "Not Provided"
                        row["client_name"] = requisition.company_client_name or "N/A"
                        row["client_id"] = requisition.client_id or "N/A"
                    else:
                        row["job_position"] = "Not Found"
                        row["client_name"] = "Not Found"
                        row["client_id"] = "Not Found"
                except Exception:
                    row["job_position"] = "Error"
                    row["client_name"] = "Error"
                    row["client_id"] = "Error"

                result.append(row)

            return Response(api_json_response_format(True, "Interviewer Bandwidth Dashboard", 200, result), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error in Interviewer Bandwidth Dashboard post method: {str(e)}", 500, {}), status=200)


class ManageRequisitionView(APIView):
    def post(self, request):
        try:
            # Extract expected fields directly from request payload
            requisition_date = request.data.get("requisition_date")
            due_requisition_date = request.data.get("due_requisition_date")
            requisition_template = request.data.get("requisition_template")
            no_of_openings = request.data.get("no_of_openings")
            hiring_plan_id = request.data.get("hiring_plan_id")

            response_data = {
                "hiring_plan_id": hiring_plan_id,
                "requisition_date": requisition_date,
                "due_requisition_date": due_requisition_date,
                "requisition_template": requisition_template,
                "no_of_openings": no_of_openings
            }

            return Response(
                api_json_response_format(True, "Manage requisition payload echoed successfully.", 200, response_data),
                status=200
            )

        except Exception as e:
            return Response(
                api_json_response_format(False, "Error echoing manage requisition payload: " + str(e), 500, {}),
                status=200
            )

def normalize_hiring_plan_payload(raw):
    def flatten_single(data, key):
        val = data.get(key)
        if isinstance(val, list) and val and isinstance(val[0], dict):
            return val[0].get("value", "")
        if isinstance(val, dict):
            return val.get("value", "")
        if isinstance(val, str):
            return val
        return ""

    def flatten_list(data, key):
        val = data.get(key)
        if isinstance(val, list):
            return ", ".join([
                item.get("value", "") if isinstance(item, dict) else str(item)
                for item in val
            ])
        if isinstance(val, str):
            return val
        return ""

    def normalize_bool_string(val):
        return "Yes" if str(val).strip().lower() == "yes" else "No"

    # 🔹 Flatten label-value fields
    flatten_map = {
        "designation": "designation",
        "education_qualification": "education_qualification",
        "shift_timings": "shift_timings",
        "location": "location",
        "working_modal": "working_model",
        "job_type": "job_type",
        "role_type": "role_type",
        "experience_range": "experience_range",
        "compensation_range": "compensation_range"
    }
    for source, target in flatten_map.items():
        raw[target] = flatten_single(raw, source)

    # 🔹 Flatten comma-separated multi-value fields
    for field in ["tech_stacks", "target_companies", "bg_verification_type", "citizen_countries"]:
        raw[field] = flatten_list(raw, field)

    # 🔹 Normalize boolean-like fields to "Yes"/"No"
    for field in [
        "relocation", "has_domain", "visa_required",
        "background_verfication", "citizen_requirement",
        "career_gap", "social_media_link"
    ]:
        raw[field] = normalize_bool_string(raw.get(field))

    # 🔹 Normalize numeric fields
    raw["relocation_amount"] = str(raw.get("relocation_amount", ""))
    raw["travel_opportunities"] = str(raw.get("travel_opportunities", ""))

    # 🔹 Communication language & proficiency (multi-entry)
    cl = raw.get("communication_language")
    if isinstance(cl, list):
        langs = []
        for entry in cl:
            lang = flatten_single(entry, "language")
            prof = flatten_single(entry, "proficiency")
            if lang or prof:
                langs.append(f"{lang}:{prof}")
        raw["communication_language"] = ", ".join(langs)

    # 🔹 Social media links
    sm = raw.get("social_media_data")
    if isinstance(sm, list):
        raw["social_media_data"] = "\n".join([
            f"{item.get('media_type', '')}: {item.get('media_link', '')}"
            for item in sm if isinstance(item, dict)
        ])

    if not raw.get("job_position"):
        job_roles_raw = raw.get("job_role", [])
        job_roles = []

        for item in job_roles_raw:
            val = item.get("value") if isinstance(item, dict) else str(item)
            if val:
                job_roles.extend([r.strip() for r in val.split(",") if r.strip()])

        # Deduplicate while preserving order
        unique_roles = list(dict.fromkeys(job_roles))
        raw["job_position"] = ", ".join(unique_roles)

    # 🔹 Direct mappings
    raw["currency_type"] = raw.get("currency_type", "")
    raw["relocation_currency_type"] = raw.get("relocation_currency_type", "")
    raw["sub_domain_name"] = raw.get("sub_domain_name", "")
    raw["domain_name"] = raw.get("domain_name", "")
    raw["jd_details"] = raw.get("jd_details", "")

    return raw


@api_view(["GET"])
def get_bg_package_dropdown(request):
    packages = BgPackage.objects.all().order_by("name")
    data = [{"label": p.name, "value": p.name} for p in packages]
    return Response(api_json_response_format(True, "Packages retrieved.", 200, data))


CONFIG_MAP = {
    "job_role": "Position Role",
    "tech_stacks": "Tech Stack",
    "experience_range": "Experience",
    "designation": "Designation",
    "target_companies": "Target Companies",
    "location": "Location",
    "working_modal": "Working Model",
    "job_type": "Job Type",
    "role_type": "Role Type",
    "shift_timings": "Shift Timings",
    "education_qualification": "Education Qualification",
    "bg_verification_type": "Background Verification",
    "citizen_countries": "Citizen Country",
    "communication_language": "Communication Language"
}

def ensure_config_value(category_name, category_value):
    if category_name and category_value:
        category_name = category_name.strip()
        category_value = category_value.strip()
        exists = ConfigHiringData.objects.filter(
            category_name=category_name,
            category_values__iexact=category_value
        ).exists()
        if not exists:
            ConfigHiringData.objects.create(
                category_name=category_name,
                category_values=category_value
            )

def extract_values(raw):
    if isinstance(raw, list):
        return [
            item.get("value") or item.get("label") if isinstance(item, dict)
            else item.strip()
            for item in raw if item
        ]
    elif isinstance(raw, dict):
        return [raw.get("value") or raw.get("label")]
    elif isinstance(raw, str):
        return [v.strip() for v in raw.split(",") if v.strip()]
    return []

def process_config_from_payload(payload):
    for key, category in CONFIG_MAP.items():
        raw_value = payload.get(key)

        # Special handling for communication_language
        if key == "communication_language":
            # Case 1: List of dicts with nested language objects
            if isinstance(raw_value, list):
                for entry in raw_value:
                    lang = entry.get("language", {})
                    lang_val = lang.get("value") or lang.get("label")
                    if lang_val:
                        lang_val = lang_val.strip()
                        if lang_val:
                            ensure_config_value(category, lang_val)
                continue

            # Case 2: Colon-separated string like "English:Advanced, Tamil:Intermediate"
            elif isinstance(raw_value, str):
                parts = [item.split(":")[0].strip() for item in raw_value.split(",") if ":" in item]
                for lang in parts:
                    if lang:
                        ensure_config_value(category, lang)
                continue

        # Standard extraction and ingestion
        values = extract_values(raw_value)
        for val in values:
            val = val.strip()
            if val:
                ensure_config_value(category, val)




class HiringPlanOverviewDetails(APIView):
    def get(self, request):
        try:
            raw_fields = request.query_params.get("fields")
            selected_fields = [f.strip() for f in raw_fields.split(",") if f.strip()] if raw_fields else []

            if selected_fields:
                planning_fields_map = {
                    key: val for key, val in DISPLAY_TO_MODEL_FIELD.items()
                    if (isinstance(val, str) and val.startswith("Planning_id__")) or
                       (isinstance(val, list) and all(f.startswith("Planning_id__") for f in val))
                }

                queryset = HiringPlan.objects.select_related("JobRequisition_id", "JobRequisition_id__details")
                result_data = extract_requested_fields(queryset, selected_fields, planning_fields_map)

                return Response(api_json_response_format(
                    True, "Filtered hiring plan data retrieved successfully!", 200, {
                        "hiring_plans": result_data,
                        "selected_fields": selected_fields
                    }), status=status.HTTP_200_OK)

            hiring_plan = HiringPlan.objects.all()
            serializer = HiringPlanSerializer(hiring_plan, many=True)
            return Response(api_json_response_format(
                True, "Hiring plans retrieved successfully.", 200, {
                    "hiring_plans": serializer.data
                }), status=status.HTTP_200_OK)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error retrieving hiring plans. {str(e)}", 500, {}), status=status.HTTP_200_OK)

    def post(self, request):
        last_plan = HiringPlan.objects.order_by("-id").first()
        next_id = f"PL{int(last_plan.hiring_plan_id[2:]) + 1:04d}" if last_plan and last_plan.hiring_plan_id.startswith("PL") else "PL0001"

        raw = request.data.copy()
        raw["hiring_plan_id"] = next_id

        # 🔁 Normalize dropdowns directly in view
        def flatten_dropdowns(payload, fields):
            def flatten(raw):
                values = []
                if isinstance(raw, list):
                    for item in raw:
                        if isinstance(item, dict) and "value" in item:
                            values.append(item["value"].strip())
                        elif isinstance(item, str):
                            values.append(item.strip())
                elif isinstance(raw, str):
                    values = [v.strip() for v in raw.split(",") if v.strip()]
                return ", ".join(values)

            for field in fields:
                if field in payload:
                    payload[field] = flatten(payload[field])

        dropdown_fields = [
            "tech_stacks", "job_role", "designation", "experience_range", "target_companies",
            "compensation_range", "location", "job_type", "role_type", "shift_timings",
            "education_qualification", "bg_verification_type", "citizen_countries",
            "working_modal", "communication_language"
        ]
        flatten_dropdowns(raw, dropdown_fields)

        # 🧠 Normalize communication_language → language_proficiency
        cl_list = request.data.get("communication_language")
        if isinstance(cl_list, list):
            langs, profs = [], []
            for cl in cl_list:
                lang = cl.get("language", {}).get("value")
                prof = cl.get("proficiency", {}).get("value")
                if lang and prof:
                    langs.append(f"{lang}:{prof}")
                    profs.append(prof)
            raw["communication_language"] = ", ".join(langs)
            raw["language_proficiency"] = ", ".join(profs)

        # 🧼 Fix typo: doamin_details → domain_details
        if "doamin_details" in raw:
            raw["domain_details"] = raw.pop("doamin_details")

        # 🧼 Normalize social_media_data → social_media_links
        sm_list = request.data.get("social_media_data")
        if isinstance(sm_list, list):
            links = [f"{sm.get('media_type', '')}: {sm.get('media_link', '')}" for sm in sm_list]
            raw["social_media_links"] = "; ".join(links)

        # 🔁 Normalize payload
        normalized = normalize_hiring_plan_payload(raw)
        # 🧠 Derive job_position from job_role
        raw_roles = normalized.get("job_role", "")
        if isinstance(raw_roles, str):
            roles = [r.strip() for r in raw_roles.split(",") if r.strip()]
        elif isinstance(raw_roles, list):
            roles = [str(r).strip() for r in raw_roles]
        else:
            roles = []

        normalized["job_position"] = ", ".join(dict.fromkeys(roles))


        # 🔐 Client ID logic
        client_name = normalized.get("client_name")
        if client_name:
            cleaned_name = client_name.strip().title()
            existing_client = HiringPlan.objects.filter(client_name=cleaned_name, client_id__isnull=False).first()
            if existing_client:
                normalized["client_id"] = existing_client.client_id
            else:
                last_client = HiringPlan.objects.exclude(client_id__isnull=True).order_by("-id").first()
                next_client_id = f"CL{int(str(last_client.client_id).replace('CL', '')) + 1:04d}" if last_client and str(last_client.client_id).startswith("CL") else "CL0001"
                normalized["client_id"] = next_client_id
            normalized["client_name"] = cleaned_name
        else:
            normalized["client_id"] = None
            normalized["client_name"] = None

        # 🔧 Config hook
        process_config_from_payload(raw)

        # 🚀 Create hiring plan
        serializer = HiringPlanSerializer(data=normalized)
        if serializer.is_valid():
            instance = serializer.save()
            ConfigHiringData.objects.get_or_create(
                category_name="planning_templates",
                category_values=instance.hiring_plan_id
            )
            return Response(api_json_response_format(
                True, "Hiring plan created successfully.", 200, {
                    "hiring_plan_id": instance.hiring_plan_id,
                    "job_position": instance.job_position
                }), status=status.HTTP_200_OK)

        return Response(api_json_response_format(
            False, "Validation error while creating hiring plan.", 400, {
                "errors": serializer.errors
            }), status=status.HTTP_200_OK)


    def put(self, request):
        hiring_plan_id = request.data.get("hiring_plan_id")
        if not hiring_plan_id:
            return Response(api_json_response_format(
                False, "Hiring Plan ID is required in request body.", 400, {}
            ), status=status.HTTP_200_OK)

        try:
            instance = HiringPlan.objects.get(hiring_plan_id=hiring_plan_id)
        except HiringPlan.DoesNotExist:
            return Response(api_json_response_format(
                False, "Hiring plan not found.", 404, {}
            ), status=status.HTTP_200_OK)

        raw = request.data.copy()

        # 🔁 Normalize dropdowns
        def flatten_dropdowns(payload, fields):
            def flatten(raw):
                values = []
                if isinstance(raw, list):
                    for item in raw:
                        if isinstance(item, dict) and "value" in item:
                            values.append(item["value"].strip())
                        elif isinstance(item, str):
                            values.append(item.strip())
                elif isinstance(raw, str):
                    values = [v.strip() for v in raw.split(",") if v.strip()]
                return ", ".join(values)

            for field in fields:
                if field in payload:
                    payload[field] = flatten(payload[field])

        dropdown_fields = [
            "tech_stacks", "job_role", "designation", "experience_range", "target_companies",
            "compensation_range", "location", "job_type", "role_type", "shift_timings",
            "education_qualification", "bg_verification_type", "citizen_countries",
            "working_modal", "communication_language"
        ]
        flatten_dropdowns(raw, dropdown_fields)

        # 🧠 Normalize communication_language → language_proficiency
        cl_list = request.data.get("communication_language")
        if isinstance(cl_list, list):
            langs, profs = [], []
            for cl in cl_list:
                lang = cl.get("language", {}).get("value")
                prof = cl.get("proficiency", {}).get("value")
                if lang and prof:
                    langs.append(f"{lang}:{prof}")
                    profs.append(prof)
            raw["communication_language"] = ", ".join(langs)
            raw["language_proficiency"] = ", ".join(profs)

        # 🧼 Fix typo: doamin_details → domain_details
        if "doamin_details" in raw:
            raw["domain_details"] = raw.pop("doamin_details")

        # 🧼 Normalize social_media_data → social_media_links
        sm_list = request.data.get("social_media_data")
        if isinstance(sm_list, list):
            links = [f"{sm.get('media_type', '')}: {sm.get('media_link', '')}" for sm in sm_list]
            raw["social_media_links"] = "; ".join(links)

        # 🔁 Normalize payload
        normalized = normalize_hiring_plan_payload(raw)

        # 🧠 Derive job_position from job_role
        raw_roles = normalized.get("job_role", "")
        if isinstance(raw_roles, str):
            roles = [r.strip() for r in raw_roles.split(",") if r.strip()]
        elif isinstance(raw_roles, list):
            roles = [str(r).strip() for r in raw_roles]
        else:
            roles = []
        normalized["job_position"] = ", ".join(dict.fromkeys(roles))

        # 🔧 Config hook
        process_config_from_payload(raw)

        # 🚀 Update hiring plan
        serializer = HiringPlanSerializer(instance, data=normalized, partial=True)
        if serializer.is_valid():
            rs = serializer.save()
            return Response(api_json_response_format(
                True, "Hiring plan updated successfully.", 200, {
                    "hiring_plan_id": rs.hiring_plan_id,
                    "job_position": rs.job_position
                }), status=status.HTTP_200_OK)

        return Response(api_json_response_format(
            False, "Validation error while updating hiring plan.", 400, {
                "errors": serializer.errors
            }), status=status.HTTP_200_OK)

    def delete(self, request):
        hiring_plan_id = request.data.get("hiring_plan_id")
        if not hiring_plan_id:
            return Response(api_json_response_format(
                False, "Hiring Plan ID is required in request body.", 400, {}
            ), status=status.HTTP_200_OK)

        obj = get_object_or_404(HiringPlan, hiring_plan_id=hiring_plan_id)
        obj.delete()
        ConfigHiringData.objects.filter(
            category_name="planning_templates",
            category_values=hiring_plan_id
        ).delete()

        return Response(api_json_response_format(
            True, "Hiring plan deleted successfully.", 204, {}
        ), status=status.HTTP_200_OK)

@api_view(['GET'])
def get_all_compensation_ranges(request):
    plans = HiringPlan.objects.exclude(compensation_range__isnull=True).exclude(compensation_range__exact='')

    unique_ranges = set(plan.compensation_range.strip() for plan in plans if plan.compensation_range.strip())

    formatted = [{"label": r, "value": r} for r in sorted(unique_ranges)]

    return Response({
        "success": True,
        "message": "All compensation ranges retrieved successfully",
        "error_code": 200,
        "data": {
            "compensation_range": formatted
        }
    }, status=200)


@api_view(['GET'])
def get_hiring_plans(request):
    try:
        hiring_plans = HiringPlan.objects.all()
        serializer = HiringPlanSerializer(hiring_plans, many=True)

        filtered_data = [
            {'hiring_plan_id': item['hiring_plan_id'], 'job_position': item['job_position']}
            for item in serializer.data
        ]

        return Response(api_json_response_format(
            True,
            "Hiring plan detail fetched successfully!",
            200,
            {"plans": filtered_data}
        ), status=200)

    except Exception as e:
        return Response(api_json_response_format(
            False,
            "Error retrieving hiring plan list."+str(e),500,
            {}
        ), status=200)

@api_view(['POST'])
def get_hiring_plans(request):
    try:
        selected_fields = request.data.get("fields", [])
        if not isinstance(selected_fields, list) or not selected_fields:
            return Response(api_json_response_format(
                False, "Missing or invalid 'fields' parameter.", 400, {}), status=200)

        queryset = HiringPlan.objects.select_related(
            "JobRequisition_id",
            "JobRequisition_id__details"
        ).order_by('-created_at')  # 👈 Latest first


        # Filtered field map (Planning_id only)
        planning_fields_map = {
            "id": "hiring_plan_id",  # ensure this field is manually injected
            **{
                key: val for key, val in DISPLAY_TO_MODEL_FIELD1.items()
                if isinstance(val, str) or (isinstance(val, list) and all(isinstance(f, str) for f in val))
            }
        }


        result_data = extract_requested_fields(queryset, selected_fields, planning_fields_map)
        for item in result_data:
            if "job_position" in item and isinstance(item["job_position"], str):
                item["job_position"] = item["job_position"].replace(" - ", ", ")



        return Response(api_json_response_format(
            True, "Filtered hiring plan data retrieved successfully!", 200, {
                "data": result_data,
                "selected_fields": selected_fields
            }), status=200)

    except Exception as e:
        return Response(api_json_response_format(
            False, "Error retrieving hiring plan data. " + str(e), 500, {}), status=200)


@api_view(['GET'])
def get_all_req_ids(request):
    req_ids = JobRequisition.objects.all().values_list('RequisitionID', flat=True)
    job_positions = JobRequisition.objects.all().values_list('PositionTitle', flat=True)
    combined_data = [
        {"req_ids": req_id, "job_position": position}
        for req_id, position in zip(req_ids, job_positions)
    ]

    return Response(api_json_response_format(
        True,
        "All requisition data fetched successfully!",
        200,
        combined_data
    ), status=200)


def parse_social_media(raw_string):
    try:
        return json.loads(raw_string.replace("'", '"'))  # replace single quotes with double
    except json.JSONDecodeError:
        return []


@api_view(['POST'])
def get_hiring_plan_details(request):
    hiring_plan_id = request.data.get('hiring_plan_id')

    if not hiring_plan_id:
        return Response(api_json_response_format(
            False, "No hiring_plan_id provided.", 400, {}
        ), status=200)

    try:
        plan = HiringPlan.objects.get(hiring_plan_id=hiring_plan_id)

        def safe_split(value):
            return [v.strip() for v in value.split(",") if v.strip()] if isinstance(value, str) else []

        def format_label_value(value):
            return [{"label": v, "value": v} for v in safe_split(value)]


        def format_list_label_value(field):
            return [{"label": v, "value": v} for v in safe_split(field)]

        def parse_social_media(value):
            lines = value.split("\n") if isinstance(value, str) else []
            return [
                {
                    "media_type": line.split(":")[0].strip(),
                    "media_link": line.split(":")[1].strip() if ":" in line else ""
                }
                for line in lines if line.strip()
            ]

        def parse_communication_language(value):
            entries = safe_split(value)
            return [
                {
                    "language": {"label": lang.split(":")[0], "value": lang.split(":")[0]},
                    "proficiency": {"label": lang.split(":")[1], "value": lang.split(":")[1]}
                }
                for lang in entries if ":" in lang
            ]

        structured_data = {
            "job_role": [{"label": role, "value": role} for role in safe_split(plan.job_position)],
            "no_of_openings": plan.no_of_openings,
            "tech_stacks": format_list_label_value(plan.tech_stacks),
            "experience_range": format_label_value(plan.experience_range),
            "designation": format_label_value(plan.designation),
            "currency_type": plan.currency_type or "",
            "target_companies": format_list_label_value(plan.target_companies),
            "compensation_range": format_label_value(plan.compensation_range),
            "location": format_label_value(plan.location),
            "working_modal": format_list_label_value(plan.working_model),
            "job_type": format_label_value(plan.job_type),
            "role_type": format_label_value(plan.role_type),
            "relocation": plan.relocation or "No",
            "relocation_amount": str(plan.relocation_amount) if plan.relocation_amount else "",
            "relocation_currency_type": plan.relocation_currency_type or "",
            "has_domain": plan.domain_yn or "No",
            "doamin_details": plan.domain_details if isinstance(plan.domain_details, list) else [],
            "visa_details": plan.visa_details if isinstance(plan.visa_details, list) else [],
            "shift_timings": format_label_value(plan.shift_timings),
            "education_qualification": format_label_value(plan.education_qualification),
            "travel_opportunities": plan.travel_opportunities or 0,
            "visa_required": plan.visa_requirements or "No",
            "visa_country": plan.visa_country or "",
            "visa_type": plan.visa_type or "",
            "background_verfication": plan.background_verification or "No",
            "bg_verification_type": format_list_label_value(plan.bg_verification_type),
            "communication_language": parse_communication_language(plan.communication_language),
            "citizen_requirement": plan.citizen_requirement or "No",
            "citizen_countries": safe_split(plan.citizen_countries),
            "career_gap": plan.career_gap or "No",
            "social_media_link": "Yes" if plan.social_media_links else "No",
            "social_media_data": plan.social_media_data if isinstance(plan.social_media_data, list) else [],
            "jd_details": plan.jd_details or "",
            "hiring_plan_id": plan.hiring_plan_id,
            "client_name": plan.client_name or ""
        }

        return Response(api_json_response_format(
            True, "Hiring plan detail fetched successfully!", 200, structured_data
        ), status=200)

    except HiringPlan.DoesNotExist:
        return Response(api_json_response_format(
            False, "Hiring plan not found.", "HIRING_PLAN_NOT_FOUND", {}
        ), status=200)

    except Exception as e:
        return Response(api_json_response_format(
            False, f"Error retrieving hiring plan. {str(e)}", 500, {}
        ), status=200)


class HiringPlanDetailView(APIView):
    def post(self, request):
        try:
            hiring_plan_id = request.data.get("Planning_id")
            if not hiring_plan_id:
                return Response(api_json_response_format(
                    False, "Missing hiring_plan_id", 400, {}), status=200)

            try:
                obj = HiringPlan.objects.get(hiring_plan_id=hiring_plan_id)
                serializer = HiringPlanSerializer(obj)
                return Response(api_json_response_format(
                    True, "Hiring plan fetched successfully", 200, serializer.data), status=200)
            except HiringPlan.DoesNotExist:
                return Response(api_json_response_format(
                    False, "Hiring plan not found", 404, {}), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, "Error fetching hiring plan: " + str(e), 500, {}), status=200)


class HiringInterviewRounds(APIView):
    def get(self, request):
        try:
            hiring_rounds = InterviewRounds.objects.all()
            serializer = HiringInterviewRoundsSerializer(hiring_rounds, many=True)
            return Response(api_json_response_format(
                True,
                "Interview rounds fetched successfully.",
                200,
                {"rounds": serializer.data}
            ), status=200)
        except Exception as e:
            return Response(api_json_response_format(
                False,
                "Error fetching interview rounds."+str(e),
                500,
                {"details": str(e)}
            ), status=200)

    def post(self, request):
        try:
            requisition_id = request.data.get('plan_id')
            round_name_list = request.data.get('round_name', [])

            if not requisition_id or not round_name_list:
                return Response(api_json_response_format(
                    False,
                    "Missing requisition_id or round_name list.",
                    400,
                    {}
                ), status=200)

            data_to_insert = [
                {"plan_id": requisition_id, "round_name": name}
                for name in round_name_list
            ]

            serializer = HiringInterviewRoundsSerializer(data=data_to_insert, many=True)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(
                    True,
                    "Interview rounds created successfully.",
                    200,
                    serializer.data
                ), status=200)
            return Response(api_json_response_format(
                False,
                "Validation failed while creating interview rounds."+serializer.errors,
                400,
                {}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                "Unexpected error during round creation."+str(e),
                500,
                {}
            ), status=200)

    def put(self, request):
        requisition_id = request.data.get('id')
        if not requisition_id:
            return Response(api_json_response_format(
                False,
                "Hiring Plan ID is required in request body.",
                400,
                {}
            ), status=200)

        try:
            instance = InterviewRounds.objects.get(id=requisition_id)
        except InterviewRounds.DoesNotExist:
            return Response(api_json_response_format(
                False,
                "Interview round not found.",
                404,
                {}
            ), status=200)

        serializer = HiringInterviewRoundsSerializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(api_json_response_format(
                True,
                "Interview round updated successfully.",
                200,
                serializer.data
            ), status=200)
        return Response(api_json_response_format(
            False,
            "Validation error while updating interview round."+serializer.errors,
            400,
            {}
        ), status=200)
    
class HiringInterviewSkills(APIView):
    def post(self, request):
        try:
            requisition_id = request.data.get('plan_id')
            skill_name_list = request.data.get('skill_name', [])
            skill_value_list = request.data.get('skill_value', [])

            if not requisition_id or not skill_name_list or not skill_value_list:
                return Response(api_json_response_format(
                    False,
                    "Missing requisition_id or skill data.",
                    400,
                    {}
                ), status=200)

            if len(skill_name_list) != len(skill_value_list):
                return Response(api_json_response_format(
                    False,
                    "skill_name and skill_value must have same number of items.",
                    400,
                    {}
                ), status=200)

            data_to_insert = [
                {"plan_id": requisition_id, "skill_name": name, "skill_value": value}
                for name, value in zip(skill_name_list, skill_value_list)
            ]

            serializer = HiringSkillsSerializer(data=data_to_insert, many=True)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(
                    True,
                    "Interview skills created successfully!",
                    200,
                    serializer.data
                ), status=200)

            return Response(api_json_response_format(
                False,
                "Validation error while creating interview skills."+serializer.errors,
                400,
                {}
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                "Unexpected error during interview skills creation."+str(e),
                500,
                {}
            ), status=200)
        

class InterviewDesignScreenView(APIView):
    def get(self, request):
        try:
            interview_design_id = request.query_params.get('interview_design_id')
            if interview_design_id:
                hiring_plan = InterviewDesignParameters.objects.filter(interview_design_id=interview_design_id)
                serializer = InterviewDesignParametersSerializer(hiring_plan, many=True)
                updated_data = []
                for item in serializer.data:
                    item['score_card'] = item.pop('score_card_name')
                    item['weightage'] = item.pop('weightage')
                    updated_data.append(item)
 
            else:
                db_model = InterviewDesignScreen.objects.all()
                serializer = InterviewDesignScreenSerializer(db_model, many=True)  
                updated_data = serializer.data
          
            return Response(api_json_response_format(True, "InterviewDesignScreenView.", 200, updated_data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False,"Error InterviewDesignScreenView get method ()."+str(e),500,{}), status=200)
    # def get(self, request):
    #     try:
    #         db_model = InterviewDesignScreen.objects.all()
    #         serializer = InterviewDesignScreenSerializer(db_model, many=True)
    #         return Response(api_json_response_format(True, "InterviewDesignScreenView.", 200, serializer.data), status=200)
    #     except Exception as e:
    #         return Response(api_json_response_format(False, f"Error InterviewDesignScreenView get method: {str(e)}", 500, {}), status=200)

        
    def post(self, request, *args, **kwargs):
        try:
            design_params = request.data.get("params", [])
            interview_design = request.data.copy()

            req_id = interview_design.get("req_id")
            plan_id = interview_design.get("plan_id")

            if not req_id:
                return Response(api_json_response_format(False, "req_id is required.", 400, {}), status=200)

            interview_design["req_id"] = req_id
            interview_design["hiring_plan_id"] = plan_id or ""

            # 🧠 Auto-fill position_role from JobRequisition
            requisition = JobRequisition.objects.filter(RequisitionID=req_id).first()
            if requisition:
                interview_design["position_role"] = requisition.PositionTitle

            # Remove nested params before main serializer
            interview_design.pop("params", None)

            serializer = InterviewDesignScreenSerializer(data=interview_design)
            if serializer.is_valid():
                instance = serializer.save()
                interview_design_id = instance.interview_design_id

                # 🔁 Normalize and attach nested parameters
                params_data = []
                for obj in design_params:
                    score_card_name = obj.get("score_card_name", "").strip()
                    score_card = score_card_name if score_card_name else "Untitled"

                    min_questions_raw = obj.get("min_questions", 0)
                    min_questions = int(min_questions_raw) if str(min_questions_raw).isdigit() else 0

                    param = {
                        "interview_design_id": interview_design_id,
                        "score_card": score_card,
                        "options": obj.get("options", ""),
                        "guideline": obj.get("guideline", ""),
                        "min_questions": min_questions,
                        "screen_type": obj.get("screen_type", ""),
                        "duration": obj.get("duration", 0),
                        "weightage": obj.get("weightage", 0),
                        "round_no": obj.get("roundno", 0),
                        "mode": obj.get("mode", ""),
                        "feedback": obj.get("feedback", ""),
                        "duration_metric": obj.get("duration_metric", ""),
                        "skills": obj.get("skills", "")
                    }

                    param_serializer = InterviewDesignParametersSerializer(data=param)
                    if param_serializer.is_valid():
                        param_serializer.save()
                        params_data.append(param_serializer.data)
                    else:
                        return Response(api_json_response_format(
                            False, "Could not save Interview Design parameters", 400, param_serializer.errors
                        ), status=200)

                # Pull final values from instance to avoid undefined variables
                response_data = {
                    "plan_id": instance.hiring_plan_id,
                    "req_id": instance.req_id,
                    "interview_design_id": interview_design_id,
                    "tech_stacks": instance.tech_stacks,
                    "screening_type": instance.screening_type,
                    "no_of_interview_round": instance.no_of_interview_round,
                    "final_rating": instance.final_rating,
                    "status": instance.status,
                    "feedback": instance.feedback,
                    "params": params_data
                }

                return Response(api_json_response_format(
                    True, "Interview Design Screen Details Saved Successfully!", 200, response_data
                ), status=200)

            return Response(api_json_response_format(
                False, "Could not save Interview Design", 400, serializer.errors
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Could not save Interview Design: {str(e)}", 500, {}
            ), status=200)


    def put(self, request, *args, **kwargs):
        try:
            interview_design_id = request.data.get("interview_design_id")
            design_params = request.data.get("params", [])
            interview_design = request.data.copy()

            if not interview_design_id:
                return Response(api_json_response_format(
                    False, "interview_design_id is required.", status.HTTP_400_BAD_REQUEST, {}
                ), status=200)

            instance = InterviewDesignScreen.objects.filter(interview_design_id=interview_design_id).first()
            if not instance:
                return Response(api_json_response_format(
                    False, "No matching InterviewDesignScreen found.", 404, {}
                ), status=200)

            req_id = interview_design.get("req_id")
            plan_id = interview_design.get("plan_id")

            interview_design["hiring_plan_id"] = plan_id
            interview_design["req_id"] = req_id

            requisition = JobRequisition.objects.filter(RequisitionID=req_id).first()
            if requisition:
                interview_design["position_role"] = requisition.PositionTitle

            interview_design.pop("plan_id", None)
            interview_design.pop("params", None)

            serializer = InterviewDesignScreenSerializer(instance, data=interview_design, partial=True)
            if serializer.is_valid():
                updated_instance = serializer.save()

                InterviewDesignParameters.objects.filter(interview_design_id=interview_design_id).delete()

                normalized_params = []
                for obj in design_params:
                    param = {
                        "score_card": obj.get("score_card_name", ""),
                        "options": obj.get("options", ""),
                        "guideline": obj.get("guideline", ""),
                        "min_questions": int(obj.get("min_questions") or 0),
                        "screen_type": obj.get("screen_type", ""),
                        "duration": obj.get("duration", ""),
                        "duration_metric": obj.get("duration_metric", ""),
                        "weightage": obj.get("weightage") or obj.get("Weightage", 0),
                        "round_no": obj.get("roundno", 0),
                        "feedback": obj.get("feedback", ""),
                        "skills": obj.get("skills", ""),
                        "mode": obj.get("mode", ""),
                        "interview_design_id": interview_design_id,
                        "hiring_plan_id": updated_instance.hiring_plan_id or ""
                    }
                    normalized_params.append(param)

                serializer_params = InterviewDesignParametersSerializer(data=normalized_params, many=True)
                if serializer_params.is_valid():
                    serializer_params.save()

                    response_data = serializer.data.copy()
                    response_data["req_id"] = req_id

                    return Response(api_json_response_format(
                        True, "Interview Design Screen Updated Successfully!", 200, response_data
                    ))

                return Response(api_json_response_format(
                    False, f"Failed to update parameters: {serializer_params.errors}",
                    status.HTTP_400_BAD_REQUEST, {}
                ))

            return Response(api_json_response_format(
                False, f"Failed to update Interview Design: {serializer.errors}",
                status.HTTP_400_BAD_REQUEST, {}
            ))

        except Exception as e:
            return Response(api_json_response_format(
                False, f"Error updating Interview Design: {str(e)}", 500, {}
            ), status=200)


    def delete(self, request):
        interview_design_id = request.data.get('interview_design_id')
        if not interview_design_id:
            return Response(api_json_response_format(False,"interview_design_id is required in request body.",400,{}), status=200)

        obj = get_object_or_404(InterviewDesignScreen, interview_design_id=interview_design_id)
        obj.delete()
        InterviewDesignParameters.objects.filter(interview_design_id=interview_design_id).delete()
        return Response(api_json_response_format(True,"Interview Desing screen Details are deleted successfully.",200,{}), status=202)  


@api_view(['POST'])
def interview_design_by_id(request):
    try:
        interview_design_id = request.data.get("interview_design_id")
        if not interview_design_id:
            return Response(api_json_response_format(False, "interview_design_id is required in request body.", 400, {}), status=200)

        instance = InterviewDesignScreen.objects.filter(interview_design_id=interview_design_id).first()
        if not instance:
            return Response(api_json_response_format(False, "Design not found.", 404, {}), status=200)

        serializer = InterviewDesignScreenSerializer(instance)
        return Response(api_json_response_format(True, "Interview design retrieved successfully.", 200, serializer.data), status=200)

    except Exception as e:
        return Response(api_json_response_format(False, f"Error retrieving interview design: {str(e)}", 500, {}), status=200)

@api_view(['POST'])
def interview_planner_by_id(request):
    try:
        planner_id = request.data.get("interview_plan_id")
        if not planner_id:
            return Response(api_json_response_format(False, "interview_plan_id is required in request body.", 400, {}), status=200)

        instance = InterviewPlanner.objects.filter(interview_plan_id=planner_id).first()
        if not instance:
            return Response(api_json_response_format(False, "Interview Planner not found.", 404, {}), status=200)

        serializer = InterviewPlannerSerializer(instance)
        return Response(api_json_response_format(True, "Interview Planner retrieved successfully.", 200, serializer.data), status=200)

    except Exception as e:
        return Response(api_json_response_format(False, f"Error retrieving Interview Planner: {str(e)}", 500, {}), status=200)

      
class StateAlertResposibilityView(APIView):
    def get(self, request):
        try:
            hiring_plan_id = request.query_params.get('hiring_plan_id')  # Changed from request.data.get
            if hiring_plan_id:
                queryset = StageAlertResponsibility.objects.filter(hiring_plan_id=hiring_plan_id)
            else:
                queryset = StageAlertResponsibility.objects.all()

            serializer = StageAlertResponsibilitySerializer(queryset, many=True)
            return Response(api_json_response_format(
                True,
                "Stage alert and responsibility settings details retrieved successfully.",
                200,
                {"stage_alert_and_responsibility": serializer.data}
            ), status=200)
        except Exception as e:
            return Response(api_json_response_format(
                False,
                "Error retrieving stage alert and responsibility settings: " + str(e),
                500,
                {}
            ), status=200)

    def post(self, request):
        try:      
            serializer = StageAlertResponsibilitySerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(True,"Stage alert and responsibility settings Details updated Successfully!.",200,serializer.data), status=200)

            return Response(api_json_response_format(
                False,
                "Could not update Stage alert and responsibility settings Details",
                200,
                {
                }
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                "Error during Could not update Stage alert and responsibility settings Details "+str(e),
                500,
                {}
            ), status=200)
        
@action(detail=False, methods=["post"], url_path="get-candidates-by-req")
def get_candidates_by_req_id(self, request):
    req_id = request.data.get("req_id")

    if not req_id:
        return Response(api_json_response_format(False, "req_id is required.", 400, {}), status=200)

    candidates = Candidate.objects.filter(Req_id_fk=req_id)

    if not candidates.exists():
        return Response(api_json_response_format(False, "No candidates found for the given req_id.", 404, {}), status=200)

    serializer = CandidateDetailWithInterviewSerializer(candidates, many=True)
    return Response(api_json_response_format(True, "Candidates retrieved successfully.", 200, serializer.data), status=200)

from collections import defaultdict


class CandidateInterviewStagesView(APIView):
    def get(self, request):
        try:
            stage_groups = defaultdict(list)
            for stage in CandidateInterviewStages.objects.all():
                key = (stage.candidate_id, stage.Req_id_id)
                stage_groups[key].append(stage)

            # Enrich data per group
            enriched_data = []
            for (cid, req_id), stages in stage_groups.items():
                candidate = Candidate.objects.select_related('Req_id_fk').filter(CandidateID=cid).first()
                requisition = candidate.Req_id_fk if candidate else None

                last_completed = None
                current_stage = None

                for s in sorted(stages, key=lambda x: x.interview_date, reverse=True):
                    if s.status == "Completed" and not last_completed:
                        last_completed = s
                    elif not current_stage:
                        current_stage = s

                enriched_data.append({
                    "Req ID":  requisition.RequisitionID if requisition else "",
                    "Candidate ID": candidate.CandidateID if candidate else "",
                    "Candidate First Name": candidate.candidate_first_name if candidate else "",
                    "Candidate Last Name": candidate.candidate_last_name if candidate else "",
                        
                    "Req_ID": requisition.RequisitionID if requisition else "",
                    "Client": requisition.company_client_name if requisition else "",
                    "Candidate_ID": candidate.CandidateID if candidate else "",
                    "Candidate_Name": candidate.candidate_first_name if candidate else "",
                    "Interview_Date": current_stage.interview_date if current_stage else "",
                    "Last_Interview_Stage": last_completed.interview_stage if last_completed else "",
                    "Current_Interview_Stage": current_stage.interview_stage if current_stage else "",
                    "Current_Interview_Status": current_stage.status if current_stage else "",
                    "Interview_Details": {
                        "Mode": current_stage.mode_of_interview if current_stage else "",
                        "Feedback": current_stage.feedback if current_stage else ""
                    }
                })
                if current_stage and current_stage.status == "Completed" and requisition:
                    offer_data = {
                        "client_name": requisition.company_client_name,
                        "client_id": requisition.client_id,
                        "first_name": candidate.candidate_first_name,
                        "last_name": candidate.candidate_last_name,
                        "position_applied": requisition.PositionTitle
                    }

                    OfferNegotiation.objects.update_or_create(
                        candidate=candidate,
                        requisition=requisition,
                        first_name=offer_data["first_name"],
                        last_name=offer_data["last_name"],
                        client_name=offer_data["client_name"],
                        client_id=offer_data["client_id"],
                        defaults={"position_applied": offer_data["position_applied"]}
                    )




            return Response(api_json_response_format(True, "Candidate interview stage data retrieved", 200, enriched_data), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error retrieving interview stage data: {str(e)}", 500, {}), status=200)

    def post(self, request):
        try:             
            serializer = CandidateInterviewStagesSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(True,"Interviwer Calender Details updated Successfully!.",200,serializer.data), status=200)
            return Response(api_json_response_format(False,"Could not update Interviwer Calender Details Details",200,{}), status=200)

        except Exception as e:
            return Response(api_json_response_format(False,"Error during Could not update Interviwer Calender Details Details "+str(e),500,{}), status=200)
     
class InterviewJourneyView(APIView):
    def post(self, request):
        try:
            candidate_id = request.data.get("candidate_id")

            if not candidate_id:
                return Response(api_json_response_format(False, "candidate_id is required", 400, {}), status=200)

            candidate = Candidate.objects.select_related('Req_id_fk').filter(CandidateID=candidate_id).first()
            if not candidate:
                return Response(api_json_response_format(False, "Candidate not found", 404, {}), status=200)

            stages = CandidateInterviewStages.objects.filter(candidate_id=candidate_id).order_by('interview_date')
            journey = []
            screening_review = CandidateReview.objects.filter(CandidateID=candidate).order_by("Created_at").first()
            screening_date = screening_review.Created_at.strftime("%Y-%m-%d") if screening_review else "N/A"
            screening_feedback = candidate.Feedback or "N/A"
            screening_status = candidate.Result or "Pending"

            # Recruiter info from UserDetails via Candidate.created_by or custom mapping
            # Assuming Email is unique and maps to Candidate.created_by or inferred somehow
            requisition = JobRequisition.objects.filter(RequisitionID=candidate.Req_id_fk.RequisitionID).first()
            recruiter_name = requisition.Recruiter if requisition and requisition.Recruiter else "Not Assigned"
            if "," in recruiter_name:
                recruiter_name = recruiter_name.split(",")[0].strip()



            # Construct the screening stage
            screening_stage = {
                "Stage": "Screening",
                "Date": screening_date,
                "Mode": "online",  # Optional, populate if available
                "Interviewer": recruiter_name,
                "Feedback": screening_feedback,
                "Status": screening_status
            }

            journey.append(screening_stage)

            approval_decisions = CandidateApproval.objects.filter(candidate_id=candidate_id).order_by("reviewed_at")

            for approval in approval_decisions:
                approver = approval.approver  # Assuming FK to UserDetails or similar
                approver_name = approver.first_name+" "+approver.last_name if approver and approver.first_name else "N/A"

                approval_stage = {
                    "Stage": f"Approval by {approver_name}",
                    "Date": approval.reviewed_at.strftime("%Y-%m-%d") if approval.reviewed_at else "Pending",
                    "Mode": "System",
                    "Interviewer": approver_name,
                    "Feedback": approval.comment if approval.comment else "",  # Optional
                    "Status": approval.decision or "Pending"
                }
                journey.append(approval_stage)



            for stage in stages:
                # Find corresponding InterviewSchedule
                interview_schedule = InterviewSchedule.objects.filter(
                    candidate__CandidateID=candidate_id,
                    candidate__Req_id_fk_id=stage.Req_id_id,
                    round_name=stage.interview_stage
                ).first()
                interviewer_name = ""
                interviewer = Interviewer.objects.filter(
                    req_id=candidate.Req_id_fk.RequisitionID,
                    interviewer_stage=stage.interview_stage
                ).first()

                if interviewer:
                    interviewer_name = f"{interviewer.first_name or ''} {interviewer.last_name or ''}".strip()
                # Get the latest review if available
                latest_review = interview_schedule.reviews.latest("reviewed_at") if interview_schedule and interview_schedule.reviews.exists() else None

                journey.append({
                    "Stage": stage.interview_stage,
                    "Date": stage.interview_date.strftime("%Y-%m-%d"),
                    "Mode": stage.mode_of_interview,
                    "Interviewer": interviewer_name,  # Optional — fill if you have a related Interviewer field
                    "Feedback": stage.feedback,
                    "Status": stage.result

                })
               
            interview_completed = stages.filter(status__iexact="Completed").exists()

            if interview_completed:
                # Attach Negotiation Stage
                negotiation = OfferNegotiation.objects.filter(candidate__CandidateID=candidate.CandidateID).first()
                if negotiation:
                    negotiation_stage = {
                        "Stage": "Recruiter Negotiation",
                        "Date": negotiation.created_at.strftime("%Y-%m-%d"),
                        "Mode": "Negotiation Portal",
                        "Interviewer": recruiter_name,  # same recruiter from earlier
                        "Feedback": negotiation.comments or "N/A",
                        "Status": negotiation.negotiation_status
                    }
                    journey.append(negotiation_stage)

                    # If negotiation is successful, include approvals
                    if negotiation.negotiation_status == "Successful":
                        approvals = ApprovalStatus.objects.filter(offer_negotiation=negotiation)

                        for approval in approvals:
                            approver = approval.approver
                            approver_name = f"{approver.first_name} {approver.last_name}" if approver else "N/A"

                            approval_stage = {
                                "Stage": f"Approval by {approver_name}",
                                "Date": approval.updated_at.strftime("%Y-%m-%d"),
                                "Mode": "System",
                                "Interviewer": approver_name,
                                "Feedback": "",  # Add remarks if available in model
                                "Status": approval.status
                            }
                            journey.append(approval_stage)

                        # If all approvals passed, show generated offer
                        if approvals.exists() and approvals.filter(status="Pending").count() == 0 and \
                        approvals.filter(status="Rejected").count() == 0:

                            generated_offer = GeneratedOffer.objects.filter(
                                candidate=candidate,
                                requisition=candidate.Req_id_fk
                            ).first()

                            if generated_offer:
                                offer_stage = {
                                    "Stage": "Offer Generated",
                                    "Date": generated_offer.created_at.strftime("%Y-%m-%d"),
                                    "Mode": "System",
                                    "Interviewer": recruiter_name,
                                    "Feedback": f"Job Title: {generated_offer.job_title}, Location: {generated_offer.job_city}, {generated_offer.job_country}",
                                    "Status": generated_offer.negotiation_status
                                }
                                journey.append(offer_stage)



            response = {
                "Candidate_ID": candidate.CandidateID,
                "Candidate_Name": f"{candidate.candidate_first_name} {candidate.candidate_last_name}",
                "Requisition_ID": candidate.Req_id_fk.RequisitionID,
                "Client": candidate.Req_id_fk.company_client_name,
                "Interview_Journey": journey
            }

            return Response(api_json_response_format(True, "Interview journey fetched", 200, response), status=200)

        except Exception as e:
            return Response(api_json_response_format(False, f"Error: {str(e)}", 500, {}), status=200)

@api_view(["POST"])
def get_plan_id_position_role(request):
    try:
        hiring_plan_id = request.data.get('hiring_plan_id')
        if hiring_plan_id:          
            hiring_plan = HiringPlan.objects.filter(hiring_plan_id=hiring_plan_id)
            serializer = HiringPlanSerializer(hiring_plan, many=True)            
            position_list = [item['job_position'] for item in serializer.data]
            return Response(api_json_response_format(True,"Job Position Details",200,position_list))
        else:            
            hiring_plans = list(HiringPlan.objects.values_list('hiring_plan_id', flat=True))
            return Response(api_json_response_format(True,"Candidate Details",200,hiring_plans))
    except Exception as e:
        return Response(api_json_response_format(False,"Error in get_plan_id_position_role  .",500,{}), status=200)
        
        
class InterviewScreenDashboardView(APIView):
    def get(self, request):
        try:
            model_data = InterviewDesignScreen.objects.all()
            design_screen_dashboard_data = []

            for design in model_data:
                score_card_qs = InterviewDesignParameters.objects.filter(
                    interview_design_id=design.interview_design_id
                )
                score_cards = [param.score_card for param in score_card_qs]
                screening_type = score_card_qs.first().screen_type if score_card_qs.exists() else "Not Found"

                # 🔍 Get requisition info
                requisition = JobRequisition.objects.filter(RequisitionID=design.req_id).first()

                design_screen_data = {
                    "interview_design_id":design.interview_design_id,
                    "plan_id": design.hiring_plan_id,
                    "position_role": requisition.PositionTitle if requisition else "Not Found",
                    "tech_stacks": design.tech_stacks,
                    "screening_type": screening_type,
                    "interview_rounds": score_cards,
                    "status": design.status,
                    "requisition_id": design.req_id,
                    "client_name": requisition.company_client_name if requisition else "Not Found",
                    "client_id": requisition.client_id if requisition else "Not Found"
                }

                design_screen_dashboard_data.append(design_screen_data)

            return Response(
                api_json_response_format(True, "Interview Design Screen Dashboard Details", 200, design_screen_dashboard_data),
                status=200
            )

        except Exception as e:
            return Response(
                api_json_response_format(False, f"Error in InterviewScreenDashboardView get method: {str(e)}", 500, {}),
                status=200
            )

@api_view(["POST"])
def filter_candidates_dashboard(request):
    try:
        status_filter = request.data.get('status')       # e.g., 'In-Progress', 'Selected', 'Rejected'
        req_id_filter = request.data.get('req_id')       # e.g., 1234
        interview_stage_filter = request.data.get('stage')  # e.g., 'Screening', 'Final Round'

        candidates = Candidate.objects.all()

        if status_filter and status_filter != "All":
            candidates = candidates.filter(final_status__iexact=status_filter)

        if req_id_filter:
            candidates = candidates.filter(req_id=req_id_filter)

        if interview_stage_filter:
            candidates = candidates.filter(candidate_current_stage__iexact=interview_stage_filter)

        candidate_list = []
        for candidate in candidates:
            candidate_list.append({
                "req_id": candidate.Req_id_fk,
                "candidate_id": candidate.CandidateID,
                "candidate_name": candidate.Name,
                "applied_position": candidate.applied_position,
                "time_in_stage": candidate.time_in_stage,
                "jd_url": candidate.jd_file.url if candidate.jd_file else None,
                "cv_url": candidate.cv_file.url if candidate.cv_file else None,
                "cover_letter_url": candidate.cover_letter.url if candidate.cover_letter else None,
                "current_stage": candidate.candidate_current_stage,
                "next_stage": candidate.candidate_next_stage,
                "overall_score": candidate.Final_rating,
                "final_status": candidate.Result,
                "source": candidate.source,
            })

        return Response(api_json_response_format(True, "Filtered Candidate List", 200, candidate_list), status=200)

    except Exception as e:
        return Response(api_json_response_format(False, f"Error in filter_candidates_dashboard: {str(e)}", 500, {}), status=200)            
    
def admin_configuration(request):
    return render(request, "admin_configuration.html")

class ConfigPositionRoleSearchView(APIView):
    def post(self, request):
        try:
            search_query = request.data.get('query', '').strip()
            if not search_query:
                return Response(
                    api_json_response_format(False, "Search query is required in body.", status.HTTP_400_BAD_REQUEST, {}),
                    status=200
                )

            matched_roles = ConfigPositionRole.objects.filter(position_role__icontains=search_query)
            serializer = ConfigPositionRoleSerializer(matched_roles, many=True)
            return Response(
                api_json_response_format(True, "Search results", 200, serializer.data),
                status=200
            )
        except Exception as e:
            return Response(
                api_json_response_format(False, f"Error during search: {str(e)}", 500, {}),
                status=200
            )
        
class ConfigScreeningTypeSearchView(APIView):
    def post(self, request):
        try:
            search_query = request.data.get('query', '').strip()
            if not search_query:
                return Response(
                    api_json_response_format(False, "Search query is required in body.", status.HTTP_400_BAD_REQUEST, {}),
                    status=200
                )

            matched_types = ConfigScreeningType.objects.filter(screening_type_name__icontains=search_query)
            serializer = ConfigScreeningTypeSerializer(matched_types, many=True)
            return Response(
                api_json_response_format(True, "Search results", 200, serializer.data),
                status=200
            )
        except Exception as e:
            return Response(
                api_json_response_format(False, f"Error during search: {str(e)}", 500, {}),
                status=200
            )
        
class ConfigScoreCardSearchView(APIView):
    def post(self, request):
        try:
            search_query = request.data.get('query', '').strip()
            if not search_query:
                return Response(
                    api_json_response_format(False, "Search query is required in body.", status.HTTP_400_BAD_REQUEST, {}),
                    status=200
                )

            matched_scorecards = ConfigScoreCard.objects.filter(score_card_name__icontains=search_query)
            serializer = ConfigScoreCardSerializer(matched_scorecards, many=True)
            return Response(
                api_json_response_format(True, "Search results", 200, serializer.data),
                status=200
            )
        except Exception as e:
            return Response(
                api_json_response_format(False, f"Error during search: {str(e)}", 500, {}),
                status=200
            )
       
class ConfigPositionRoleView(APIView):
    def get(self, request):
        try:           

            db_model = ConfigPositionRole.objects.all()
            serializer = ConfigPositionRoleSerializer(db_model, many=True)            
            return Response(api_json_response_format(True,"Job Position Role details",200,serializer.data), status=200)
        except Exception as e:
            return Response(api_json_response_format(False,"Error in Job Position Role details get method  "+str(e),500,{}), status=200)
            
    def post(self, request):
        try:
            table_value = (request.data.get('table_value') )
            json_data = {"position_role" : table_value}
            serializer = ConfigPositionRoleSerializer(data=json_data)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(True,"Job Position Role details updated Successfully!.",200,serializer.data), status=200)
            return Response(api_json_response_format(False,"Could not update Job Position Role details",200,{}), status=200)

        except Exception as e:
            return Response(api_json_response_format(False,"Error during Could not update Job Position Role details post method "+str(e),500,{}), status=200)
        
    def put(self, request):
        ref_id = request.data.get('id')        
        if not ref_id:
            return Response(api_json_response_format(False,"ref_id is required.",status.HTTP_400_BAD_REQUEST,{}), status=200)       
        try:
            instance = ConfigPositionRole.objects.get(id=ref_id)
            serializer = ConfigPositionRoleSerializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(True,"Job Position Role Details updated successfully : ",status.HTTP_200_OK,serializer.data), status=200)
            return Response(api_json_response_format(False,"Could not update Job Position Role error : "+str(serializer.errors),500,{}), status=200)       
                
        except Exception as error:
            return Response(api_json_response_format(False,"Could not update Job Position Role error in put method : "+str(error),500,{}), status=200)            
        

    def delete(self, request):
        try:
            ref_id = request.data.get('id') 
            if not ref_id:
                return Response(api_json_response_format(False,"id is required.",status.HTTP_400_BAD_REQUEST,{}), status=200)
            obj = ConfigPositionRole.objects.get(id=ref_id)
            obj.delete()
            return Response(api_json_response_format(True,"Job Position Role deleted successfully.",200,{}), status=200)             
        except Exception as error:
            return Response(api_json_response_format(False,"Could not delete Job Position Role ."+str(error),500,{}), status=200) 
        
class ConfigScreeningTypeView(APIView):
    def get(self, request):
        try:           

            db_model = ConfigScreeningType.objects.all()
            serializer = ConfigScreeningTypeSerializer(db_model, many=True)            
            return Response(api_json_response_format(True,"Job Position Role details",200,serializer.data), status=200)
        except Exception as e:
            return Response(api_json_response_format(False,"Error in Job Position Role details get method  "+str(e),500,{}), status=200)
            
    def post(self, request):
        try:
            table_value = (request.data.get('table_value') )
            json_data = {"screening_type_name" : table_value}
            serializer = ConfigScreeningTypeSerializer(data=json_data)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(True,"Job Position Role details updated Successfully!.",200,serializer.data), status=200)
            return Response(api_json_response_format(False,"Could not update Job Position Role details",200,{}), status=200)

        except Exception as e:
            return Response(api_json_response_format(False,"Error during Could not update Job Position Role details post method "+str(e),500,{}), status=200)
        
    def put(self, request):
        ref_id = request.data.get('id')        
        if not ref_id:
            return Response(api_json_response_format(False,"ref_id is required.",status.HTTP_400_BAD_REQUEST,{}), status=200)       
        try:
            instance = ConfigScreeningType.objects.get(id=ref_id)
            serializer = ConfigScreeningTypeSerializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(True,"Job Position Role Details updated successfully : ",status.HTTP_200_OK,serializer.data), status=200)
            return Response(api_json_response_format(False,"Could not update Job Position Role error : "+str(serializer.errors),500,{}), status=200)       
                
        except Exception as error:
            return Response(api_json_response_format(False,"Could not update Job Position Role error in put method : "+str(error),500,{}), status=200)            
        

    def delete(self, request):
        try:
            ref_id = request.data.get('id') 
            if not ref_id:
                return Response(api_json_response_format(False,"id is required.",status.HTTP_400_BAD_REQUEST,{}), status=200)
            obj = ConfigScreeningType.objects.get(id=ref_id)
            obj.delete()
            return Response(api_json_response_format(True,"Job Position Role deleted successfully.",200,{}), status=200)             
        except Exception as error:
            return Response(api_json_response_format(False,"Could not delete Job Position Role ."+str(error),500,{}), status=200) 
        
class ConfigScoreCardView(APIView):
    def get(self, request):
        try:           
            # screening_type_name_list = list(ConfigScoreCard.objects.values_list('score_card_name', flat=True))
            db_model = ConfigScoreCard.objects.all()
            serializer = ConfigScoreCardSerializer(db_model, many=True)            
            return Response(api_json_response_format(True,"Job Position Role details",200,serializer.data), status=200)
        except Exception as e:
            return Response(api_json_response_format(False,"Error in Job Position Role details get method  "+str(e),500,{}), status=200)
            
    def post(self, request):
        try:
            table_value = (request.data.get('table_value') )
            json_data = {"score_card_name" : table_value}
            serializer = ConfigScoreCardSerializer(data=json_data)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(True,"Job Position Role details updated Successfully!.",200,serializer.data), status=200)
            return Response(api_json_response_format(False,"Could not update Job Position Role details",200,{}), status=200)

        except Exception as e:
            return Response(api_json_response_format(False,"Error during Could not update Job Position Role details post method "+str(e),500,{}), status=200)
        
    def put(self, request):
        ref_id = request.data.get('id')        
        if not ref_id:
            return Response(api_json_response_format(False,"ref_id is required.",status.HTTP_400_BAD_REQUEST,{}), status=200)       
        try:
            instance = ConfigScoreCard.objects.get(id=ref_id)
            serializer = ConfigScoreCardSerializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(True,"Job Position Role Details updated successfully : ",status.HTTP_200_OK,serializer.data), status=200)
            return Response(api_json_response_format(False,"Could not update Job Position Role error : "+str(serializer.errors),500,{}), status=200)       
                
        except Exception as error:
            return Response(api_json_response_format(False,"Could not update Job Position Role error in put method : "+str(error),500,{}), status=200)            
        

    def delete(self, request):
        try:
            ref_id = request.data.get('id') 
            if not ref_id:
                return Response(api_json_response_format(False,"id is required.",status.HTTP_400_BAD_REQUEST,{}), status=200)
            obj = ConfigScoreCard.objects.get(id=ref_id)
            obj.delete()
            return Response(api_json_response_format(True,"Job Position Role deleted successfully.",200,{}), status=200)             
        except Exception as error:
            return Response(api_json_response_format(False,"Could not delete Job Position Role ."+str(error),500,{}), status=200) 

class AdminConfigurationView(APIView):
    def get(self, request):
        try:        
            category_name = request.query_params.get('category_name')
            if not category_name:
                categories = ConfigHiringData.objects.values_list('category_name', flat=True).distinct()
                data = {}
                for category in categories:
                    values = ConfigHiringData.objects.filter(category_name=category).values_list('category_values', flat=True)
                    data[category] = list(values)         
                return Response(api_json_response_format(True,"Admin Configuration Details details",200,data), status=200)
            else:
                print(category_name)
                queryset = ConfigHiringData.objects.filter(category_name=category_name)
                serializer = ConfigHiringDataSerializer(queryset, many=True)
                return Response(api_json_response_format(True,"Admin Configuration Details details",200,serializer.data), status=200)                

        except Exception as e:
            return Response(api_json_response_format(False,"Error in Admin Configuration Details details get method  "+str(e),500,{}), status=status.HTTP_200_OK)
            
    def post(self, request):
        try:            
            serializer = ConfigHiringDataSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(True,"Admin Configuration Details updated Successfully!.",201,serializer.data), status=200)
            return Response(api_json_response_format(False,"Could not update Admin Configuration details",201,{}), status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(api_json_response_format(False,"Error during Could not update Admin Configuration details post method "+str(e),500,{}), status=status.HTTP_200_OK)
        
    def put(self, request):
        ref_id = request.data.get('id')        
        if not ref_id:
            return Response(api_json_response_format(False,"ref_id is required.",status.HTTP_400_BAD_REQUEST,{}), status=200)       
        try:
            instance = ConfigHiringData.objects.get(id=ref_id)
            serializer = ConfigHiringDataSerializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(api_json_response_format(True,"Admin Configuration Details Details updated successfully : ",status.HTTP_200_OK,serializer.data), status=200)
            return Response(api_json_response_format(False,"Could not update Admin Configuration Details error : "+str(serializer.errors),500,{}), status=200)       
                
        except Exception as error:
            return Response(api_json_response_format(False,"Could not update Admin Configuration Details error in put method : "+str(error),500,{}), status=200)            
        

    def delete(self, request):
        try:
            ref_id = request.data.get('id') 
            if not ref_id:
                return Response(api_json_response_format(False,"id is required.",status.HTTP_400_BAD_REQUEST,{}), status=200)
            obj = ConfigHiringData.objects.get(id=ref_id)
            obj.delete()
            return Response(api_json_response_format(True,"Admin Configuration Details  deleted successfully.",200,{}), status=200)             
        except Exception as error:
            return Response(api_json_response_format(False,"Could not delete Admin Configuration Details ."+str(error),500,{}), status=204) 

class MappedAdminConfigView(APIView):
    def get(self, request):
        try:
            categories = ConfigHiringData.objects.values_list('category_name', flat=True).distinct()
            mapped_data = {}

            for category in categories:
                values = ConfigHiringData.objects.filter(category_name=category).values_list('category_values', flat=True)
                mapped_values = [{"label": val, "value": val} for val in values if val]
                mapped_data[category] = mapped_values

            return Response(api_json_response_format(
                True,
                "Mapped admin configuration data retrieved successfully.",
                200,
                mapped_data
            ), status=200)

        except Exception as e:
            return Response(api_json_response_format(
                False,
                f"Error retrieving mapped admin configurations. {str(e)}",
                500,
                {}
            ), status=200)



@api_view(['GET'])
def get_report_dropdowns(request):
    try:
        # 🔹 Clients
        clients = JobRequisition.objects.values_list('company_client_name', flat=True).distinct()
        client_dropdown = [{"label": c, "value": c} for c in clients if c]

        # 🔹 Recruiters
        recruiters = JobRequisition.objects.values_list('Recruiter', flat=True).distinct()
        recruiter_dropdown = [{"label": r, "value": r} for r in recruiters if r and r != "Not Assigned"]

        # 🔹 Positions
        positions = ConfigHiringData.objects.filter(category_name="Position Role").values_list('category_values', flat=True).distinct()
        position_dropdown = [{"label": p, "value": p} for p in positions if p]

        # 🔹 Locations
        locations = ConfigHiringData.objects.filter(category_name="Location").values_list('category_values', flat=True).distinct()
        location_dropdown = [{"label": l, "value": l} for l in locations if l]

        dropdowns = {
            "client_name": client_dropdown,
            "recruiter_name": recruiter_dropdown,
            "position_offered": position_dropdown,
            "location": location_dropdown
        }

        return Response(api_json_response_format(True, "Dropdown values fetched", 200, dropdowns), status=200)

    except Exception as e:
        return Response(api_json_response_format(False, f"Error fetching dropdowns: {e}", 500, {}), status=200)


@api_view(['GET'])
def design_screen_list_data(request):
    try:
        position_role_list = list(ConfigPositionRole.objects.values_list('position_role', flat=True))
        hiring_plan_id_list = list(HiringPlan.objects.values_list('hiring_plan_id', flat=True))
        screening_type_name_list = list(ConfigScreeningType.objects.values_list('screening_type_name', flat=True))

        # 🎯 Filter requisitions present in InterviewDesignScreen and not incomplete
        # valid_req_ids = InterviewDesignScreen.objects.values_list('req_id', flat=True)
        requisition_id_list = list(
            JobRequisition.objects
            .exclude(Status="Incomplete form")
            # .filter(RequisitionID__in=valid_req_ids)
            .values_list('RequisitionID', flat=True)
        )

        result_data = {
            "position_role": position_role_list,
            "plan_id": hiring_plan_id_list,
            "screening_type": screening_type_name_list,
            "requisition_id": requisition_id_list
        }

        return Response(api_json_response_format(True, "Design Screen Dropdown details", 200, result_data), status=200)

    except Exception as e:
        return Response(api_json_response_format(False, f"Error: {str(e)}", 500, {}), status=200)

@api_view(['GET'])
def design_screen_list_data_interviewer(request):
    try:
        position_role_list = list(ConfigPositionRole.objects.values_list('position_role', flat=True))
        hiring_plan_id_list = list(HiringPlan.objects.values_list('hiring_plan_id', flat=True))
        screening_type_name_list = list(ConfigScreeningType.objects.values_list('screening_type_name', flat=True))

        # 🎯 Filter requisitions present in InterviewDesignScreen and not incomplete
        valid_req_ids = InterviewDesignScreen.objects.values_list('req_id', flat=True)
        requisition_id_list = list(
            JobRequisition.objects
            .exclude(Status="Incomplete form")
            .filter(RequisitionID__in=valid_req_ids)
            .values_list('RequisitionID', flat=True)
        )

        result_data = {
            "position_role": position_role_list,
            "plan_id": hiring_plan_id_list,
            "screening_type": screening_type_name_list,
            "requisition_id": requisition_id_list
        }

        return Response(api_json_response_format(True, "Design Screen Dropdown details", 200, result_data), status=200)

    except Exception as e:
        return Response(api_json_response_format(False, f"Error: {str(e)}", 500, {}), status=200)

@api_view(['POST'])
def client_lookup_from_plan(request):
    try:
        req_id = request.data.get("req_id")
        plan_id = request.data.get("plan_id")

        if not req_id:
            return Response(api_json_response_format(
                False, "'req_id' is required.", 400, {}
            ), status=200)

        filter_kwargs = {"RequisitionID": req_id}
        if plan_id:
            filter_kwargs["Planning_id__hiring_plan_id"] = plan_id

        requisition = JobRequisition.objects.filter(**filter_kwargs).first()

        if not requisition:
            return Response(api_json_response_format(
                False, "No matching requisition found.", 404, {}
            ), status=200)

        data = {
            "client_name": requisition.company_client_name or "Not Provided",
            "client_id": requisition.client_id or "Not Provided"
        }

        return Response(api_json_response_format(
            True, "Client info retrieved.", 200, data
        ), status=200)

    except Exception as e:
        return Response(api_json_response_format(
            False, f"Error retrieving client info. {str(e)}", 500, {}
        ), status=200)

@api_view(['POST'])
def job_metadata_lookup(request):
    try:
        req_id = request.data.get("req_id")
        plan_id = request.data.get("plan_id")

        if not req_id:
            return Response(api_json_response_format(
                False, "'req_id' is required.", 400, {}
            ), status=200)

        # 🔍 Find the requisition
        filter_kwargs = {"RequisitionID": req_id}
        if plan_id:
            filter_kwargs["Planning_id__hiring_plan_id"] = plan_id

        requisition = JobRequisition.objects.filter(**filter_kwargs).first()

        if not requisition:
            return Response(api_json_response_format(
                False, "No matching requisition found.", 404, {}
            ), status=200)

        # ✅ Extract hiring_plan_id from requisition (via Planning_id)
        # hiring_plan_id = requisition.Planning_id.hiring_plan_id if requisition.Planning_id else None
        interview_design_id = InterviewDesignScreen.objects.get(req_id=req_id).interview_design_id
        # 🔍 Query InterviewDesignParameters for metadata
        design_params = InterviewDesignParameters.objects.filter(interview_design_id=interview_design_id).order_by("interview_desing_params_id")

        # 💡 Map score_card = stage, screen_type = mode
        interview_design = [
            {
                "interviewer_stage": param.score_card,
                "interview_mode": param.screen_type
            }
            for param in design_params
        ]

        data = {
            "job_position": requisition.PositionTitle or "Not Provided",
            "interview_design": interview_design
        }

        return Response(api_json_response_format(
            True, "Job metadata retrieved.", 200, data
        ), status=200)

    except Exception as e:
        return Response(api_json_response_format(
            False, f"Error retrieving job metadata: {str(e)}", 500, {}
        ), status=200)


def parse_nested_form_data(flat_data):
    nested = {}

    for key, value in flat_data.items():
        if "[" in key and "]" in key:
            parts = key.replace("]", "").split("[")
            current = nested

            for i, part in enumerate(parts):
                is_last = i == len(parts) - 1

                # If next part is an index
                if not is_last:
                    next_part = parts[i + 1]
                    is_index = next_part.isdigit()

                    if part not in current:
                        current[part] = [] if is_index else {}

                    current = current[part]

                    # If it's a list, ensure index exists
                    if isinstance(current, list):
                        index = int(next_part)
                        while len(current) <= index:
                            current.append({})
                        current = current[index]
                else:
                    current[part] = value
        else:
            nested[key] = value

    return nested



@api_view(['POST'])
def submit_pre_onboarding_form(request):
    flat_data = request.data
    data = parse_nested_form_data(flat_data)

    try:
        token = data.get("token")
        if not token:
            return Response(api_json_response_format(False, "Token is required.", 400, {}))

        invite = get_object_or_404(CandidateFormInvite, token=token)

        if invite.expires_at < timezone.now():
            return Response(api_json_response_format(False, "Token has expired.", 400, {}))

        candidate = invite.candidate
        candidate_id = candidate.CandidateID

        candidate_info = data.get("candidateInfo", {})
        personal_details = data.get("personalDetails", {})
        reference_check = data.get("referenceCheck", [])
        banking_details = data.get("bankingDetails", {})
        financial_documents = data.get("financialDocuments", {})
        nominee_details = data.get("nomineeDetails", [])
        insurance_details = data.get("insuranceDetails", [])
        uploaded_documents = data.get("uploadedDocuments", {})

        # Candidate Profile
        candidate_obj, _ = CandidateProfile.objects.update_or_create(
            candidate_id=candidate_id,
            defaults={
                "first_name": candidate_info.get("firstName", candidate.candidate_first_name),
                "last_name": candidate_info.get("lastName", candidate.candidate_last_name),
                "date_of_joining": candidate_info.get("dateOfJoining") or None
            }
        )

        def normalize_personal_details(data):
            return {
                "dob": data.get("dob"),
                "marital_status": data.get("maritalStatus"),
                "gender": data.get("gender"),
                "blood_group": data.get("bloodGroup"),
                "permanent_address": data.get("permanentAddress"),
                "present_address": data.get("presentAddress"),
                "emergency_contact_name": data.get("emergencyPOCName"),
                "emergency_contact_number": data.get("emergencyContactNumber"),
                "photograph": data.get("photographs")
            }

        PersonalDetails.objects.update_or_create(
            candidate=candidate_obj,
            defaults=normalize_personal_details(personal_details)
        )

        flat_references = [list(ref.values())[0] for ref in reference_check]

        def normalize_reference(ref):
            return {
                "first_name": ref.get("firstName"),
                "last_name": ref.get("lastName"),
                "designation": ref.get("designation"),
                "reporting_manager_name": ref.get("reportingManagerName"),
                "official_email": ref.get("officialEmailId"),
                "phone_number": ref.get("phoneNumber")
            }

        ReferenceCheck.objects.bulk_create([
            ReferenceCheck(candidate=candidate_obj, **normalize_reference(ref))
            for ref in flat_references if ref.get("firstName") or ref.get("lastName")
        ])

        def normalize_banking_details(data):
            return {
                "bank_name": data.get("bankName"),
                "account_number": data.get("accountNumber"),
                "ifsc_code": data.get("ifscCode"),
                "branch_address": data.get("branchAddress"),
                "bank_statement": data.get("bankStatement"),
                "cancel_cheque": data.get("cancelCheque")
            }

        BankingDetails.objects.update_or_create(
            candidate=candidate_obj,
            defaults=normalize_banking_details(banking_details)
        )

        def normalize_financial_documents(data):
            return {
                "pf_number": data.get("pfNumber"),
                "uan_number": data.get("uanNumber"),
                "pran_number": data.get("pranNumber"),
                "form_16": data.get("form16"),
                "salary_slips": data.get("salarySlips")
            }

        FinancialDocuments.objects.update_or_create(
            candidate=candidate_obj,
            defaults=normalize_financial_documents(financial_documents)
        )

        flat_nominees = [list(nom.values())[0] for nom in nominee_details]

        def normalize_nominee(nom):
            return {
                "first_name": nom.get("firstName"),
                "last_name": nom.get("lastName"),
                "share_percentage": int(nom.get("share", 0))
            }

        Nominee.objects.bulk_create([
            Nominee(candidate=candidate_obj, **normalize_nominee(nom))
            for nom in flat_nominees if nom.get("firstName") or nom.get("lastName")
        ])

        flat_insurance = [list(ins.values())[0] for ins in insurance_details]

        def normalize_insurance(ins):
            return {
                "first_name": ins.get("firstName"),
                "last_name": ins.get("lastName"),
                "dob": ins.get("dob")
            }

        InsuranceDetail.objects.bulk_create([
            InsuranceDetail(candidate=candidate_obj, **normalize_insurance(ins))
            for ins in flat_insurance if ins.get("firstName") or ins.get("lastName")
        ])

        def normalize_category(raw_category):
            mapping = {
                "education": "Education",
                "employment": "Employment",
                "mandatory": "Mandatory"
            }
            return mapping.get(raw_category.lower(), "Mandatory")

        DocumentItem.objects.filter(candidate=candidate_obj).delete()

        for raw_category, items in uploaded_documents.items():
            category = normalize_category(raw_category)
            DocumentItem.objects.bulk_create([
                DocumentItem(
                    candidate=candidate_obj,
                    category=category,
                    type=doc_type,
                    uploaded_file=file,
                    document_name=f"{doc_type} Document",
                    document_status="Submitted",
                    institution_name="N/A"
                )
                for doc_type, file in items.items() if file
            ])

        invite.completed = True
        invite.save()

        return Response(api_json_response_format(
            True,
            "Pre-onboarding form submitted successfully.",
            200,
            {"candidate_id": candidate_id}
        ))

    except Exception as e:
        return Response(api_json_response_format(
            False,
            f"Error processing form: {str(e)}",
            500,
            {}
        ))


@api_view(['GET'])
def get_all_pre_onboarding_forms(request):
    candidates = CandidateProfile.objects.all()
    all_data = []

    for candidate_obj in candidates:
        entry = {
            "candidate_id": candidate_obj.candidate_id,
            "first_name": candidate_obj.first_name,
            "last_name": candidate_obj.last_name,
            "date_of_joining": candidate_obj.date_of_joining,
            # "personal_details": PersonalDetails.objects.filter(candidate=candidate_obj).values().first(),
            # "references": list(ReferenceCheck.objects.filter(candidate=candidate_obj).values()),
            # "banking_details": BankingDetails.objects.filter(candidate=candidate_obj).values().first(),
            # "financial_documents": FinancialDocuments.objects.filter(candidate=candidate_obj).values().first(),
            # "nominee_details": list(Nominee.objects.filter(candidate=candidate_obj).values()),
            # "insurance_details": list(InsuranceDetail.objects.filter(candidate=candidate_obj).values()),
            # "uploaded_documents": {
            #     category: list(DocumentItem.objects.filter(candidate=candidate_obj, category=category).values())
            #     for category, _ in DocumentItem.CATEGORY_CHOICES
            # }
        }
        all_data.append(entry)

    return Response(api_json_response_format(True, "All pre-onboarding forms retrieved.", 200, all_data))